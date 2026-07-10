"""
DCF Excel 样式辅助函数
=====================

集中管理 DCF 估值工作簿中使用的单元格样式、图例、列宽和注释。
被 dcf_excel_sheets.py 中的各 create_*_sheet() 函数调用。
"""

from typing import Dict

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def apply_title_style(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_header_style(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4F81BD")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D9E2F3")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def apply_input_style(cell) -> None:
    thin = Side(style="thin", color="C9A227")
    cell.fill = PatternFill("solid", fgColor="FFE699")
    cell.font = Font(color="7F6000", bold=True)
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_formula_style(cell) -> None:
    thin = Side(style="thin", color="A9D18E")
    cell.fill = PatternFill("solid", fgColor="E2F0D9")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def apply_output_style(cell) -> None:
    thin = Side(style="thin", color="9EADCC")
    cell.fill = PatternFill("solid", fgColor="DDEBF7")
    cell.font = Font(bold=True, color="1F4E78")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def add_legend(ws, start_cell: str) -> None:
    row = int("".join(ch for ch in start_cell if ch.isdigit()))
    col = ord("".join(ch for ch in start_cell if ch.isalpha()).upper()) - ord("A") + 1
    ws.cell(row=row, column=col, value="颜色说明")
    apply_header_style(ws.cell(row=row, column=col))
    ws.cell(row=row + 1, column=col, value="可编辑输入")
    apply_input_style(ws.cell(row=row + 1, column=col))
    ws.cell(row=row + 2, column=col, value="公式/联动")
    apply_formula_style(ws.cell(row=row + 2, column=col))
    ws.cell(row=row + 3, column=col, value="关键输出")
    apply_output_style(ws.cell(row=row + 3, column=col))


def set_col_widths(ws, widths: Dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_note(ws, cell_ref: str, text: str) -> None:
    ws[cell_ref] = text
    ws[cell_ref].font = Font(italic=True, color="666666")
