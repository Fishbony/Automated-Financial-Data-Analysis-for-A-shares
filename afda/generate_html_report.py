"""
Step 9/9 - offline HTML dashboard and interactive DCF valuation tool.

The generated HTML has independent tabs for:
1. Financial overview: historical financial metrics and charts.
2. Balance sheet, income statement, and cash flow statement views.
3. Valuation model: editable assumptions, full DCF bridge, forecast table, and
   instantly recalculated intrinsic share price.
"""

from __future__ import annotations

import argparse
import csv
import html
import json
import math
import os
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, Optional

from openpyxl import load_workbook


OUTPUT_FILE_NAME = "financial_dcf_dashboard.html"
LOCAL_ECHARTS_SOURCE = Path(__file__).resolve().parents[1] / "assets" / "echarts.min.js"
STATEMENT_SOURCES = {
    "balance_sheet": ("资产负债表", "balance_sheet", "2_standardized_bs_wide.csv"),
    "income_statement": ("利润表", "income_statement", "2_standardized_pl_wide.csv"),
    "cash_flow": ("现金流量表", "cash_flow", "2_standardized_cf_wide.csv"),
}
BALANCE_SHEET_EMPHASIS_ROWS = {
    "Total Current Assets",
    "Total Non-current Assets",
    "Total Assets",
    "Total Current Liabilities",
    "Total Non-current Liabilities",
    "Total Liabilities",
    "Total Equity",
}
STATEMENT_CHARTS = {
    "balance_sheet": [
        {
            "id": "balanceSheetScaleChart",
            "title": "资产、负债与资产负债率",
            "items": ["Total Assets", "Total Liabilities"],
            "types": {"Total Assets": "bar", "Total Liabilities": "bar"},
            "barMode": "grouped",
            "derived": [
                {
                    "name": "资产负债率",
                    "type": "debt_to_assets",
                    "numerator": "Total Liabilities",
                    "denominator": "Total Assets",
                    "chartType": "line",
                    "yAxisIndex": 1,
                }
            ],
            "dualAxis": True,
        },
        {
            "id": "balanceSheetStructureChart",
            "title": "流动与非流动资产结构",
            "items": ["Total Current Assets", "Total Non-current Assets"],
            "types": {"Total Current Assets": "bar", "Total Non-current Assets": "bar"},
            "stack": "assets",
        },
        {
            "id": "currentAssetsStructureChart",
            "title": "流动资产结构",
            "items": [
                "Cash & Short-term Financial Assets",
                "Core Operating Current Assets",
                "Non-operating Misc. Current Assets",
            ],
            "types": {
                "Cash & Short-term Financial Assets": "bar",
                "Core Operating Current Assets": "bar",
                "Non-operating Misc. Current Assets": "bar",
            },
            "stack": "current_assets",
        },
        {
            "id": "netDebtRatioDebtGrowthChart",
            "title": "净负债率与有息负债增速",
            "customType": "net_debt_ratio_debt_growth",
            "dualAxis": True,
        },
    ],
    "income_statement": [
        {
            "id": "incomeProfitChart",
            "title": "收入、营业利润与归母净利润",
            "items": ["Revenue", "Operating Profit", "Parent Net Profit"],
            "types": {"Revenue": "bar", "Operating Profit": "line", "Parent Net Profit": "line"},
            "yAxisIndex": {"Revenue": 0, "Operating Profit": 1, "Parent Net Profit": 1},
            "dualAxis": True,
            "rightAxisName": "利润",
        },
        {
            "id": "incomeExpenseChart",
            "title": "成本与期间费用",
            "items": ["COGS", "Selling Expense", "Admin Expense", "R&D Expense", "Financial Expense"],
            "types": {
                "COGS": "bar",
                "Selling Expense": "bar",
                "Admin Expense": "bar",
                "R&D Expense": "bar",
                "Financial Expense": "bar",
            },
            "stack": "expenses",
        },
    ],
    "cash_flow": [
        {
            "id": "cashFlowStatementChart",
            "title": "经营、投资、筹资现金流",
            "items": ["Operating Cash Flow", "Investing Cash Flow", "Financing Cash Flow", "Net Change in Cash"],
            "types": {
                "Operating Cash Flow": "bar",
                "Investing Cash Flow": "bar",
                "Financing Cash Flow": "bar",
                "Net Change in Cash": "line",
            },
        },
        {
            "id": "cashPositionChart",
            "title": "期初与期末现金余额",
            "items": ["Beginning Cash", "Ending Cash"],
            "types": {"Beginning Cash": "line", "Ending Cash": "line"},
        },
    ],
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate an offline HTML financial and DCF dashboard.")
    parser.add_argument(
        "data_dir",
        nargs="?",
        default=None,
        help="Directory containing Info.csv and source files. If omitted, a folder picker will open.",
    )
    parser.add_argument("--data-dir", dest="data_dir_flag", default=None, help="Same as positional data_dir.")
    return parser.parse_args()


def configure_paths(data_dir_value: Optional[str]) -> Path:
    from afda.pipeline_utils import prompt_data_dir_with_dialog, resolve_data_dir, set_results_dir

    data_dir = resolve_data_dir(data_dir_value) if data_dir_value else prompt_data_dir_with_dialog()
    data_dir = data_dir.expanduser().resolve()
    set_results_dir(data_dir / "results")
    return data_dir


def load_dataset(data_dir: Path) -> Dict[str, object]:
    from afda.generate_dcf_valuation import build_historical_dataset
    from afda.pipeline_utils import find_info_file

    info_path = find_info_file(data_dir)
    if info_path is None:
        raise FileNotFoundError(f"Info.csv not found in {data_dir}. Cannot generate DCF HTML dashboard.")
    return build_historical_dataset(data_dir=data_dir, info_path=info_path)


def as_float(value: object, fallback: float = 0.0) -> float:
    try:
        if value is None:
            return fallback
        out = float(value)
        if math.isnan(out) or math.isinf(out):
            return fallback
        return out
    except (TypeError, ValueError):
        return fallback


def read_assumptions(data: Dict[str, object], workbook_path: Path) -> Dict[str, object]:
    config = data["valuation_config"]
    assumptions = {
        "shares_outstanding": float(data["shares_outstanding"]),
        "current_price": float(data["current_price"]),
        "net_cash": float(data["cash"][-1] - data["short_debt"][-1] - data["long_debt"][-1]),
        "wacc": float(config["dcf"]["wacc"]),
        "terminal_growth": float(config["dcf"]["terminal_growth"]),
        "minority_interest": float(data["minority_interest"][-1]),
        "long_term_investments": float(data["long_term_investments"][-1]),
        "non_op_current_assets": float(data["non_op_current_assets"][-1]),
        "growths": [float(x) for x in data["default_growths"]],
        "ebit_margins": [float(data["base_ebit_margin"])] * 5,
        "tax_rates": [float(data["base_tax_rate"])] * 5,
        "da_ratios": [float(data["base_da_ratio"])] * 5,
        "capex_ratios": [float(data["base_capex_ratio"])] * 5,
        "nwc_ratios": [float(data["base_nwc_ratio"])] * 5,
    }
    if not workbook_path.exists():
        return assumptions

    wb = load_workbook(workbook_path, data_only=False, read_only=True)
    if "Assumptions" not in wb.sheetnames:
        return assumptions
    ws = wb["Assumptions"]
    for key, cell in {
        "shares_outstanding": "B3",
        "current_price": "B4",
        "net_cash": "B5",
        "wacc": "B6",
        "terminal_growth": "B7",
        "minority_interest": "B8",
        "long_term_investments": "B9",
        "non_op_current_assets": "B10",
    }.items():
        assumptions[key] = as_float(ws[cell].value, float(assumptions[key]))

    for key, row in {
        "growths": 14,
        "ebit_margins": 15,
        "tax_rates": 16,
        "da_ratios": 17,
        "capex_ratios": 18,
        "nwc_ratios": 19,
    }.items():
        assumptions[key] = [
            as_float(ws.cell(row=row, column=col).value, float(assumptions[key][col - 2]))
            for col in range(2, 7)
        ]
    return assumptions


def compute_dcf(data: Dict[str, object], assumptions: Dict[str, object]) -> Dict[str, object]:
    forecast_rows = []
    prev_revenue = float(data["revenue"][-1])
    prev_nwc = float(data["nwc"][-1])
    for idx, year in enumerate(data["forecast_years"]):
        revenue = prev_revenue * (1 + float(assumptions["growths"][idx]))
        ebit = revenue * float(assumptions["ebit_margins"][idx])
        tax_on_ebit = ebit * float(assumptions["tax_rates"][idx])
        nopat = ebit - tax_on_ebit
        da = revenue * float(assumptions["da_ratios"][idx])
        capex = revenue * float(assumptions["capex_ratios"][idx])
        operating_nwc = revenue * float(assumptions["nwc_ratios"][idx])
        change_nwc = operating_nwc - prev_nwc
        fcff = nopat + da - capex - change_nwc
        forecast_rows.append(
            {
                "year": int(year),
                "revenue": revenue,
                "growth": float(assumptions["growths"][idx]),
                "ebit_margin": float(assumptions["ebit_margins"][idx]),
                "tax_rate": float(assumptions["tax_rates"][idx]),
                "da_ratio": float(assumptions["da_ratios"][idx]),
                "capex_ratio": float(assumptions["capex_ratios"][idx]),
                "nwc_ratio": float(assumptions["nwc_ratios"][idx]),
                "ebit": ebit,
                "nopat": nopat,
                "da": da,
                "capex": capex,
                "operating_nwc": operating_nwc,
                "change_nwc": change_nwc,
                "fcff": fcff,
            }
        )
        prev_revenue = revenue
        prev_nwc = operating_nwc

    wacc = float(assumptions["wacc"])
    terminal_growth = float(assumptions["terminal_growth"])
    if wacc <= terminal_growth:
        terminal_growth = max(wacc - 0.005, 0.0)
    terminal_value = forecast_rows[-1]["fcff"] * (1 + terminal_growth) / (wacc - terminal_growth)
    pv_fcff = [row["fcff"] / ((1 + wacc) ** (idx + 1)) for idx, row in enumerate(forecast_rows)]
    pv_terminal = terminal_value / ((1 + wacc) ** len(forecast_rows))
    enterprise_value = sum(pv_fcff) + pv_terminal
    equity_value = (
        enterprise_value
        + float(assumptions["net_cash"])
        - float(assumptions["minority_interest"])
        + float(assumptions["long_term_investments"])
        + float(assumptions["non_op_current_assets"])
    )
    intrinsic_price = equity_value / float(assumptions["shares_outstanding"]) if assumptions["shares_outstanding"] else 0.0
    current_price = float(assumptions["current_price"])
    upside = intrinsic_price / current_price - 1 if current_price else 0.0
    safety_margin = 1 - current_price / intrinsic_price if intrinsic_price > 0 else 0.0
    return {
        "forecast_rows": forecast_rows,
        "pv_fcff": pv_fcff,
        "pv_terminal": pv_terminal,
        "terminal_value": terminal_value,
        "enterprise_value": enterprise_value,
        "equity_value": equity_value,
        "intrinsic_price": intrinsic_price,
        "upside": upside,
        "safety_margin": safety_margin,
    }


def ensure_echarts_asset(assets_dir: Path) -> str:
    assets_dir.mkdir(parents=True, exist_ok=True)
    target = assets_dir / "echarts.min.js"
    if not LOCAL_ECHARTS_SOURCE.exists():
        raise FileNotFoundError(f"Offline ECharts asset not found: {LOCAL_ECHARTS_SOURCE}")
    shutil.copy2(LOCAL_ECHARTS_SOURCE, target)
    return "_assets/echarts.min.js"


def money(value: float) -> str:
    abs_value = abs(value)
    if abs_value >= 1e8:
        return f"{value / 1e8:,.2f} 亿"
    if abs_value >= 1e4:
        return f"{value / 1e4:,.2f} 万"
    return f"{value:,.2f}"


def percent(value: float) -> str:
    return f"{value * 100:,.1f}%"


def esc(value: object) -> str:
    return html.escape(str(value), quote=True)


def statement_value(value: str) -> str:
    text = (value or "").strip()
    if not text:
        return ""
    try:
        return money(float(text))
    except ValueError:
        return esc(text)


def load_statement_tables(results_dir: Path) -> Dict[str, Dict[str, object]]:
    statements: Dict[str, Dict[str, object]] = {}
    for key, (title, folder, filename) in STATEMENT_SOURCES.items():
        path = results_dir / "04_rebuilt_statements" / folder / filename
        rel_path = f"04_rebuilt_statements/{folder}/{filename}"
        if not path.exists():
            statements[key] = {
                "title": title,
                "csvPath": rel_path,
                "csvText": "",
                "years": [],
                "series": {},
                "charts": STATEMENT_CHARTS[key],
                "headers": ["提示"],
                "rows": [[esc(f"未找到文件：{path}")]],
            }
            continue

        csv_text = path.read_text(encoding="utf-8-sig")
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.reader(fh)
            headers = next(reader, [])
            years = headers[1:]
            series = {}
            rows = []
            for raw_row in reader:
                if not raw_row or not any(cell.strip() for cell in raw_row):
                    continue
                item = esc(raw_row[0]) if raw_row else ""
                item_name = raw_row[0].strip()
                series[item_name] = [as_float(cell) for cell in raw_row[1:]]
                values = [statement_value(cell) for cell in raw_row[1:]]
                row_class = "emphasis-row" if key == "balance_sheet" and item_name in BALANCE_SHEET_EMPHASIS_ROWS else ""
                rows.append({"cells": [item, *values], "class": row_class})
        statements[key] = {
            "title": title,
            "csvPath": rel_path,
            "csvText": csv_text,
            "years": years,
            "series": series,
            "charts": STATEMENT_CHARTS[key],
            "headers": headers,
            "rows": rows,
        }
    return statements


def table(headers: Iterable[str], rows: Iterable[Iterable[str]]) -> str:
    head = "".join(f"<th>{esc(h)}</th>" for h in headers)
    body = ""
    for row in rows:
        if isinstance(row, dict):
            cells = row.get("cells", [])
            row_class = f' class="{esc(row.get("class", ""))}"' if row.get("class") else ""
        else:
            cells = row
            row_class = ""
        body += f"<tr{row_class}>" + "".join(f"<td>{cell}</td>" for cell in cells) + "</tr>"
    return f"<table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>"


def statement_section(statement: Dict[str, object]) -> str:
    charts = "".join(chart_container(chart["id"], chart["title"]) for chart in statement["charts"])
    return f"""
      <section>
        <h2>{esc(statement["title"])}可视化</h2>
        <div class="grid charts statement-charts">
          {charts}
        </div>
      </section>
      <section>
        <h2>{esc(statement["title"])}</h2>
        <div class="statement-table">
          {table(statement["headers"], statement["rows"])}
        </div>
      </section>
    """


def chart_container(chart_id: str, title: str) -> str:
    return f"""
    <div class="chart-card">
      <div class="chart-title">{esc(title)}</div>
      <div id="{esc(chart_id)}" class="echart"></div>
    </div>
    """


def build_assumption_inputs(years: list[int], assumptions: Dict[str, object]) -> str:
    rows = [
        ("growths", "Revenue Growth"),
        ("ebit_margins", "EBIT Margin"),
        ("tax_rates", "Tax Rate"),
        ("da_ratios", "D&A / Revenue"),
        ("capex_ratios", "Capex / Revenue"),
        ("nwc_ratios", "NWC / Revenue"),
    ]
    body = []
    for key, label in rows:
        cells = [f"<td>{esc(label)}</td>"]
        for idx, value in enumerate(assumptions[key]):
            cells.append(
                f'<td><input class="assumption-input" data-array="{key}" data-index="{idx}" '
                f'type="number" step="0.1" value="{float(value) * 100:.2f}"><span class="unit">%</span></td>'
            )
        body.append("<tr>" + "".join(cells) + "</tr>")
    header = "".join(f"<th>{year}E</th>" for year in years)
    return f"""
    <table class="input-table">
      <thead><tr><th>Assumption</th>{header}</tr></thead>
      <tbody>{''.join(body)}</tbody>
    </table>
    """


def build_html(
    data: Dict[str, object],
    assumptions: Dict[str, object],
    dcf: Dict[str, object],
    echarts_src: str,
    statements: Dict[str, Dict[str, object]],
) -> str:
    years = [int(y) for y in data["years"]]
    forecast_years = [int(y) for y in data["forecast_years"]]
    revenue = [float(x) for x in data["revenue"]]
    ebit = [float(x) for x in data["ebit"]]
    cfo = [float(x) for x in data["cfo"]]
    fcff = [float(x) for x in data["fcff_proxy"]]
    net_profit = [float(x) for x in data["parent_net_profit"]]
    ebit_margin = [float(x) for x in data["ebit_margin"]]

    hist_rows = [
        [esc(year), esc(money(revenue[i])), esc(money(ebit[i])), esc(percent(ebit_margin[i])), esc(money(net_profit[i])), esc(money(cfo[i])), esc(money(fcff[i]))]
        for i, year in enumerate(years)
    ]
    statement_payload = {
        key: {
            "title": statement["title"],
            "csvPath": statement["csvPath"],
            "csvText": statement["csvText"],
            "years": statement["years"],
            "series": statement["series"],
            "charts": statement["charts"],
        }
        for key, statement in statements.items()
    }

    payload = {
        "company": {"ticker": data["ticker"], "name": data["company_name"], "valuationDate": data["valuation_date"]},
        "historical": {
            "years": years,
            "revenue": revenue,
            "ebit": ebit,
            "netProfit": net_profit,
            "cfo": cfo,
            "fcffProxy": fcff,
            "ebitMargin": ebit_margin,
            "baseRevenue": revenue[-1],
            "baseNwc": float(data["nwc"][-1]),
        },
        "forecastYears": forecast_years,
        "assumptions": assumptions,
        "initialDcf": dcf,
        "statements": statement_payload,
    }
    payload_json = json.dumps(payload, ensure_ascii=False)
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    initial_badge = "good" if dcf["upside"] >= 0.15 else "neutral" if dcf["upside"] >= -0.10 else "risk"

    cards = [
        ("当前股价", f"{assumptions['current_price']:.2f}", "来自 Info.csv / Assumptions"),
        ("DCF 每股内在价值", f"{dcf['intrinsic_price']:.2f}", f"安全边际 {percent(dcf['safety_margin'])}"),
        ("企业价值 EV", money(dcf["enterprise_value"]), "5年 FCFF + Terminal Value"),
        (f"{years[-1]}A 营收", money(revenue[-1]), f"EBIT Margin {percent(ebit_margin[-1])}"),
        (f"{years[-1]}A 经营现金流", money(cfo[-1]), f"FCFF Proxy {money(fcff[-1])}"),
        ("股东权益价值", money(dcf["equity_value"]), "EV 到 Equity Bridge"),
    ]
    card_html = "".join(
        f'<article class="metric-card"><span>{esc(label)}</span><strong>{esc(value)}</strong><small>{esc(note)}</small></article>'
        for label, value, note in cards
    )

    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{esc(data['company_name'])} 财务与DCF估值</title>
  <script src="{esc(echarts_src)}"></script>
  <style>
    :root {{
      --bg:#f6f7fb; --panel:#fff; --ink:#172033; --muted:#687385; --line:#dfe5ef;
      --accent:#2563eb; --teal:#0f766e; --red:#b91c1c; --shadow:0 18px 44px rgba(23,32,51,.08);
    }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; background:var(--bg); color:var(--ink); font-family:"Segoe UI","Microsoft YaHei",Arial,sans-serif; line-height:1.5; }}
    header {{ background:#0f172a; color:#fff; padding:30px 28px 24px; }}
    .wrap {{ max-width:1220px; margin:0 auto; }}
    .topline {{ display:flex; justify-content:space-between; gap:18px; align-items:flex-start; flex-wrap:wrap; }}
    h1 {{ margin:0; font-size:29px; letter-spacing:0; }}
    h2 {{ margin:0 0 14px; font-size:20px; }}
    h3 {{ margin:0 0 10px; font-size:16px; }}
    .subtitle {{ color:#b9c4d8; margin-top:8px; }}
    .badge {{ padding:7px 12px; border-radius:999px; background:rgba(255,255,255,.12); border:1px solid rgba(255,255,255,.18); font-weight:700; }}
    .badge.good {{ color:#bbf7d0; }} .badge.neutral {{ color:#fde68a; }} .badge.risk {{ color:#fecaca; }}
    main {{ padding:22px 28px 42px; }}
    .tabs {{ display:flex; gap:8px; margin:0 auto 16px; max-width:1220px; }}
    .tab-btn {{ border:1px solid var(--line); background:#fff; color:var(--ink); border-radius:8px; padding:10px 14px; font-weight:740; cursor:pointer; }}
    .tab-btn.active {{ background:#172033; color:#fff; border-color:#172033; }}
    .tab-pane {{ display:none; }} .tab-pane.active {{ display:block; }}
    .grid {{ display:grid; gap:16px; }}
    .metrics {{ grid-template-columns:repeat(3,minmax(0,1fr)); }}
    .two {{ grid-template-columns:1.05fr .95fr; }}
    .charts {{ grid-template-columns:repeat(2,minmax(0,1fr)); }}
    .three {{ grid-template-columns:repeat(3,minmax(0,1fr)); }}
    section, .metric-card, .chart-card {{ background:var(--panel); border:1px solid var(--line); border-radius:8px; box-shadow:var(--shadow); }}
    section {{ padding:20px; margin-top:16px; overflow:hidden; }}
    .metric-card {{ padding:17px; min-height:116px; }}
    .metric-card span, .metric-card small {{ color:var(--muted); display:block; }}
    .metric-card strong {{ display:block; font-size:25px; margin:7px 0 5px; }}
    .chart-card {{ padding:16px; box-shadow:none; }}
    .chart-title {{ font-weight:760; margin-bottom:8px; }}
    .echart {{ width:100%; height:320px; }}
    table {{ width:100%; border-collapse:collapse; font-size:14px; }}
    th,td {{ padding:9px 10px; border-bottom:1px solid var(--line); text-align:right; white-space:nowrap; }}
    th:first-child,td:first-child {{ text-align:left; }}
    th {{ color:#465366; background:#f3f6fb; font-weight:740; }}
    .statement-table {{ overflow:auto; border:1px solid var(--line); border-radius:8px; }}
    .statement-table table {{ min-width:980px; }}
    .statement-table th:first-child,.statement-table td:first-child {{ position:sticky; left:0; z-index:1; background:#fff; min-width:260px; max-width:360px; white-space:normal; }}
    .statement-table th:first-child {{ z-index:2; background:#f3f6fb; }}
    .statement-table tr.emphasis-row td {{ font-weight:800; background:#eef4ff; color:#0f172a; }}
    .statement-table tr.emphasis-row td:first-child {{ background:#e4edff; }}
    .analysis-panel section {{ margin-top:18px; }}
    .analysis-hero {{ background:#111827; color:#fff; border-color:#111827; }}
    .analysis-hero h2 {{ color:#fff; }}
    .analysis-hero .note {{ color:#cbd5e1; }}
    .analysis-cards {{ grid-template-columns:repeat(4,minmax(0,1fr)); }}
    .analysis-card {{ background:#fff; border:1px solid var(--line); border-radius:8px; padding:14px; min-height:118px; }}
    .analysis-card span {{ display:block; color:var(--muted); font-size:12px; }}
    .analysis-card strong {{ display:block; font-size:20px; margin:6px 0 4px; }}
    .analysis-card small {{ display:block; color:var(--muted); }}
    .status-pill {{ display:inline-block; margin-top:8px; padding:3px 8px; border-radius:999px; font-size:12px; font-weight:750; background:#eef2ff; color:#1d4ed8; }}
    .status-pill.good {{ background:#dcfce7; color:#166534; }}
    .status-pill.warn {{ background:#fef3c7; color:#92400e; }}
    .status-pill.risk {{ background:#fee2e2; color:#991b1b; }}
    .analysis-text {{ margin-top:12px; padding:12px 14px; background:#f8fafc; border:1px solid var(--line); border-radius:8px; color:#334155; }}
    .risk-list {{ display:grid; gap:10px; margin:0; padding:0; list-style:none; }}
    .risk-list li {{ padding:11px 12px; border:1px solid var(--line); border-left:4px solid #f97316; border-radius:8px; background:#fff; }}
    .risk-list li.good {{ border-left-color:#10b981; }}
    .analysis-table {{ overflow:auto; border:1px solid var(--line); border-radius:8px; }}
    #valuation {{ display:grid; grid-template-columns:minmax(0,1fr) 260px; gap:12px; align-items:start; }}
    #valuation section {{ margin-top:16px; }}
    #valuation section:nth-of-type(1) {{ grid-column:1; grid-row:1; }}
    #valuation section:nth-of-type(2) {{ grid-column:2; grid-row:1; padding:14px; }}
    #valuation section:nth-of-type(n+3) {{ grid-column:1 / -1; }}
    #valuation section:nth-of-type(2) h2 {{ font-size:20px; margin-bottom:10px; }}
    #valuation section:nth-of-type(2) .three {{ grid-template-columns:1fr; gap:8px; }}
    #valuation section:nth-of-type(2) .three > div {{ background:#f7f9fc; border:1px solid var(--line); border-radius:8px; padding:9px 10px; }}
    #valuation section:nth-of-type(2) .note {{ font-size:12px; margin-top:0; }}
    .input-grid {{ display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:8px; }}
    label.input-card {{ display:block; background:#f7f9fc; border:1px solid var(--line); border-radius:8px; padding:8px 9px; }}
    label.input-card span {{ display:block; color:var(--muted); font-size:11px; margin-bottom:4px; line-height:1.25; }}
    input {{ width:100%; border:1px solid #cbd5e1; border-radius:7px; padding:6px 7px; font:inherit; font-size:13px; text-align:right; background:#fff; }}
    .input-table {{ font-size:12px; }}
    .input-table th,.input-table td {{ padding:6px 7px; }}
    .input-table input {{ width:64px; display:inline-block; }}
    .unit {{ color:var(--muted); margin-left:4px; }}
    .kpi-large {{ font-size:22px; font-weight:800; }}
    .note {{ color:var(--muted); font-size:13px; margin-top:10px; }}
    .toolbar {{ display:flex; gap:10px; flex-wrap:wrap; margin-top:12px; }}
    .action-btn {{ border:1px solid var(--line); background:#fff; border-radius:8px; padding:9px 12px; cursor:pointer; font-weight:700; }}
    .action-btn.primary {{ background:var(--accent); color:#fff; border-color:var(--accent); }}
    @media(max-width:1100px) {{ #valuation {{ display:block; }} }}
    @media(max-width:900px) {{ .metrics,.two,.charts,.three,.input-grid,.analysis-cards {{ grid-template-columns:1fr; }} section {{ overflow-x:auto; }} }}
  </style>
</head>
<body>
  <header>
    <div class="wrap topline">
      <div>
        <h1>{esc(data['company_name'])} ({esc(data['ticker'])}) 财务与 DCF 估值</h1>
        <div class="subtitle">估值日期：{esc(data['valuation_date'])} · 生成时间：{esc(generated_at)} · 离线 HTML</div>
      </div>
      <div id="headerBadge" class="badge {initial_badge}">目标价 <span id="headerPrice">{dcf['intrinsic_price']:.2f}</span> · <span id="headerUpside">{percent(dcf['upside'])}</span></div>
    </div>
  </header>
  <main>
    <div class="tabs">
      <button class="tab-btn active" data-tab="financial">财务基本情况</button>
      <button class="tab-btn" data-tab="analysis">财务分析可视化面板</button>
      <button class="tab-btn" data-tab="balanceSheet">资产负债表</button>
      <button class="tab-btn" data-tab="incomeStatement">利润表</button>
      <button class="tab-btn" data-tab="cashFlow">现金流量表</button>
      <button class="tab-btn" data-tab="valuation">估值过程与假设联动</button>
    </div>

    <div id="financial" class="tab-pane active wrap">
      <div class="grid metrics">{card_html}</div>
      <section>
        <h2>历史核心趋势</h2>
        <div class="grid charts">
          {chart_container("revenueTrendChart", "收入、EBIT、归母净利润")}
          {chart_container("cashFlowChart", "CFO 与 FCFF Proxy")}
        </div>
      </section>
      <section>
        <h2>历史关键财务数据</h2>
        {table(["年度", "Revenue", "EBIT", "EBIT Margin", "归母净利润", "CFO", "FCFF Proxy"], hist_rows)}
      </section>
    </div>

    <div id="analysis" class="tab-pane wrap analysis-panel">
      <section class="analysis-hero">
        <h2>财务分析可视化面板</h2>
        <div class="note">基于三张标准化财务报表 CSV 动态计算成长、盈利、运营效率、偿债、现金流质量、三表联动和风险提示。</div>
      </section>
      <section>
        <h2>顶部概览</h2>
        <div id="analysisError" class="note"></div>
        <div id="analysisCards" class="grid analysis-cards"></div>
      </section>
      <section>
        <h2>成长能力</h2>
        <div class="grid charts">
          {chart_container("analysisRevenueTrend", "Revenue 趋势")}
          {chart_container("analysisRevenueYoY", "Revenue YoY 趋势")}
          {chart_container("analysisProfitTrend", "Net Profit / Adjusted Net Profit")}
          {chart_container("analysisProfitGrowth", "利润增速对比")}
        </div>
        <div id="growthAnalysisText" class="analysis-text"></div>
      </section>
      <section>
        <h2>盈利能力</h2>
        <div class="grid charts">
          {chart_container("analysisMargins", "毛利率、经营利润率、净利率")}
          {chart_container("analysisReturns", "ROE / ROA")}
          {chart_container("analysisExpenseRatios", "费用率趋势")}
          {chart_container("analysisDupont", "杜邦分解")}
        </div>
        <div id="profitabilityAnalysisText" class="analysis-text"></div>
      </section>
      <section>
        <h2>资产负债与偿债能力</h2>
        <div class="grid charts">
          {chart_container("analysisAssetStructure", "资产结构变化")}
          {chart_container("analysisLiabilityStructure", "负债结构")}
          {chart_container("analysisDebtRatios", "资产负债率与有息负债率")}
          {chart_container("analysisDebtTrend", "有息负债与净债务")}
          {chart_container("analysisCashDebt", "现金短债比")}
        </div>
        <div id="solvencyAnalysisText" class="analysis-text"></div>
      </section>
      <section>
        <h2>现金流质量</h2>
        <div class="grid charts">
          {chart_container("analysisOCFProfit", "经营现金流 vs 净利润")}
          {chart_container("analysisFCFTrend", "自由现金流趋势")}
          {chart_container("analysisCashQualityRatios", "现金流质量指标")}
          {chart_container("analysisCapexRevenue", "Capex / Revenue")}
        </div>
        <div id="cashFlowAnalysisText" class="analysis-text"></div>
      </section>
      <section>
        <h2>三表联动</h2>
        {chart_container("analysisThreeStatement", "收入、利润、现金流、资产与有息负债")}
        <div id="threeStatementAnalysisText" class="analysis-text"></div>
      </section>
      <section>
        <h2>财务风险提示</h2>
        <ul id="riskWarnings" class="risk-list"></ul>
      </section>
      <section>
        <h2>初步估值输入指标</h2>
        <div id="valuationInputTable" class="analysis-table"></div>
      </section>
    </div>

    <div id="balanceSheet" class="tab-pane wrap">
      {statement_section(statements["balance_sheet"])}
    </div>

    <div id="incomeStatement" class="tab-pane wrap">
      {statement_section(statements["income_statement"])}
    </div>

    <div id="cashFlow" class="tab-pane wrap">
      {statement_section(statements["cash_flow"])}
    </div>

    <div id="valuation" class="tab-pane wrap">
      <section>
        <h2>可编辑 Assumptions</h2>
        <div class="input-grid">
          <label class="input-card"><span>当前股价</span><input id="currentPrice" type="number" step="0.01" value="{float(assumptions['current_price']):.4f}"></label>
          <label class="input-card"><span>总股本</span><input id="sharesOutstanding" type="number" step="1" value="{float(assumptions['shares_outstanding']):.4f}"></label>
          <label class="input-card"><span>WACC (%)</span><input id="wacc" type="number" step="0.1" value="{float(assumptions['wacc']) * 100:.2f}"></label>
          <label class="input-card"><span>永续增长率 (%)</span><input id="terminalGrowth" type="number" step="0.1" value="{float(assumptions['terminal_growth']) * 100:.2f}"></label>
          <label class="input-card"><span>净现金 / (净债务)</span><input id="netCash" type="number" step="10000" value="{float(assumptions['net_cash']):.4f}"></label>
          <label class="input-card"><span>少数股东权益</span><input id="minorityInterest" type="number" step="10000" value="{float(assumptions['minority_interest']):.4f}"></label>
          <label class="input-card"><span>长期金融及股权投资</span><input id="longTermInvestments" type="number" step="10000" value="{float(assumptions['long_term_investments']):.4f}"></label>
          <label class="input-card"><span>非经营流动资产</span><input id="nonOpCurrentAssets" type="number" step="10000" value="{float(assumptions['non_op_current_assets']):.4f}"></label>
        </div>
        <div style="margin-top:14px">{build_assumption_inputs(forecast_years, assumptions)}</div>
        <div class="toolbar">
          <button id="resetAssumptions" class="action-btn">恢复初始假设</button>
          <button id="copyAssumptions" class="action-btn primary">使用当前预测数据</button>
        </div>
        <div class="note">修改任意输入后，下面的预测表、估值桥、目标价和图表会自动联动重算。百分比请输入百分数，例如 10 表示 10%。</div>
      </section>

      <section>
        <h2>估值结论</h2>
        <div class="grid three">
          <div><span class="note">DCF 每股内在价值</span><div id="valuationPrice" class="kpi-large"></div></div>
          <div><span class="note">相对当前股价空间</span><div id="valuationUpside" class="kpi-large"></div></div>
          <div><span class="note">企业价值 EV</span><div id="valuationEv" class="kpi-large"></div></div>
          <div><span class="note">安全边际</span><div id="valuationMos" class="kpi-large"></div></div>
        </div>
      </section>

      <section>
        <h2>估值全过程</h2>
        <div class="grid two">
          <div>{chart_container("forecastChart", "预测期 Revenue 与 FCFF")}</div>
          <div>
            <h3>DCF Bridge</h3>
            <table><tbody id="bridgeBody"></tbody></table>
          </div>
        </div>
      </section>

      <section>
        <h2>预测明细</h2>
        <table>
          <thead><tr><th>年度</th><th>Revenue</th><th>增长率</th><th>EBIT Margin</th><th>NOPAT</th><th>D&A</th><th>Capex</th><th>ΔNWC</th><th>FCFF</th><th>PV FCFF</th></tr></thead>
          <tbody id="forecastBody"></tbody>
        </table>
      </section>
    </div>
  </main>

  <script>
    const model = {payload_json};
    let initialAssumptions = JSON.parse(JSON.stringify(model.assumptions));
    const chartRefs = {{}};
    const analysisState = {{ initialized: false }};

    const num = (id) => Number(document.getElementById(id).value || 0);
    const pct = (id) => num(id) / 100;
    const money = (value) => {{
      const n = Number(value || 0), abs = Math.abs(n);
      if (abs >= 100000000) return (n / 100000000).toFixed(2) + ' 亿';
      if (abs >= 10000) return (n / 10000).toFixed(2) + ' 万';
      return n.toLocaleString('zh-CN', {{ maximumFractionDigits: 2 }});
    }};
    const price = (value) => Number(value || 0).toFixed(2);
    const percent = (value) => (Number(value || 0) * 100).toFixed(1) + '%';

    function collectAssumptions() {{
      const a = {{
        current_price: num('currentPrice'),
        shares_outstanding: num('sharesOutstanding'),
        wacc: pct('wacc'),
        terminal_growth: pct('terminalGrowth'),
        net_cash: num('netCash'),
        minority_interest: num('minorityInterest'),
        long_term_investments: num('longTermInvestments'),
        non_op_current_assets: num('nonOpCurrentAssets'),
        growths: [],
        ebit_margins: [],
        tax_rates: [],
        da_ratios: [],
        capex_ratios: [],
        nwc_ratios: []
      }};
      document.querySelectorAll('.assumption-input').forEach(input => {{
        a[input.dataset.array][Number(input.dataset.index)] = Number(input.value || 0) / 100;
      }});
      return a;
    }}

    function compute(a) {{
      const rows = [];
      let prevRevenue = model.historical.baseRevenue;
      let prevNwc = model.historical.baseNwc;
      model.forecastYears.forEach((year, i) => {{
        const revenue = prevRevenue * (1 + a.growths[i]);
        const ebit = revenue * a.ebit_margins[i];
        const taxOnEbit = ebit * a.tax_rates[i];
        const nopat = ebit - taxOnEbit;
        const da = revenue * a.da_ratios[i];
        const capex = revenue * a.capex_ratios[i];
        const operatingNwc = revenue * a.nwc_ratios[i];
        const changeNwc = operatingNwc - prevNwc;
        const fcff = nopat + da - capex - changeNwc;
        rows.push({{ year, revenue, growth:a.growths[i], ebitMargin:a.ebit_margins[i], nopat, da, capex, changeNwc, fcff }});
        prevRevenue = revenue;
        prevNwc = operatingNwc;
      }});
      let terminalGrowth = a.terminal_growth;
      if (a.wacc <= terminalGrowth) terminalGrowth = Math.max(a.wacc - 0.005, 0);
      const terminalValue = rows[rows.length - 1].fcff * (1 + terminalGrowth) / (a.wacc - terminalGrowth);
      const pvFcff = rows.map((row, i) => row.fcff / Math.pow(1 + a.wacc, i + 1));
      const pvTerminal = terminalValue / Math.pow(1 + a.wacc, rows.length);
      const enterpriseValue = pvFcff.reduce((x, y) => x + y, 0) + pvTerminal;
      const equityValue = enterpriseValue + a.net_cash - a.minority_interest + a.long_term_investments + a.non_op_current_assets;
      const intrinsicPrice = a.shares_outstanding ? equityValue / a.shares_outstanding : 0;
      const upside = a.current_price ? intrinsicPrice / a.current_price - 1 : 0;
      const safetyMargin = intrinsicPrice > 0 ? 1 - a.current_price / intrinsicPrice : 0;
      return {{ rows, pvFcff, pvTerminal, terminalValue, enterpriseValue, equityValue, intrinsicPrice, upside, safetyMargin }};
    }}

    function initChart(id, option) {{
      const el = document.getElementById(id);
      if (!el || !window.echarts) return null;
      const chart = chartRefs[id] || echarts.init(el, null, {{ renderer: 'canvas' }});
      chartRefs[id] = chart;
      chart.setOption(option);
      return chart;
    }}

    function chartBase(xData, series) {{
      return {{
        color: ['#2563eb', '#10b981', '#f97316', '#7c3aed'],
        tooltip: {{ trigger: 'axis', valueFormatter: money }},
        legend: {{ top: 0, textStyle: {{ color: '#687385' }} }},
        grid: {{ left: 72, right: 26, top: 42, bottom: 42 }},
        xAxis: {{ type: 'category', data: xData, axisLabel: {{ color: '#687385' }} }},
        yAxis: {{ type: 'value', axisLabel: {{ color: '#687385', formatter: money }}, splitLine: {{ lineStyle: {{ color: '#e8edf5' }} }} }},
        series
      }};
    }}

    function forecastComboChart(xData, rows) {{
      return {{
        color: ['#2563eb', '#7c3aed'],
        tooltip: {{ trigger: 'axis', valueFormatter: money }},
        legend: {{ top: 0, textStyle: {{ color: '#687385' }} }},
        grid: {{ left: 72, right: 72, top: 42, bottom: 42 }},
        xAxis: {{ type: 'category', data: xData, axisLabel: {{ color: '#687385' }} }},
        yAxis: [
          {{ type: 'value', name: 'Revenue', axisLabel: {{ color: '#687385', formatter: money }}, splitLine: {{ lineStyle: {{ color: '#e8edf5' }} }} }},
          {{ type: 'value', name: 'FCFF', axisLabel: {{ color: '#687385', formatter: money }}, splitLine: {{ show: false }} }}
        ],
        series: [
          {{ name:'Revenue', type:'line', yAxisIndex:0, data:rows.map(r=>r.revenue), smooth:true, symbolSize:7 }},
          {{ name:'FCFF', type:'bar', yAxisIndex:1, data:rows.map(r=>r.fcff), barMaxWidth:38 }}
        ]
      }};
    }}

    function cleanNumber(value) {{
      if (value === null || value === undefined) return null;
      const text = String(value).trim().replace(/,/g, '');
      if (!text || text === '--' || text.toLowerCase() === 'nan') return null;
      const numValue = Number(text);
      return Number.isFinite(numValue) ? numValue : null;
    }}

    function parseCSV(text) {{
      const rows = [];
      let row = [], cell = '', inQuotes = false;
      for (let i = 0; i < text.length; i++) {{
        const ch = text[i], next = text[i + 1];
        if (ch === '"' && inQuotes && next === '"') {{
          cell += '"'; i += 1;
        }} else if (ch === '"') {{
          inQuotes = !inQuotes;
        }} else if (ch === ',' && !inQuotes) {{
          row.push(cell); cell = '';
        }} else if ((ch === '\\n' || ch === '\\r') && !inQuotes) {{
          if (ch === '\\r' && next === '\\n') i += 1;
          row.push(cell);
          if (row.some(x => String(x).trim())) rows.push(row);
          row = []; cell = '';
        }} else {{
          cell += ch;
        }}
      }}
      row.push(cell);
      if (row.some(x => String(x).trim())) rows.push(row);
      return rows;
    }}

    async function loadCSV(source) {{
      if (source.csvPath) {{
        try {{
          const response = await fetch(source.csvPath);
          if (response.ok) return await response.text();
        }} catch (err) {{
          // Offline file:// openings often block fetch; embedded CSV keeps the dashboard usable.
        }}
      }}
      if (source.csvText) return source.csvText;
      throw new Error('CSV 数据不足：' + (source.title || source.csvPath || 'unknown'));
    }}

    function parseWideFinancialTable(csvText) {{
      const rows = parseCSV(csvText);
      if (!rows.length) return {{ years: [], series: {{}}, rawRows: [] }};
      const headers = rows[0].map(x => String(x || '').replace(/^\\uFEFF/, '').trim());
      const years = headers.slice(1).map(x => String(x).trim());
      const series = {{}};
      rows.slice(1).forEach(row => {{
        const item = String(row[0] || '').trim();
        if (!item) return;
        series[item] = years.map((_, i) => cleanNumber(row[i + 1]));
      }});
      return {{ years, series, rawRows: rows.slice(1), originalHeaders: headers }};
    }}

    const FIELD_ALIASES = {{
      revenue: ['Revenue'],
      cogs: ['COGS'],
      operatingProfit: ['Operating Profit'],
      netProfit: ['Net Profit'],
      parentNetProfit: ['Parent Net Profit'],
      adjustedNetProfit: ['Adjusted Net Profit'],
      sellingExpense: ['Selling Expense'],
      adminExpense: ['Admin Expense'],
      rdExpense: ['R&D Expense'],
      financialExpense: ['Financial Expense'],
      totalAssets: ['Total Assets'],
      totalEquity: ['Total Equity'],
      totalLiabilities: ['Total Liabilities'],
      cash: ['Cash & Short-term Financial Assets'],
      coreCurrentAssets: ['Core Operating Current Assets'],
      longTermCoreAssets: ['Long-term Core Operating Assets'],
      longTermInvestments: ['Long-term Financial & Equity Investments'],
      riskAssets: ['Risk & Amortizing Assets'],
      shortDebt: ['Interest-bearing Short-term Debt'],
      longDebt: ['Long-term Interest-bearing Debt'],
      operatingCurrentLiabilities: ['Operating Non-interest-bearing Current Liabilities'],
      operatingCashFlow: ['Operating Cash Flow'],
      investingCashFlow: ['Investing Cash Flow'],
      financingCashFlow: ['Financing Cash Flow'],
      capex: ['Capex'],
      cashFromCustomers: ['Cash From Customers'],
      dividendInterestCashOut: ['Dividend & Interest Cash Out']
    }};

    const emptySeries = (len) => Array.from({{ length: len }}, () => null);
    const pickSeries = (table, key, fallbackLen) => {{
      const aliases = FIELD_ALIASES[key] || [key];
      for (const name of aliases) {{
        if (Array.isArray(table.series[name])) return table.series[name];
      }}
      return emptySeries(fallbackLen);
    }};
    const hasField = (table, key) => (FIELD_ALIASES[key] || [key]).some(name => Array.isArray(table.series[name]));
    const mapSeries = (a, fn) => a.map((x, i) => x === null || x === undefined ? null : fn(x, i));
    const binarySeries = (a, b, fn) => a.map((x, i) => {{
      const y = b[i];
      if (x === null || x === undefined || y === null || y === undefined) return null;
      return fn(x, y, i);
    }});
    const ratioSeries = (a, b) => binarySeries(a, b, (x, y) => y ? x / y : null);
    const addSeries = (a, b) => binarySeries(a, b, (x, y) => x + y);
    const subSeries = (a, b) => binarySeries(a, b, (x, y) => x - y);
    const yoySeries = (a) => a.map((x, i) => {{
      const prev = a[i - 1];
      if (i === 0 || x === null || prev === null || prev === undefined || !prev) return null;
      return x / prev - 1;
    }});
    const averageSeries = (a) => a.map((x, i) => {{
      const prev = a[i - 1];
      if (i === 0 || x === null || prev === null || prev === undefined) return null;
      return (x + prev) / 2;
    }});
    const latest = (a) => [...a].reverse().find(x => x !== null && x !== undefined);
    const latestIndex = (a) => {{
      for (let i = a.length - 1; i >= 0; i--) if (a[i] !== null && a[i] !== undefined) return i;
      return -1;
    }};
    const lastN = (a, n) => a.slice(Math.max(0, a.length - n));
    const isRising = (a) => {{
      const xs = a.filter(x => x !== null && x !== undefined);
      return xs.length >= 3 && xs[xs.length - 1] > xs[xs.length - 2] && xs[xs.length - 2] > xs[xs.length - 3];
    }};
    const isFalling = (a) => {{
      const xs = a.filter(x => x !== null && x !== undefined);
      return xs.length >= 3 && xs[xs.length - 1] < xs[xs.length - 2] && xs[xs.length - 2] < xs[xs.length - 3];
    }};

    function calculateGrowthMetrics(ctx) {{
      const m = ctx.metrics;
      m.grossProfit = subSeries(m.revenue, m.cogs); // Gross Profit = Revenue - COGS
      m.revenueYoY = yoySeries(m.revenue);
      m.grossProfitYoY = yoySeries(m.grossProfit);
      m.operatingProfitYoY = yoySeries(m.operatingProfit);
      m.netProfitYoY = yoySeries(m.netProfit);
      m.parentNetProfitYoY = yoySeries(m.parentNetProfit);
      m.adjustedNetProfitYoY = yoySeries(m.adjustedNetProfit);
    }}

    function calculateProfitabilityMetrics(ctx) {{
      const m = ctx.metrics;
      m.avgTotalEquity = averageSeries(m.totalEquity); // Average Total Equity = (current + previous) / 2
      m.avgTotalAssets = averageSeries(m.totalAssets); // Average Total Assets = (current + previous) / 2
      m.grossMargin = ratioSeries(m.grossProfit, m.revenue);
      m.operatingMargin = ratioSeries(m.operatingProfit, m.revenue);
      m.netMargin = ratioSeries(m.netProfit, m.revenue);
      m.parentNetMargin = ratioSeries(m.parentNetProfit, m.revenue);
      m.adjustedNetMargin = ratioSeries(m.adjustedNetProfit, m.revenue);
      m.roe = ratioSeries(m.parentNetProfit, m.avgTotalEquity); // ROE = Parent Net Profit / Average Total Equity
      m.roa = ratioSeries(m.netProfit, m.avgTotalAssets); // ROA = Net Profit / Average Total Assets
      m.sellingExpenseRatio = ratioSeries(m.sellingExpense, m.revenue);
      m.adminExpenseRatio = ratioSeries(m.adminExpense, m.revenue);
      m.rdExpenseRatio = ratioSeries(m.rdExpense, m.revenue);
      m.financialExpenseRatio = ratioSeries(m.financialExpense, m.revenue);
    }}

    function calculateBalanceSheetMetrics(ctx) {{
      const m = ctx.metrics;
      m.interestBearingDebt = addSeries(m.shortDebt, m.longDebt); // Interest-bearing Debt = short debt + long debt
      m.netDebt = subSeries(m.interestBearingDebt, m.cash); // Net Debt = Interest-bearing Debt - cash
      m.debtToAssetRatio = ratioSeries(m.totalLiabilities, m.totalAssets);
      m.interestBearingDebtRatio = ratioSeries(m.interestBearingDebt, m.totalAssets);
      m.netDebtToEquity = ratioSeries(m.netDebt, m.totalEquity);
      m.cashToShortDebt = ratioSeries(m.cash, m.shortDebt);
      m.cashAssetRatio = ratioSeries(m.cash, m.totalAssets);
      m.coreCurrentAssetRatio = ratioSeries(m.coreCurrentAssets, m.totalAssets);
      m.longTermCoreAssetRatio = ratioSeries(m.longTermCoreAssets, m.totalAssets);
      m.longTermInvestmentRatio = ratioSeries(m.longTermInvestments, m.totalAssets);
      m.riskAssetRatio = ratioSeries(m.riskAssets, m.totalAssets);
    }}

    function calculateCashFlowMetrics(ctx) {{
      const m = ctx.metrics;
      m.freeCashFlow = subSeries(m.operatingCashFlow, m.capex); // Free Cash Flow = Operating Cash Flow - Capex
      m.ocfToNetProfit = ratioSeries(m.operatingCashFlow, m.cfNetProfit);
      m.fcfMargin = ratioSeries(m.freeCashFlow, m.revenue);
      m.fcfToNetProfit = ratioSeries(m.freeCashFlow, m.cfNetProfit);
      m.cashCollectionRatio = ratioSeries(m.cashFromCustomers, m.revenue);
      m.capexRevenue = ratioSeries(m.capex, m.revenue);
    }}

    function calculateDuPontMetrics(ctx) {{
      const m = ctx.metrics;
      m.dupontNetMargin = ratioSeries(m.parentNetProfit, m.revenue);
      m.assetTurnover = ratioSeries(m.revenue, m.avgTotalAssets);
      m.equityMultiplier = ratioSeries(m.avgTotalAssets, m.avgTotalEquity);
      m.dupontRoe = binarySeries(binarySeries(m.dupontNetMargin, m.assetTurnover, (x, y) => x * y), m.equityMultiplier, (x, y) => x * y);
    }}

    async function buildAnalysisContext() {{
      const pl = parseWideFinancialTable(await loadCSV(model.statements.income_statement));
      const bs = parseWideFinancialTable(await loadCSV(model.statements.balance_sheet));
      const cf = parseWideFinancialTable(await loadCSV(model.statements.cash_flow));
      const years = pl.years.length ? pl.years : (bs.years.length ? bs.years : cf.years);
      const len = years.length;
      const metrics = {{
        revenue: pickSeries(pl, 'revenue', len),
        cogs: pickSeries(pl, 'cogs', len),
        operatingProfit: pickSeries(pl, 'operatingProfit', len),
        netProfit: pickSeries(pl, 'netProfit', len),
        parentNetProfit: pickSeries(pl, 'parentNetProfit', len),
        adjustedNetProfit: pickSeries(pl, 'adjustedNetProfit', len),
        sellingExpense: pickSeries(pl, 'sellingExpense', len),
        adminExpense: pickSeries(pl, 'adminExpense', len),
        rdExpense: pickSeries(pl, 'rdExpense', len),
        financialExpense: pickSeries(pl, 'financialExpense', len),
        totalAssets: pickSeries(bs, 'totalAssets', len),
        totalEquity: pickSeries(bs, 'totalEquity', len),
        totalLiabilities: pickSeries(bs, 'totalLiabilities', len),
        cash: pickSeries(bs, 'cash', len),
        coreCurrentAssets: pickSeries(bs, 'coreCurrentAssets', len),
        longTermCoreAssets: pickSeries(bs, 'longTermCoreAssets', len),
        longTermInvestments: pickSeries(bs, 'longTermInvestments', len),
        riskAssets: pickSeries(bs, 'riskAssets', len),
        shortDebt: pickSeries(bs, 'shortDebt', len),
        longDebt: pickSeries(bs, 'longDebt', len),
        operatingCurrentLiabilities: pickSeries(bs, 'operatingCurrentLiabilities', len),
        operatingCashFlow: pickSeries(cf, 'operatingCashFlow', len),
        investingCashFlow: pickSeries(cf, 'investingCashFlow', len),
        financingCashFlow: pickSeries(cf, 'financingCashFlow', len),
        capex: pickSeries(cf, 'capex', len),
        cfNetProfit: pickSeries(cf, 'netProfit', len),
        cashFromCustomers: pickSeries(cf, 'cashFromCustomers', len),
        dividendInterestCashOut: pickSeries(cf, 'dividendInterestCashOut', len)
      }};
      const requiredFields = [
        [pl, 'revenue'], [pl, 'cogs'], [pl, 'operatingProfit'], [pl, 'netProfit'], [pl, 'parentNetProfit'], [pl, 'adjustedNetProfit'],
        [bs, 'totalAssets'], [bs, 'totalEquity'], [bs, 'totalLiabilities'], [bs, 'cash'], [bs, 'shortDebt'], [bs, 'longDebt'],
        [cf, 'operatingCashFlow'], [cf, 'capex'], [cf, 'cashFromCustomers']
      ];
      const missing = requiredFields.filter(([table, key]) => !hasField(table, key)).map(([, key]) => (FIELD_ALIASES[key] || [key])[0]);
      const ctx = {{ years, tables: {{ pl, bs, cf }}, metrics, missing }};
      calculateGrowthMetrics(ctx);
      calculateProfitabilityMetrics(ctx);
      calculateBalanceSheetMetrics(ctx);
      calculateCashFlowMetrics(ctx);
      calculateDuPontMetrics(ctx);
      return ctx;
    }}

    function statusFor(metric, value) {{
      if (value === null || value === undefined || Number.isNaN(value)) return ['数据不足', 'warn'];
      if (['Revenue YoY', 'ROE', 'OCF / Net Profit', '现金短债比'].includes(metric)) return value >= 1 || (metric !== '现金短债比' && value > 0.1) ? ['良好', 'good'] : ['承压', 'warn'];
      if (metric === '资产负债率') return value > 0.7 ? ['风险较高', 'risk'] : value > 0.55 ? ['承压', 'warn'] : ['良好', 'good'];
      if (metric.includes('Margin') || metric.includes('率')) return value > 0 ? ['改善', 'good'] : ['承压', 'warn'];
      return ['良好', 'good'];
    }}

    function renderMetricCards(ctx) {{
      const m = ctx.metrics;
      const specs = [
        ['Revenue', m.revenue, money], ['Revenue YoY', m.revenueYoY, percent],
        ['Parent Net Profit', m.parentNetProfit, money], ['Adjusted Net Profit', m.adjustedNetProfit, money],
        ['Gross Margin', m.grossMargin, percent], ['Net Margin', m.netMargin, percent],
        ['ROE', m.roe, percent], ['Operating Cash Flow', m.operatingCashFlow, money],
        ['Free Cash Flow', m.freeCashFlow, money], ['资产负债率', m.debtToAssetRatio, percent],
        ['现金短债比', m.cashToShortDebt, percent]
      ];
      document.getElementById('analysisCards').innerHTML = specs.map(([label, arr, fmt]) => {{
        const idx = latestIndex(arr), value = idx >= 0 ? arr[idx] : null;
        const prev = idx > 0 ? arr[idx - 1] : null;
        const change = value !== null && prev !== null && prev ? value / prev - 1 : null;
        const [status, cls] = statusFor(label, value);
        return `<article class="analysis-card"><span>${{label}}</span><strong>${{value === null ? '数据不足' : fmt(value)}}</strong><small>同比变化：${{change === null ? '数据不足' : percent(change)}}</small><em class="status-pill ${{cls}}">${{status}}</em></article>`;
      }}).join('');
    }}

    function metricChart(id, titleYears, series, mode = 'money') {{
      const isRate = mode === 'rate';
      const formatter = value => (isRate || mode === 'ratio') ? percent(value) : money(value);
      initChart(id, {{
        color: ['#2563eb', '#0f766e', '#f97316', '#7c3aed', '#b91c1c', '#64748b'],
        tooltip: {{ trigger: 'axis', valueFormatter: formatter }},
        legend: {{ top: 0, type: 'scroll', textStyle: {{ color: '#687385' }} }},
        grid: {{ left: 76, right: 30, top: 48, bottom: 42 }},
        xAxis: {{ type: 'category', data: titleYears, axisLabel: {{ color: '#687385' }} }},
        yAxis: {{ type: 'value', axisLabel: {{ color: '#687385', formatter }}, splitLine: {{ lineStyle: {{ color: '#e8edf5' }} }} }},
        series: series.map(s => ({{ ...s, smooth: s.type !== 'bar', symbolSize: 6, barMaxWidth: 36 }}))
      }});
    }}

    function dualAxisChart(id, years, amountSeries, rateSeries) {{
      const hasRightAxis = Array.isArray(rateSeries) && rateSeries.length > 0;
      initChart(id, {{
        color: ['#2563eb', '#0f766e', '#f97316', '#7c3aed'],
        tooltip: {{ trigger: 'axis', valueFormatter: value => hasRightAxis && Math.abs(value || 0) <= 1 ? percent(value) : money(value) }},
        legend: {{ top: 0, type: 'scroll', textStyle: {{ color: '#687385' }} }},
        grid: {{ left: 76, right: hasRightAxis ? 72 : 30, top: 48, bottom: 42 }},
        xAxis: {{ type: 'category', data: years, axisLabel: {{ color: '#687385' }} }},
        yAxis: hasRightAxis ? [
          {{ type: 'value', name:'金额', axisLabel: {{ color: '#687385', formatter: money }}, splitLine: {{ lineStyle: {{ color: '#e8edf5' }} }} }},
          {{ type: 'value', axisLabel: {{ color: '#687385', formatter: percent }}, splitLine: {{ show: false }} }}
        ] : {{ type: 'value', name:'金额', axisLabel: {{ color: '#687385', formatter: money }}, splitLine: {{ lineStyle: {{ color: '#e8edf5' }} }} }},
        series: [
          ...amountSeries.map(s => ({{ ...s, yAxisIndex: 0, barMaxWidth: 34, smooth: true }})),
          ...rateSeries.map(s => ({{ ...s, yAxisIndex: 1, smooth: true, symbolSize: 7 }}))
        ]
      }});
    }}

    function generateRiskWarnings(ctx) {{
      const m = ctx.metrics, risks = [];
      if (lastN(m.ocfToNetProfit, 2).every(x => x !== null && x < 1)) risks.push('利润现金含量偏弱：OCF / Net Profit 连续两年低于 1。');
      if (lastN(m.freeCashFlow, 2).every(x => x !== null && x < 0)) risks.push('自由现金流承压：Free Cash Flow 连续两年为负。');
      if (isRising(m.debtToAssetRatio)) risks.push('资产负债率上升：扩张过程中杠杆水平抬升。');
      if ((latest(m.cashToShortDebt) ?? 999) < 1) risks.push('短期偿债压力较高：现金短债比低于 1。');
      const i = m.revenue.length - 1;
      if (i > 0 && m.revenue[i] > m.revenue[i - 1] && m.netProfit[i] < m.netProfit[i - 1]) risks.push('增收不增利：收入增长但净利润下滑。');
      if (isFalling(m.grossMargin)) risks.push('毛利率承压：Gross Margin 连续下降。');
      if (isRising(m.financialExpenseRatio)) risks.push('融资成本压力增加：财务费用率持续上升。');
      return risks.length ? risks : ['暂无明显财务风险信号，核心指标处于相对可控区间。'];
    }}

    function generateTextAnalysis(ctx) {{
      const m = ctx.metrics;
      const rev = latest(m.revenueYoY), profit = latest(m.parentNetProfitYoY), adj = latest(m.adjustedNetProfitYoY);
      document.getElementById('growthAnalysisText').textContent = rev === null ? '收入或利润历史数据不足，暂无法形成成长能力判断。' : `最近一期收入增速为 ${{percent(rev)}}，归母净利润增速为 ${{profit === null ? '数据不足' : percent(profit)}}，扣非净利润增速为 ${{adj === null ? '数据不足' : percent(adj)}}。${{profit !== null && profit < rev ? '利润增速低于收入增速，盈利弹性需要关注。' : '利润表现与收入增长基本匹配。'}}`;
      document.getElementById('profitabilityAnalysisText').textContent = `最近一期毛利率 ${{latest(m.grossMargin) === null ? '数据不足' : percent(latest(m.grossMargin))}}，净利率 ${{latest(m.netMargin) === null ? '数据不足' : percent(latest(m.netMargin))}}，ROE ${{latest(m.roe) === null ? '数据不足' : percent(latest(m.roe))}}。${{isFalling(m.grossMargin) ? '毛利率连续走弱，需关注成本压力。' : '毛利率未出现连续恶化信号。'}}`;
      document.getElementById('cashFlowAnalysisText').textContent = `最近一期 OCF / Net Profit 为 ${{latest(m.ocfToNetProfit) === null ? '数据不足' : percent(latest(m.ocfToNetProfit))}}，自由现金流为 ${{latest(m.freeCashFlow) === null ? '数据不足' : money(latest(m.freeCashFlow))}}。${{latest(m.freeCashFlow) !== null && latest(m.freeCashFlow) < 0 ? '资本开支或营运资金占用对现金流形成压力。' : '经营现金流对利润质量形成一定支撑。'}}`;
      document.getElementById('solvencyAnalysisText').textContent = `最近一期资产负债率 ${{latest(m.debtToAssetRatio) === null ? '数据不足' : percent(latest(m.debtToAssetRatio))}}，现金短债比 ${{latest(m.cashToShortDebt) === null ? '数据不足' : percent(latest(m.cashToShortDebt))}}。${{latest(m.cashToShortDebt) !== null && latest(m.cashToShortDebt) < 1 ? '短期债务覆盖偏弱。' : '短期流动性覆盖尚可。'}}`;
      document.getElementById('threeStatementAnalysisText').textContent = `三表联动显示，收入、利润、经营现金流、自由现金流与资产负债扩张需要共同观察：若收入增长同时现金流改善且有息负债未明显膨胀，增长质量更高；反之需关注扩张消耗现金和杠杆抬升。`;
    }}

    function renderValuationInputs(ctx) {{
      const m = ctx.metrics;
      const rows = [
        ['最近一期收入', latest(m.revenue), money],
        ['归母净利润', latest(m.parentNetProfit), money],
        ['扣非净利润', latest(m.adjustedNetProfit), money],
        ['经营现金流', latest(m.operatingCashFlow), money],
        ['自由现金流', latest(m.freeCashFlow), money],
        ['ROE', latest(m.roe), percent],
        ['资产负债率', latest(m.debtToAssetRatio), percent],
        ['有息负债', latest(m.interestBearingDebt), money],
        ['净债务', latest(m.netDebt), money]
      ];
      document.getElementById('valuationInputTable').innerHTML = tableHtml(['指标', '数值'], rows.map(([label, value, fmt]) => [label, value === null || value === undefined ? '数据不足' : fmt(value)]));
    }}

    function tableHtml(headers, rows) {{
      return `<table><thead><tr>${{headers.map(h => `<th>${{h}}</th>`).join('')}}</tr></thead><tbody>${{rows.map(row => `<tr>${{row.map(c => `<td>${{c}}</td>`).join('')}}</tr>`).join('')}}</tbody></table>`;
    }}

    function renderCharts(ctx) {{
      const y = ctx.years, m = ctx.metrics;
      metricChart('analysisRevenueTrend', y, [{{ name:'Revenue', type:'bar', data:m.revenue }}]);
      metricChart('analysisRevenueYoY', y, [{{ name:'Revenue YoY', type:'line', data:m.revenueYoY }}], 'rate');
      metricChart('analysisProfitTrend', y, [{{ name:'Net Profit', type:'bar', data:m.netProfit }}, {{ name:'Adjusted Net Profit', type:'line', data:m.adjustedNetProfit }}]);
      metricChart('analysisProfitGrowth', y, [{{ name:'Gross Profit YoY', type:'line', data:m.grossProfitYoY }}, {{ name:'Operating Profit YoY', type:'line', data:m.operatingProfitYoY }}, {{ name:'Parent Net Profit YoY', type:'line', data:m.parentNetProfitYoY }}, {{ name:'Adjusted Net Profit YoY', type:'line', data:m.adjustedNetProfitYoY }}], 'rate');
      metricChart('analysisMargins', y, [{{ name:'Gross Margin', type:'line', data:m.grossMargin }}, {{ name:'Operating Margin', type:'line', data:m.operatingMargin }}, {{ name:'Net Margin', type:'line', data:m.netMargin }}], 'rate');
      metricChart('analysisReturns', y, [{{ name:'ROE', type:'line', data:m.roe }}, {{ name:'ROA', type:'line', data:m.roa }}], 'rate');
      metricChart('analysisExpenseRatios', y, [{{ name:'Selling Expense Ratio', type:'line', data:m.sellingExpenseRatio }}, {{ name:'Admin Expense Ratio', type:'line', data:m.adminExpenseRatio }}, {{ name:'R&D Expense Ratio', type:'line', data:m.rdExpenseRatio }}, {{ name:'Financial Expense Ratio', type:'line', data:m.financialExpenseRatio }}], 'rate');
      metricChart('analysisAssetStructure', y, [{{ name:'现金及短期金融资产占比', type:'bar', stack:'assets', data:m.cashAssetRatio }}, {{ name:'核心经营性流动资产占比', type:'bar', stack:'assets', data:m.coreCurrentAssetRatio }}, {{ name:'长期核心经营资产占比', type:'bar', stack:'assets', data:m.longTermCoreAssetRatio }}, {{ name:'长期金融及股权投资占比', type:'bar', stack:'assets', data:m.longTermInvestmentRatio }}, {{ name:'风险及摊销类资产占比', type:'bar', stack:'assets', data:m.riskAssetRatio }}], 'rate');
      metricChart('analysisLiabilityStructure', y, [{{ name:'有息负债', type:'bar', stack:'liab', data:m.interestBearingDebt }}, {{ name:'经营性无息流动负债', type:'bar', stack:'liab', data:m.operatingCurrentLiabilities }}]);
      metricChart('analysisDebtRatios', y, [{{ name:'Debt to Asset Ratio', type:'line', data:m.debtToAssetRatio }}, {{ name:'Interest-bearing Debt Ratio', type:'line', data:m.interestBearingDebtRatio }}, {{ name:'Net Debt to Equity', type:'line', data:m.netDebtToEquity }}], 'rate');
      metricChart('analysisDebtTrend', y, [{{ name:'Interest-bearing Debt', type:'bar', data:m.interestBearingDebt }}, {{ name:'Net Debt', type:'line', data:m.netDebt }}]);
      metricChart('analysisCashDebt', y, [{{ name:'Cash to Short-term Debt', type:'line', data:m.cashToShortDebt }}], 'ratio');
      metricChart('analysisOCFProfit', y, [{{ name:'Operating Cash Flow', type:'bar', data:m.operatingCashFlow }}, {{ name:'Net Profit', type:'line', data:m.netProfit }}]);
      metricChart('analysisFCFTrend', y, [{{ name:'Free Cash Flow', type:'bar', data:m.freeCashFlow }}]);
      metricChart('analysisCashQualityRatios', y, [{{ name:'OCF / Net Profit', type:'line', data:m.ocfToNetProfit }}, {{ name:'FCF Margin', type:'line', data:m.fcfMargin }}, {{ name:'Cash Collection Ratio', type:'line', data:m.cashCollectionRatio }}], 'rate');
      metricChart('analysisCapexRevenue', y, [{{ name:'Capex / Revenue', type:'line', data:m.capexRevenue }}], 'rate');
      dualAxisChart('analysisThreeStatement', y, [
        {{ name:'营业收入', type:'bar', data:m.revenue }},
        {{ name:'净利润', type:'line', data:m.netProfit }},
        {{ name:'经营现金流', type:'line', data:m.operatingCashFlow }},
        {{ name:'自由现金流', type:'line', data:m.freeCashFlow }},
        {{ name:'总资产', type:'bar', data:m.totalAssets }},
        {{ name:'有息负债', type:'bar', data:m.interestBearingDebt }}
      ], []);
      metricChart('analysisDupont', y, [{{ name:'ROE', type:'line', data:m.dupontRoe }}, {{ name:'Net Margin', type:'line', data:m.dupontNetMargin }}, {{ name:'Asset Turnover', type:'line', data:m.assetTurnover }}, {{ name:'Equity Multiplier', type:'line', data:m.equityMultiplier }}], 'rate');
    }}

    async function renderFinancialAnalysisDashboard() {{
      if (analysisState.initialized) {{
        setTimeout(() => Object.values(chartRefs).forEach(c => c && c.resize()), 0);
        return;
      }}
      try {{
        const ctx = await buildAnalysisContext();
        analysisState.initialized = true;
        document.getElementById('analysisError').textContent = ctx.missing.length ? '以下字段缺失，相关指标显示为数据不足：' + ctx.missing.join('、') : '';
        renderMetricCards(ctx);
        renderCharts(ctx);
        generateTextAnalysis(ctx);
        renderValuationInputs(ctx);
        const risks = generateRiskWarnings(ctx);
        document.getElementById('riskWarnings').innerHTML = risks.map((risk, idx) => `<li class="${{idx === 0 && risks.length === 1 && risk.startsWith('暂无') ? 'good' : ''}}">${{risk}}</li>`).join('');
      }} catch (err) {{
        document.getElementById('analysisError').textContent = '财务分析面板加载失败：' + err.message;
      }}
    }}

    function statementChartOption(statement, chartConfig) {{
      const palette = ['#2563eb', '#0f766e', '#f97316', '#7c3aed', '#b91c1c', '#64748b'];
      if (chartConfig.customType === 'net_debt_ratio_debt_growth') {{
        const cash = statement.series['Cash & Short-term Financial Assets'] || [];
        const shortDebt = statement.series['Interest-bearing Short-term Debt'] || [];
        const longDebt = statement.series['Long-term Interest-bearing Debt'] || [];
        const equity = statement.series['Total Equity'] || [];
        const interestDebt = statement.years.map((_, i) => (shortDebt[i] ?? 0) + (longDebt[i] ?? 0));
        const netDebtRatio = statement.years.map((_, i) => equity[i] ? (interestDebt[i] - (cash[i] ?? 0)) / equity[i] : null);
        const debtGrowth = interestDebt.map((value, i) => i > 0 && interestDebt[i - 1] ? value / interestDebt[i - 1] - 1 : null);
        return {{
          color: ['#2563eb', '#f97316'],
          tooltip: {{ trigger: 'axis', valueFormatter: percent }},
          legend: {{ top: 0, textStyle: {{ color: '#687385' }} }},
          grid: {{ left: 76, right: 72, top: 48, bottom: 42 }},
          xAxis: {{ type: 'category', data: statement.years, axisLabel: {{ color: '#687385' }} }},
          yAxis: [
            {{ type: 'value', name: '净负债率', axisLabel: {{ color: '#687385', formatter: percent }}, splitLine: {{ lineStyle: {{ color: '#e8edf5' }} }} }},
            {{ type: 'value', name: '有息负债增速', axisLabel: {{ color: '#687385', formatter: percent }}, splitLine: {{ show: false }} }}
          ],
          series: [
            {{ name:'净负债率', type:'line', yAxisIndex:0, data:netDebtRatio, smooth:true, symbolSize:7 }},
            {{ name:'有息负债增速', type:'line', yAxisIndex:1, data:debtGrowth, smooth:true, symbolSize:7 }}
          ]
        }};
      }}
      const series = chartConfig.items
        .filter(name => Array.isArray(statement.series[name]))
        .map(name => {{
          const item = {{
            name,
            type: (chartConfig.types && chartConfig.types[name]) || 'line',
            data: statement.series[name],
            smooth: true,
            symbolSize: 6,
            barMaxWidth: 34,
            barGap: chartConfig.barMode === 'grouped' ? '12%' : undefined,
            barCategoryGap: chartConfig.barMode === 'grouped' ? '34%' : undefined,
            stack: chartConfig.stack || undefined
          }};
          if (chartConfig.yAxisIndex && chartConfig.yAxisIndex[name] !== undefined) {{
            item.yAxisIndex = chartConfig.yAxisIndex[name];
          }}
          return item;
        }});
      if (Array.isArray(chartConfig.derived)) {{
        chartConfig.derived.forEach(item => {{
          const numerator = statement.series[item.numerator] || [];
          const denominator = statement.series[item.denominator] || [];
          const data = denominator.map((value, i) => value ? numerator[i] / value : 0);
          series.push({{
            name: item.name,
            type: item.chartType || 'line',
            yAxisIndex: item.yAxisIndex || 0,
            data,
            smooth: true,
            symbolSize: 7
          }});
        }});
      }}
      const yAxis = chartConfig.dualAxis ? [
        {{ type: 'value', name: '金额', axisLabel: {{ color: '#687385', formatter: money }}, splitLine: {{ lineStyle: {{ color: '#e8edf5' }} }} }},
        {{ type: 'value', name: chartConfig.rightAxisName || '资产负债率', axisLabel: {{ color: '#687385', formatter: chartConfig.rightAxisName ? money : percent }}, splitLine: {{ show: false }} }}
      ] : {{ type: 'value', axisLabel: {{ color: '#687385', formatter: money }}, splitLine: {{ lineStyle: {{ color: '#e8edf5' }} }} }};
      return {{
        color: palette,
        tooltip: {{
          trigger: 'axis',
          valueFormatter: (value) => typeof value === 'number' && Math.abs(value) <= 1 ? percent(value) : money(value)
        }},
        legend: {{ top: 0, type: 'scroll', textStyle: {{ color: '#687385' }} }},
        grid: {{ left: 76, right: chartConfig.dualAxis ? 72 : 26, top: 48, bottom: 42 }},
        xAxis: {{ type: 'category', data: statement.years, axisLabel: {{ color: '#687385' }} }},
        yAxis,
        series
      }};
    }}

    function renderFinancialCharts() {{
      initChart('revenueTrendChart', chartBase(model.historical.years, [
        {{ name:'Revenue', type:'line', data:model.historical.revenue, smooth:true, symbolSize:7 }},
        {{ name:'EBIT', type:'line', data:model.historical.ebit, smooth:true, symbolSize:7 }},
        {{ name:'Parent Net Profit', type:'line', data:model.historical.netProfit, smooth:true, symbolSize:7 }}
      ]));
      initChart('cashFlowChart', chartBase(model.historical.years, [
        {{ name:'CFO', type:'bar', data:model.historical.cfo, barMaxWidth:38 }},
        {{ name:'FCFF Proxy', type:'line', data:model.historical.fcffProxy, smooth:true, symbolSize:7 }}
      ]));
    }}

    function renderStatementCharts(tabId) {{
      const tabToStatement = {{
        balanceSheet: 'balance_sheet',
        incomeStatement: 'income_statement',
        cashFlow: 'cash_flow'
      }};
      const key = tabToStatement[tabId];
      if (!key || !model.statements[key]) return;
      const statement = model.statements[key];
      statement.charts.forEach(chartConfig => {{
        initChart(chartConfig.id, statementChartOption(statement, chartConfig));
      }});
    }}

    function renderValuation() {{
      const assumptions = collectAssumptions();
      const dcf = compute(assumptions);
      document.getElementById('valuationPrice').textContent = price(dcf.intrinsicPrice);
      document.getElementById('valuationUpside').textContent = percent(dcf.upside);
      document.getElementById('valuationEv').textContent = money(dcf.enterpriseValue);
      document.getElementById('valuationMos').textContent = percent(dcf.safetyMargin);
      document.getElementById('headerPrice').textContent = price(dcf.intrinsicPrice);
      document.getElementById('headerUpside').textContent = percent(dcf.upside);
      const badge = document.getElementById('headerBadge');
      badge.className = 'badge ' + (dcf.upside >= 0.15 ? 'good' : dcf.upside >= -0.10 ? 'neutral' : 'risk');

      const bridge = [
        ['5年 FCFF 现值合计', money(dcf.pvFcff.reduce((x,y)=>x+y,0))],
        ['终值现值', money(dcf.pvTerminal)],
        ['企业价值 EV', money(dcf.enterpriseValue)],
        ['加：净现金 / 减：净债务', money(assumptions.net_cash)],
        ['减：少数股东权益', money(-assumptions.minority_interest)],
        ['加：长期金融及股权投资', money(assumptions.long_term_investments)],
        ['加：非经营流动资产', money(assumptions.non_op_current_assets)],
        ['股东权益价值', money(dcf.equityValue)],
        ['每股内在价值', price(dcf.intrinsicPrice)]
      ];
      document.getElementById('bridgeBody').innerHTML = bridge.map(r => `<tr><td>${{r[0]}}</td><td>${{r[1]}}</td></tr>`).join('');

      document.getElementById('forecastBody').innerHTML = dcf.rows.map((row, i) => `
        <tr>
          <td>${{row.year}}</td><td>${{money(row.revenue)}}</td><td>${{percent(row.growth)}}</td>
          <td>${{percent(row.ebitMargin)}}</td><td>${{money(row.nopat)}}</td><td>${{money(row.da)}}</td>
          <td>${{money(row.capex)}}</td><td>${{money(row.changeNwc)}}</td><td>${{money(row.fcff)}}</td><td>${{money(dcf.pvFcff[i])}}</td>
        </tr>`).join('');

      initChart('forecastChart', forecastComboChart(model.forecastYears, dcf.rows));
    }}

    document.querySelectorAll('.tab-btn').forEach(btn => {{
      btn.addEventListener('click', () => {{
        document.querySelectorAll('.tab-btn').forEach(x => x.classList.remove('active'));
        document.querySelectorAll('.tab-pane').forEach(x => x.classList.remove('active'));
        btn.classList.add('active');
        document.getElementById(btn.dataset.tab).classList.add('active');
        if (btn.dataset.tab === 'analysis') renderFinancialAnalysisDashboard();
        renderStatementCharts(btn.dataset.tab);
        setTimeout(() => Object.values(chartRefs).forEach(c => c && c.resize()), 0);
      }});
    }});
    document.querySelectorAll('input').forEach(input => input.addEventListener('input', renderValuation));
    document.getElementById('resetAssumptions').addEventListener('click', () => {{
      document.getElementById('currentPrice').value = initialAssumptions.current_price;
      document.getElementById('sharesOutstanding').value = initialAssumptions.shares_outstanding;
      document.getElementById('wacc').value = (initialAssumptions.wacc * 100).toFixed(2);
      document.getElementById('terminalGrowth').value = (initialAssumptions.terminal_growth * 100).toFixed(2);
      document.getElementById('netCash').value = initialAssumptions.net_cash;
      document.getElementById('minorityInterest').value = initialAssumptions.minority_interest;
      document.getElementById('longTermInvestments').value = initialAssumptions.long_term_investments;
      document.getElementById('nonOpCurrentAssets').value = initialAssumptions.non_op_current_assets;
      document.querySelectorAll('.assumption-input').forEach(input => {{
        input.value = (initialAssumptions[input.dataset.array][Number(input.dataset.index)] * 100).toFixed(2);
      }});
      renderValuation();
    }});
    document.getElementById('copyAssumptions').addEventListener('click', () => {{
      initialAssumptions = JSON.parse(JSON.stringify(collectAssumptions()));
      renderValuation();
      alert('已使用当前预测数据作为本页面新的基准假设');
    }});
    window.addEventListener('resize', () => Object.values(chartRefs).forEach(c => c && c.resize()));
    renderFinancialCharts();
    renderValuation();
  </script>
</body>
</html>
"""


def main() -> None:
    args = parse_args()
    data_dir = configure_paths(args.data_dir_flag or args.data_dir)
    from afda.pipeline_utils import ASSETS_DIR, RESULTS_DIR, VALUATION_DIR

    data = load_dataset(data_dir)
    workbook_path = VALUATION_DIR / "DCF_valuation_model.xlsx"
    assumptions = read_assumptions(data, workbook_path)
    dcf = compute_dcf(data, assumptions)
    statements = load_statement_tables(RESULTS_DIR)
    echarts_src = ensure_echarts_asset(ASSETS_DIR)
    output_path = RESULTS_DIR / OUTPUT_FILE_NAME
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(build_html(data, assumptions, dcf, echarts_src, statements), encoding="utf-8")
    print(f"HTML dashboard generated: {output_path}")


if __name__ == "__main__":
    main()
