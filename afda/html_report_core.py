"""Support code for the offline HTML dashboard generator."""

from __future__ import annotations

import csv
import html
import math
import shutil
from pathlib import Path
from typing import Dict, Iterable

from openpyxl import load_workbook

OUTPUT_FILE_NAME = "financial_dcf_dashboard.html"
LOCAL_ECHARTS_SOURCE = Path(__file__).resolve().parents[1] / "assets" / "echarts.min.js"
STATEMENT_SOURCES = {
    "balance_sheet": ("资产负债表", "balance_sheet", "2_standardized_bs_wide.csv"),
    "income_statement": ("利润表", "income_statement", "2_standardized_pl_wide.csv"),
    "cash_flow": ("现金流量表", "cash_flow", "2_standardized_cf_wide.csv"),
}
BALANCE_SHEET_EMPHASIS_ROWS = {
    "Total Current Assets",
    "Total Non-current Assets",
    "Total Assets",
    "Total Current Liabilities",
    "Total Non-current Liabilities",
    "Total Liabilities",
    "Total Equity",
}
STATEMENT_CHARTS = {
    "balance_sheet": [
        {
            "id": "balanceSheetScaleChart",
            "title": "资产、负债与资产负债率",
            "items": ["Total Assets", "Total Liabilities"],
            "types": {"Total Assets": "bar", "Total Liabilities": "bar"},
            "barMode": "grouped",
            "derived": [
                {
                    "name": "资产负债率",
                    "type": "debt_to_assets",
                    "numerator": "Total Liabilities",
                    "denominator": "Total Assets",
                    "chartType": "line",
                    "yAxisIndex": 1,
                }
            ],
            "dualAxis": True,
        },
        {
            "id": "balanceSheetStructureChart",
            "title": "流动与非流动资产结构",
            "items": ["Total Current Assets", "Total Non-current Assets"],
            "types": {"Total Current Assets": "bar", "Total Non-current Assets": "bar"},
            "stack": "assets",
        },
        {
            "id": "currentAssetsStructureChart",
            "title": "流动资产结构",
            "items": [
                "Cash & Short-term Financial Assets",
                "Core Operating Current Assets",
                "Non-operating Misc. Current Assets",
            ],
            "types": {
                "Cash & Short-term Financial Assets": "bar",
                "Core Operating Current Assets": "bar",
                "Non-operating Misc. Current Assets": "bar",
            },
            "stack": "current_assets",
        },
        {
            "id": "netDebtRatioDebtGrowthChart",
            "title": "净负债率与有息负债增速",
            "customType": "net_debt_ratio_debt_growth",
            "dualAxis": True,
        },
    ],
    "income_statement": [
        {
            "id": "incomeProfitChart",
            "title": "收入、营业利润与归母净利润",
            "items": ["Revenue", "Operating Profit", "Parent Net Profit"],
            "types": {"Revenue": "bar", "Operating Profit": "line", "Parent Net Profit": "line"},
            "yAxisIndex": {"Revenue": 0, "Operating Profit": 1, "Parent Net Profit": 1},
            "dualAxis": True,
            "rightAxisName": "利润",
        },
        {
            "id": "incomeExpenseChart",
            "title": "成本与期间费用",
            "items": ["COGS", "Selling Expense", "Admin Expense", "R&D Expense", "Financial Expense"],
            "types": {
                "COGS": "bar",
                "Selling Expense": "bar",
                "Admin Expense": "bar",
                "R&D Expense": "bar",
                "Financial Expense": "bar",
            },
            "stack": "expenses",
        },
    ],
    "cash_flow": [
        {
            "id": "cashFlowStatementChart",
            "title": "经营、投资、筹资现金流",
            "items": ["Operating Cash Flow", "Investing Cash Flow", "Financing Cash Flow", "Net Change in Cash"],
            "types": {
                "Operating Cash Flow": "bar",
                "Investing Cash Flow": "bar",
                "Financing Cash Flow": "bar",
                "Net Change in Cash": "line",
            },
        },
        {
            "id": "cashPositionChart",
            "title": "期初与期末现金余额",
            "items": ["Beginning Cash", "Ending Cash"],
            "types": {"Beginning Cash": "line", "Ending Cash": "line"},
        },
    ],
}


def as_float(value: object, fallback: float = 0.0) -> float:
    try:
        if value is None:
            return fallback
        out = float(value)
        if math.isnan(out) or math.isinf(out):
            return fallback
        return out
    except (TypeError, ValueError):
        return fallback


def read_assumptions(data: Dict[str, object], workbook_path: Path) -> Dict[str, object]:
    config = data["valuation_config"]
    assumptions = {
        "shares_outstanding": float(data["shares_outstanding"]),
        "current_price": float(data["current_price"]),
        "net_cash": float(data["cash"][-1] - data["short_debt"][-1] - data["long_debt"][-1]),
        "wacc": float(config["dcf"]["wacc"]),
        "terminal_growth": float(config["dcf"]["terminal_growth"]),
        "minority_interest": float(data["minority_interest"][-1]),
        "long_term_investments": float(data["long_term_investments"][-1]),
        "non_op_current_assets": float(data["non_op_current_assets"][-1]),
        "growths": [float(x) for x in data["default_growths"]],
        "ebit_margins": [float(data["base_ebit_margin"])] * 5,
        "tax_rates": [float(data["base_tax_rate"])] * 5,
        "da_ratios": [float(data["base_da_ratio"])] * 5,
        "capex_ratios": [float(data["base_capex_ratio"])] * 5,
        "nwc_ratios": [float(data["base_nwc_ratio"])] * 5,
    }
    if not workbook_path.exists():
        return assumptions

    wb = load_workbook(workbook_path, data_only=False, read_only=True)
    if "Assumptions" not in wb.sheetnames:
        return assumptions
    ws = wb["Assumptions"]
    for key, cell in {
        "shares_outstanding": "B3",
        "current_price": "B4",
        "net_cash": "B5",
        "wacc": "B6",
        "terminal_growth": "B7",
        "minority_interest": "B8",
        "long_term_investments": "B9",
        "non_op_current_assets": "B10",
    }.items():
        assumptions[key] = as_float(ws[cell].value, float(assumptions[key]))

    for key, row in {
        "growths": 14,
        "ebit_margins": 15,
        "tax_rates": 16,
        "da_ratios": 17,
        "capex_ratios": 18,
        "nwc_ratios": 19,
    }.items():
        assumptions[key] = [
            as_float(ws.cell(row=row, column=col).value, float(assumptions[key][col - 2]))
            for col in range(2, 7)
        ]
    return assumptions


def compute_dcf(data: Dict[str, object], assumptions: Dict[str, object]) -> Dict[str, object]:
    forecast_rows = []
    prev_revenue = float(data["revenue"][-1])
    prev_nwc = float(data["nwc"][-1])
    for idx, year in enumerate(data["forecast_years"]):
        revenue = prev_revenue * (1 + float(assumptions["growths"][idx]))
        ebit = revenue * float(assumptions["ebit_margins"][idx])
        tax_on_ebit = ebit * float(assumptions["tax_rates"][idx])
        nopat = ebit - tax_on_ebit
        da = revenue * float(assumptions["da_ratios"][idx])
        capex = revenue * float(assumptions["capex_ratios"][idx])
        operating_nwc = revenue * float(assumptions["nwc_ratios"][idx])
        change_nwc = operating_nwc - prev_nwc
        fcff = nopat + da - capex - change_nwc
        forecast_rows.append(
            {
                "year": int(year),
                "revenue": revenue,
                "growth": float(assumptions["growths"][idx]),
                "ebit_margin": float(assumptions["ebit_margins"][idx]),
                "tax_rate": float(assumptions["tax_rates"][idx]),
                "da_ratio": float(assumptions["da_ratios"][idx]),
                "capex_ratio": float(assumptions["capex_ratios"][idx]),
                "nwc_ratio": float(assumptions["nwc_ratios"][idx]),
                "ebit": ebit,
                "nopat": nopat,
                "da": da,
                "capex": capex,
                "operating_nwc": operating_nwc,
                "change_nwc": change_nwc,
                "fcff": fcff,
            }
        )
        prev_revenue = revenue
        prev_nwc = operating_nwc

    wacc = float(assumptions["wacc"])
    terminal_growth = float(assumptions["terminal_growth"])
    if wacc <= terminal_growth:
        terminal_growth = max(wacc - 0.005, 0.0)
    terminal_value = forecast_rows[-1]["fcff"] * (1 + terminal_growth) / (wacc - terminal_growth)
    pv_fcff = [row["fcff"] / ((1 + wacc) ** (idx + 1)) for idx, row in enumerate(forecast_rows)]
    pv_terminal = terminal_value / ((1 + wacc) ** len(forecast_rows))
    enterprise_value = sum(pv_fcff) + pv_terminal
    equity_value = (
        enterprise_value
        + float(assumptions["net_cash"])
        - float(assumptions["minority_interest"])
        + float(assumptions["long_term_investments"])
        + float(assumptions["non_op_current_assets"])
    )
    intrinsic_price = equity_value / float(assumptions["shares_outstanding"]) if assumptions["shares_outstanding"] else 0.0
    current_price = float(assumptions["current_price"])
    upside = intrinsic_price / current_price - 1 if current_price else 0.0
    safety_margin = 1 - current_price / intrinsic_price if intrinsic_price > 0 else 0.0
    return {
        "forecast_rows": forecast_rows,
        "pv_fcff": pv_fcff,
        "pv_terminal": pv_terminal,
        "terminal_value": terminal_value,
        "enterprise_value": enterprise_value,
        "equity_value": equity_value,
        "intrinsic_price": intrinsic_price,
        "upside": upside,
        "safety_margin": safety_margin,
    }


def ensure_echarts_asset(assets_dir: Path) -> str:
    assets_dir.mkdir(parents=True, exist_ok=True)
    target = assets_dir / "echarts.min.js"
    if not LOCAL_ECHARTS_SOURCE.exists():
        raise FileNotFoundError(f"Offline ECharts asset not found: {LOCAL_ECHARTS_SOURCE}")
    shutil.copy2(LOCAL_ECHARTS_SOURCE, target)
    return "_assets/echarts.min.js"


def money(value: float) -> str:
    abs_value = abs(value)
    if abs_value >= 1e8:
        return f"{value / 1e8:,.2f} 亿"
    if abs_value >= 1e4:
        return f"{value / 1e4:,.2f} 万"
    return f"{value:,.2f}"


def percent(value: float) -> str:
    return f"{value * 100:,.1f}%"


def esc(value: object) -> str:
    return html.escape(str(value), quote=True)


def statement_value(value: str) -> str:
    text = (value or "").strip()
    if not text:
        return ""
    try:
        return money(float(text))
    except ValueError:
        return esc(text)


def load_statement_tables(results_dir: Path) -> Dict[str, Dict[str, object]]:
    statements: Dict[str, Dict[str, object]] = {}
    for key, (title, folder, filename) in STATEMENT_SOURCES.items():
        path = results_dir / "04_rebuilt_statements" / folder / filename
        rel_path = f"04_rebuilt_statements/{folder}/{filename}"
        if not path.exists():
            statements[key] = {
                "title": title,
                "csvPath": rel_path,
                "csvText": "",
                "years": [],
                "series": {},
                "charts": STATEMENT_CHARTS[key],
                "headers": ["提示"],
                "rows": [[esc(f"未找到文件：{path}")]],
            }
            continue

        csv_text = path.read_text(encoding="utf-8-sig")
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.reader(fh)
            headers = next(reader, [])
            years = headers[1:]
            series = {}
            rows = []
            for raw_row in reader:
                if not raw_row or not any(cell.strip() for cell in raw_row):
                    continue
                item = esc(raw_row[0]) if raw_row else ""
                item_name = raw_row[0].strip()
                series[item_name] = [as_float(cell) for cell in raw_row[1:]]
                values = [statement_value(cell) for cell in raw_row[1:]]
                row_class = "emphasis-row" if key == "balance_sheet" and item_name in BALANCE_SHEET_EMPHASIS_ROWS else ""
                rows.append({"cells": [item, *values], "class": row_class})
        statements[key] = {
            "title": title,
            "csvPath": rel_path,
            "csvText": csv_text,
            "years": years,
            "series": series,
            "charts": STATEMENT_CHARTS[key],
            "headers": headers,
            "rows": rows,
        }
    return statements


def table(headers: Iterable[str], rows: Iterable[Iterable[str]]) -> str:
    head = "".join(f"<th>{esc(h)}</th>" for h in headers)
    body = ""
    for row in rows:
        if isinstance(row, dict):
            cells = row.get("cells", [])
            row_class = f' class="{esc(row.get("class", ""))}"' if row.get("class") else ""
        else:
            cells = row
            row_class = ""
        body += f"<tr{row_class}>" + "".join(f"<td>{cell}</td>" for cell in cells) + "</tr>"
    return f"<table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>"


def statement_section(statement: Dict[str, object]) -> str:
    charts = "".join(chart_container(chart["id"], chart["title"]) for chart in statement["charts"])
    return f"""
      <section>
        <h2>{esc(statement["title"])}可视化</h2>
        <div class="grid charts statement-charts">
          {charts}
        </div>
      </section>
      <section>
        <h2>{esc(statement["title"])}</h2>
        <div class="statement-table">
          {table(statement["headers"], statement["rows"])}
        </div>
      </section>
    """


def chart_container(chart_id: str, title: str) -> str:
    return f"""
    <div class="chart-card">
      <div class="chart-title">{esc(title)}</div>
      <div id="{esc(chart_id)}" class="echart"></div>
    </div>
    """


def build_assumption_inputs(years: list[int], assumptions: Dict[str, object]) -> str:
    rows = [
        ("growths", "Revenue Growth"),
        ("ebit_margins", "EBIT Margin"),
        ("tax_rates", "Tax Rate"),
        ("da_ratios", "D&A / Revenue"),
        ("capex_ratios", "Capex / Revenue"),
        ("nwc_ratios", "NWC / Revenue"),
    ]
    body = []
    for key, label in rows:
        cells = [f"<td>{esc(label)}</td>"]
        for idx, value in enumerate(assumptions[key]):
            cells.append(
                f'<td><input class="assumption-input" data-array="{key}" data-index="{idx}" '
                f'type="number" step="0.1" value="{float(value) * 100:.2f}"><span class="unit">%</span></td>'
            )
        body.append("<tr>" + "".join(cells) + "</tr>")
    header = "".join(f"<th>{year}E</th>" for year in years)
    return f"""
    <table class="input-table">
      <thead><tr><th>Assumption</th>{header}</tr></thead>
      <tbody>{''.join(body)}</tbody>
    </table>
    """



