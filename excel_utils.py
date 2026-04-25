"""
excel_utils — Excel 工作簿公共样式工具
========================================
提供对 openpyxl Workbook 的字体后处理功能，被管道中各步骤脚本调用。

主要功能
--------
apply_bilingual_fonts(wb, english_font, chinese_font, size)
    遍历工作簿所有 Sheet 的所有非空单元格，根据单元格内容
    自动选择字体：
    - 含中文字符（CJK Unified Ideographs）→ 黑体（SimHei）
    - 其余（英文、数字、符号）→ Calibri
    保留单元格原有的 bold / italic / size / color / underline 等属性。

使用示例
--------
    from openpyxl import load_workbook
    from excel_utils import apply_bilingual_fonts

    wb = load_workbook("output.xlsx")
    apply_bilingual_fonts(wb)
    wb.save("output.xlsx")
"""

from __future__ import annotations

import copy
import re
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font

# CJK Unified Ideographs 基本区 + 扩展区 A/B + CJK 兼容、部首等常用范围
_CJK_PATTERN = re.compile(
    r"[\u2e80-\u2eff"    # CJK Radicals Supplement
    r"\u2f00-\u2fdf"     # Kangxi Radicals
    r"\u3000-\u303f"     # CJK Symbols and Punctuation
    r"\u3040-\u309f"     # Hiragana
    r"\u30a0-\u30ff"     # Katakana
    r"\u3100-\u312f"     # Bopomofo
    r"\u3200-\u32ff"     # Enclosed CJK Letters and Months
    r"\u3400-\u4dbf"     # CJK Unified Ideographs Extension A
    r"\u4e00-\u9fff"     # CJK Unified Ideographs
    r"\uf900-\ufaff"     # CJK Compatibility Ideographs
    r"\ufe30-\ufe4f"     # CJK Compatibility Forms
    r"]"
)


def _has_cjk(text: str) -> bool:
    """返回字符串中是否含有任何 CJK 字符。"""
    return bool(_CJK_PATTERN.search(text))


def _clone_font(original: Optional[Font], name: str) -> Font:
    """
    以 original 字体属性为基础克隆出一个新 Font，仅替换 name。

    Parameters
    ----------
    original : Font or None
        单元格的当前字体对象。为 None 时使用默认值。
    name : str
        新字体名称（如 "Calibri" 或 "SimHei"）。

    Returns
    -------
    Font
        属性与 original 相同、仅 name 被替换的新字体对象。
    """
    if original is None:
        return Font(name=name)

    return Font(
        name=name,
        size=original.size,
        bold=original.bold,
        italic=original.italic,
        underline=original.underline,
        strike=original.strike,
        color=copy.copy(original.color) if original.color else None,
        vertAlign=original.vertAlign,
        charset=original.charset,
        scheme=original.scheme,
    )


def apply_bilingual_fonts(
    wb: Workbook,
    english_font: str = "Calibri",
    chinese_font: str = "SimHei",
    default_size: Optional[float] = None,
) -> None:
    """
    对工作簿所有 Sheet 的非空单元格应用双语字体策略：
    - 内容含中文 → chinese_font（默认 SimHei / 黑体）
    - 其余内容 → english_font（默认 Calibri）

    已保留单元格的 bold、italic、size、color、underline 等原有属性，
    仅替换字体名称（name）。

    Parameters
    ----------
    wb : openpyxl.Workbook
        要处理的工作簿对象（就地修改）。
    english_font : str
        英文 / 数字内容使用的字体名称，默认 "Calibri"。
    chinese_font : str
        中文内容使用的字体名称，默认 "SimHei"（黑体）。
    default_size : float or None
        若不为 None，则同时统一设置字号；为 None 时保留原字号。
    """
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                text = str(cell.value)
                target_name = chinese_font if _has_cjk(text) else english_font

                new_font = _clone_font(cell.font, target_name)
                if default_size is not None:
                    new_font = Font(
                        name=new_font.name,
                        size=default_size,
                        bold=new_font.bold,
                        italic=new_font.italic,
                        underline=new_font.underline,
                        strike=new_font.strike,
                        color=new_font.color,
                        vertAlign=new_font.vertAlign,
                        charset=new_font.charset,
                        scheme=new_font.scheme,
                    )
                cell.font = new_font
