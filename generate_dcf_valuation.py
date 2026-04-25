"""
Step 8/8 — DCF 估值模型生成
============================
读取三表标准化输出和公司基础信息，自动计算历史财务驱动因子，
生成包含多个联动 Sheet 的 Excel DCF 估值工作簿。

工作簿结构（共 10 个 Sheet）
------------------------------
Summary          — 估值总览：DCF vs 相对估值、加权目标价、上行/下行空间
Historical       — 历史财务整理（Revenue、EBIT、CapEx、CFO 等）
Assumptions      — 可编辑假设区（WACC、永续增长率、5年逐年收入增速等）
Forecast         — 5 年经营预测与 FCFF 计算（公式联动 Assumptions）
DCF              — DCF 估值主表（EV → 股东权益价值 → 每股内在价值）
Sensitivity      — WACC / 永续增长率 敏感性矩阵（含条件格式）
Comparable       — 相对估值（PE / PB / EV/EBIT / EV/EBITDA）
Investment_Thesis — 投资评级与核心观点（自动联动估值结果）
Charts           — Revenue / EBIT / FCFF 趋势图 + 目标价区间图
Raw_Data         — 模型关键原始口径备查

自动计算的驱动因子
------------------
- 收入增速 Seed：基于最近 3 年 CAGR，夹在 [3%, 20%]
- EBIT Margin / Tax Rate / D&A% / CapEx% / NWC%：3 年均值，夹在合理区间
- 5 年预测增速：从 Seed 开始逐年平缓下台阶

输入
----
- results/BS_rebuilt_output/2_standardized_bs.csv
- results/PL_rebuilt_output/2_standardized_pl.csv
- results/CF_rebuilt_output/2_standardized_cf.csv
- rawdata/Info.csv     （需含：总股本、当前股价、公司简称）

输出
----
- results/valuation_output/DCF_valuation_model.xlsx

运行方式
--------
    python generate_dcf_valuation.py
    # 或通过主管道：
    python run_pipeline.py
"""

import math
import os
from datetime import date
from pathlib import Path
from typing import Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from pipeline_utils import detect_ticker
from excel_utils import apply_bilingual_fonts


BASE_DIR = Path(".")
BS_PATH = BASE_DIR / "results" / "BS_rebuilt_output" / "2_standardized_bs.csv"
PL_PATH = BASE_DIR / "results" / "PL_rebuilt_output" / "2_standardized_pl.csv"
CF_PATH = BASE_DIR / "results" / "CF_rebuilt_output" / "2_standardized_cf.csv"
INFO_PATH = BASE_DIR / "rawdata" / "Info.csv"
OUTPUT_DIR = BASE_DIR / "results" / "valuation_output"
OUTPUT_PATH = OUTPUT_DIR / "DCF_valuation_model.xlsx"


def load_item_series(df: pd.DataFrame, item_col: str, year_col: str, value_col: str) -> Dict[str, Dict[int, float]]:
    out: Dict[str, Dict[int, float]] = {}
    for _, row in df.iterrows():
        item = str(row[item_col])
        year = int(row[year_col])
        value = float(row[value_col])
        out.setdefault(item, {})[year] = value
    return out


def get_series(item_map: Dict[str, Dict[int, float]], item: str, years: List[int]) -> List[float]:
    series = item_map.get(item, {})
    return [float(series.get(year, 0.0)) for year in years]


def cagr(start_value: float, end_value: float, periods: int) -> float:
    if start_value <= 0 or end_value <= 0 or periods <= 0:
        return 0.08
    return (end_value / start_value) ** (1 / periods) - 1


def avg(values: List[float], fallback: float = 0.0) -> float:
    clean = [float(v) for v in values if pd.notna(v)]
    return sum(clean) / len(clean) if clean else fallback


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def ensure_output_dir() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def detect_company_name(ticker: str) -> str:
    info_path = BASE_DIR / "rawdata" / "Info.csv"
    if info_path.exists():
        try:
            info_df = pd.read_csv(info_path, dtype=str)
            for key in ["公司简称", "公司名称"]:
                match = info_df.loc[info_df["项目"] == key, info_df.columns[-1]]
                if not match.empty:
                    value = str(match.iloc[0]).strip()
                    if value and value.lower() != "nan":
                        return value
        except Exception:
            pass
    return ticker


def build_historical_dataset() -> Dict[str, object]:
    bs = pd.read_csv(BS_PATH)
    pl = pd.read_csv(PL_PATH)
    cf = pd.read_csv(CF_PATH)
    info = pd.read_csv(INFO_PATH)

    bs_map = load_item_series(bs, "StandardLineItem", "Year", "Value")
    pl_map = load_item_series(pl, "Standard Item", "Year", "Value")
    cf_map = load_item_series(cf, "Standard Item", "Year", "Value")

    years = sorted(set(bs["Year"]).intersection(pl["Year"]).intersection(cf["Year"]))
    years = [int(y) for y in years]
    ticker = detect_ticker(BASE_DIR / "rawdata")
    company_name = detect_company_name(ticker)
    valuation_date = date.today().isoformat()

    revenue = get_series(pl_map, "Revenue", years)
    operating_profit = get_series(pl_map, "Operating Profit", years)
    financial_expense = get_series(pl_map, "Financial Expense", years)
    ebit = [op + fin for op, fin in zip(operating_profit, financial_expense)]
    profit_before_tax = get_series(pl_map, "Profit Before Tax", years)
    income_tax = get_series(pl_map, "Income Tax", years)
    parent_net_profit = get_series(pl_map, "Parent Net Profit", years)

    depreciation = get_series(cf_map, "Depreciation", years)
    amortization = get_series(cf_map, "Amortization", years)
    capex = [abs(v) for v in get_series(cf_map, "Capex", years)]
    cfo = get_series(cf_map, "Operating Cash Flow", years)
    fcff_proxy = [c - x for c, x in zip(cfo, capex)]

    cash = get_series(bs_map, "Cash & Short-term Financial Assets", years)
    short_debt = get_series(bs_map, "Interest-bearing Short-term Debt", years)
    long_debt = get_series(bs_map, "Long-term Interest-bearing Debt", years)
    minority_interest = get_series(bs_map, "Minority Interest", years)
    total_equity = get_series(bs_map, "Total Equity", years)
    non_op_current_assets = get_series(bs_map, "Non-operating Misc. Current Assets", years)
    long_term_investments = get_series(bs_map, "Long-term Financial & Equity Investments", years)
    core_operating_current_assets = get_series(bs_map, "Core Operating Current Assets", years)
    operating_current_liab = get_series(bs_map, "Operating Non-interest-bearing Current Liabilities", years)

    nwc = [ca - cl for ca, cl in zip(core_operating_current_assets, operating_current_liab)]
    tax_rate = []
    ebit_margin = []
    da_ratio = []
    capex_ratio = []
    nwc_ratio = []
    for idx, year in enumerate(years):
        rev = revenue[idx]
        pbt = profit_before_tax[idx]
        tax_rate.append(income_tax[idx] / pbt if abs(pbt) > 1e-9 else 0.20)
        ebit_margin.append(ebit[idx] / rev if abs(rev) > 1e-9 else 0.10)
        da_ratio.append((depreciation[idx] + amortization[idx]) / rev if abs(rev) > 1e-9 else 0.03)
        capex_ratio.append(capex[idx] / rev if abs(rev) > 1e-9 else 0.03)
        nwc_ratio.append(nwc[idx] / rev if abs(rev) > 1e-9 else 0.05)

    shares_outstanding = float(info.loc[info["项目"] == "总股本", info.columns[-1]].iloc[0])
    current_price = float(info.loc[info["项目"] == "当前股价", info.columns[-1]].iloc[0])

    base_year = max(years)
    growth_seed = cagr(revenue[-4], revenue[-1], 3) if len(revenue) >= 4 else 0.10
    growth_seed = clamp(growth_seed, 0.03, 0.20)
    base_ebit_margin = clamp(avg(ebit_margin[-3:], ebit_margin[-1]), 0.05, 0.35)
    base_tax_rate = clamp(avg(tax_rate[-3:], tax_rate[-1]), 0.10, 0.30)
    base_da_ratio = clamp(avg(da_ratio[-3:], da_ratio[-1]), 0.01, 0.08)
    base_capex_ratio = clamp(avg(capex_ratio[-3:], capex_ratio[-1]), 0.01, 0.12)
    base_nwc_ratio = clamp(avg(nwc_ratio[-3:], nwc_ratio[-1]), 0.00, 0.25)

    forecast_years = [base_year + i for i in range(1, 6)]
    default_growths = [
        round(growth_seed, 4),
        round(max(growth_seed - 0.015, 0.04), 4),
        round(max(growth_seed - 0.030, 0.035), 4),
        round(max(growth_seed - 0.045, 0.030), 4),
        round(max(growth_seed - 0.055, 0.025), 4),
    ]

    return {
        "years": years,
        "forecast_years": forecast_years,
        "base_year": base_year,
        "revenue": revenue,
        "operating_profit": operating_profit,
        "financial_expense": financial_expense,
        "ebit": ebit,
        "profit_before_tax": profit_before_tax,
        "income_tax": income_tax,
        "parent_net_profit": parent_net_profit,
        "depreciation": depreciation,
        "amortization": amortization,
        "capex": capex,
        "cfo": cfo,
        "fcff_proxy": fcff_proxy,
        "cash": cash,
        "short_debt": short_debt,
        "long_debt": long_debt,
        "minority_interest": minority_interest,
        "total_equity": total_equity,
        "non_op_current_assets": non_op_current_assets,
        "long_term_investments": long_term_investments,
        "nwc": nwc,
        "tax_rate": tax_rate,
        "ebit_margin": ebit_margin,
        "da_ratio": da_ratio,
        "capex_ratio": capex_ratio,
        "nwc_ratio": nwc_ratio,
        "shares_outstanding": shares_outstanding,
        "current_price": current_price,
        "default_growths": default_growths,
        "ticker": ticker,
        "company_name": company_name,
        "valuation_date": valuation_date,
        "base_ebit_margin": round(base_ebit_margin, 4),
        "base_tax_rate": round(base_tax_rate, 4),
        "base_da_ratio": round(base_da_ratio, 4),
        "base_capex_ratio": round(base_capex_ratio, 4),
        "base_nwc_ratio": round(base_nwc_ratio, 4),
    }


def apply_title_style(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_header_style(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4F81BD")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D9E2F3")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def apply_input_style(cell) -> None:
    cell.fill = PatternFill("solid", fgColor="FFF2CC")
    cell.font = Font(color="7F6000")


def apply_formula_style(cell) -> None:
    cell.fill = PatternFill("solid", fgColor="E2F0D9")


def set_col_widths(ws, widths: Dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_note(ws, cell_ref: str, text: str) -> None:
    ws[cell_ref] = text
    ws[cell_ref].font = Font(italic=True, color="666666")


def create_summary_sheet(wb: Workbook, data: Dict[str, object]) -> None:
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "估值总览"
    apply_title_style(ws["A1"])
    ws.merge_cells("A1:J1")
    ws["A2"] = f"Ticker: {data['ticker']}    公司简称: {data['company_name']}    估值日期: {data['valuation_date']}"
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:J2")

    summary_rows = [
        ("基准年度", f"{data['base_year']}A"),
        ("当前股价", "=Assumptions!B4"),
        ("总股本", "=Assumptions!B3"),
        ("WACC", "=Assumptions!B6"),
        ("永续增长率", "=Assumptions!B7"),
        ("企业价值 EV", "=DCF!B14"),
        ("股东权益价值", "=DCF!B20"),
        ("每股内在价值", "=DCF!B21"),
        ("相对当前股价空间", "=DCF!B22"),
    ]
    ws["A3"] = "核心结果"
    apply_header_style(ws["A3"])
    ws["B3"] = "数值"
    apply_header_style(ws["B3"])
    for idx, (label, formula) in enumerate(summary_rows, start=4):
        ws[f"A{idx}"] = label
        ws[f"B{idx}"] = formula
        if isinstance(formula, str) and formula.startswith("="):
            apply_formula_style(ws[f"B{idx}"])

    ws["D3"] = "目标价汇总"
    ws["E3"] = "低"
    ws["F3"] = "中"
    ws["G3"] = "高"
    ws["H3"] = "权重"
    ws["I3"] = "加权目标价"
    for ref in ["D3", "E3", "F3", "G3", "H3", "I3"]:
        apply_header_style(ws[ref])

    target_rows = [
        (4, "DCF", "=DCF!B21*0.85", "=DCF!B21", "=DCF!B21*1.15", 0.60),
        (5, "Relative Valuation", "=MIN(Comparable!C20:C23)", "=Comparable!C25", "=MAX(Comparable!E20:E23)", 0.40),
    ]
    for row, method, low_formula, mid_formula, high_formula, weight in target_rows:
        ws[f"D{row}"] = method
        ws[f"E{row}"] = low_formula
        ws[f"F{row}"] = mid_formula
        ws[f"G{row}"] = high_formula
        ws[f"H{row}"] = weight
        ws[f"I{row}"] = f"=F{row}*H{row}"
        for ref in [f"E{row}", f"F{row}", f"G{row}", f"I{row}"]:
            apply_formula_style(ws[ref])
        apply_input_style(ws[f"H{row}"])

    ws["D7"] = "综合目标价"
    ws["E7"] = "=SUMPRODUCT(E4:E5,H4:H5)"
    ws["F7"] = "=SUMPRODUCT(F4:F5,H4:H5)"
    ws["G7"] = "=SUMPRODUCT(G4:G5,H4:H5)"
    ws["H7"] = "=SUM(H4:H5)"
    ws["I7"] = "=SUM(I4:I5)"
    for ref in ["D7", "E7", "F7", "G7", "H7", "I7"]:
        if ref == "D7":
            apply_header_style(ws[ref])
        else:
            apply_formula_style(ws[ref])

    ws["D9"] = "相对当前股价空间"
    ws["E9"] = "=E7/$B$5-1"
    ws["F9"] = "=F7/$B$5-1"
    ws["G9"] = "=G7/$B$5-1"
    for ref in ["D9", "E9", "F9", "G9"]:
        if ref == "D9":
            apply_header_style(ws[ref])
        else:
            apply_formula_style(ws[ref])

    ws["D12"] = "使用说明"
    apply_header_style(ws["D12"])
    ws["D13"] = "黄色区域为可编辑假设"
    ws["D14"] = "绿色区域为公式联动结果"
    ws["D15"] = "修改 Assumptions 或 Comparable 后本页自动刷新"
    ws["D16"] = "建议优先调整：收入增速、EBIT率、WACC、永续增长率、可比倍数"
    ws.freeze_panes = "A3"
    set_col_widths(ws, {"A": 20, "B": 16, "D": 20, "E": 12, "F": 12, "G": 12, "H": 10, "I": 14})
    ws["B4"].number_format = "0.00"       # 基准年度（文本，格式不影响显示）
    ws["B5"].number_format = "0.00"       # 当前股价（元/股，保留两位小数）
    ws["B6"].number_format = "#,##0"      # 总股本（股数，整数带千分位）
    ws["B7"].number_format = "0.0%"       # WACC（百分比）
    ws["B8"].number_format = "0.0%"       # 永续增长率（百分比）
    ws["B9"].number_format = "#,##0"      # 企业价值 EV（元，整数带千分位）
    ws["B10"].number_format = "#,##0"     # 股东权益价值（元，整数带千分位）
    ws["B11"].number_format = "0.00"      # 每股内在价值（元/股，两位小数）
    ws["B12"].number_format = "0.0%"      # 相对当前股价空间（百分比）
    for ref in ["E4", "F4", "G4", "I4", "E5", "F5", "G5", "I5", "E7", "F7", "G7", "I7"]:
        ws[ref].number_format = "0.00"
    for ref in ["H4", "H5", "H7", "E9", "F9", "G9"]:
        ws[ref].number_format = "0.0%"


def create_historical_sheet(wb: Workbook, data: Dict[str, object]) -> None:
    ws = wb.create_sheet("Historical")
    ws["A1"] = "历史财务整理"
    apply_title_style(ws["A1"])
    ws.merge_cells("A1:L1")

    years = data["years"]
    headers = ["Metric"] + years
    for col_idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=value)
        apply_header_style(cell)

    metrics = [
        ("Revenue", data["revenue"]),
        ("EBIT", data["ebit"]),
        ("EBIT Margin", data["ebit_margin"]),
        ("Income Tax", data["income_tax"]),
        ("Tax Rate", data["tax_rate"]),
        ("Parent Net Profit", data["parent_net_profit"]),
        ("Depreciation", data["depreciation"]),
        ("Amortization", data["amortization"]),
        ("Capex", data["capex"]),
        ("Operating Cash Flow", data["cfo"]),
        ("FCFF Proxy", data["fcff_proxy"]),
        ("Operating NWC", data["nwc"]),
        ("NWC / Revenue", data["nwc_ratio"]),
        ("Cash", data["cash"]),
        ("Short Debt", data["short_debt"]),
        ("Long Debt", data["long_debt"]),
        ("Total Equity", data["total_equity"]),
        ("Minority Interest", data["minority_interest"]),
        ("Long-term Investments", data["long_term_investments"]),
        ("Non-op Current Assets", data["non_op_current_assets"]),
    ]
    for row_idx, (metric, values) in enumerate(metrics, start=4):
        ws.cell(row=row_idx, column=1, value=metric)
        for col_idx, value in enumerate(values, start=2):
            ws.cell(row=row_idx, column=col_idx, value=float(value))

    ws.freeze_panes = "B4"
    set_col_widths(ws, {"A": 22, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14, "I": 14, "J": 14, "K": 14})
    for row in range(4, 24):
        for col in range(2, 2 + len(years)):
            cell = ws.cell(row=row, column=col)
            if row in {6, 8, 16}:
                cell.number_format = "0.0%"
            else:
                cell.number_format = '#,##0.00'


def create_assumptions_sheet(wb: Workbook, data: Dict[str, object]) -> None:
    ws = wb.create_sheet("Assumptions")
    ws["A1"] = "DCF假设区"
    apply_title_style(ws["A1"])
    ws.merge_cells("A1:H1")

    labels = [
        ("B3", data["shares_outstanding"], "总股本"),
        ("B4", data["current_price"], "当前股价"),
        ("B5", data["cash"][-1] - data["short_debt"][-1] - data["long_debt"][-1], "净现金/(净债务)"),
        ("B6", 0.10, "WACC"),
        ("B7", 0.03, "永续增长率"),
        ("B8", data["minority_interest"][-1], "少数股东权益"),
        ("B9", data["long_term_investments"][-1], "长期金融投资加回"),
        ("B10", data["non_op_current_assets"][-1], "非经营流动资产加回"),
    ]
    ws["A3"] = "全局假设"
    apply_header_style(ws["A3"])
    ws["B3"] = "数值"
    apply_header_style(ws["B3"])
    ws["C3"] = "说明"
    apply_header_style(ws["C3"])
    for cell_ref, value, desc in labels:
        ws[f"A{cell_ref[1:]}"] = desc
        ws[cell_ref] = float(value)
        ws[f"C{cell_ref[1:]}"] = "可编辑"
        apply_input_style(ws[cell_ref])

    forecast_years = data["forecast_years"]
    start_row = 13
    ws[f"A{start_row}"] = "逐年预测假设"
    apply_header_style(ws[f"A{start_row}"])
    for idx, year in enumerate(forecast_years, start=2):
        cell = ws.cell(row=start_row, column=idx, value=year)
        apply_header_style(cell)

    assumptions = [
        ("Revenue Growth", data["default_growths"], "收入增速"),
        ("EBIT Margin", [data["base_ebit_margin"]] * 5, "EBIT率"),
        ("Tax Rate", [data["base_tax_rate"]] * 5, "税率"),
        ("D&A % Revenue", [data["base_da_ratio"]] * 5, "折旧摊销占收入"),
        ("Capex % Revenue", [data["base_capex_ratio"]] * 5, "资本开支占收入"),
        ("NWC % Revenue", [data["base_nwc_ratio"]] * 5, "营运资本占收入"),
    ]
    for row_idx, (name, values, cn_name) in enumerate(assumptions, start=start_row + 1):
        ws.cell(row=row_idx, column=1, value=cn_name)
        for col_idx, value in enumerate(values, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=float(value))
            apply_input_style(cell)
            cell.number_format = "0.0%"

    add_note(ws, "A22", "黄色单元格可直接修改，Forecast / DCF / Summary 会自动联动。")
    ws.freeze_panes = "B14"
    set_col_widths(ws, {"A": 22, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14})
    for cell_ref in ["B3", "B5", "B8", "B9", "B10"]:
        ws[cell_ref].number_format = '#,##0.00'
    ws["B4"].number_format = "0.00"
    ws["B6"].number_format = "0.0%"
    ws["B7"].number_format = "0.0%"


def create_forecast_sheet(wb: Workbook, data: Dict[str, object]) -> None:
    ws = wb.create_sheet("Forecast")
    ws["A1"] = "5年经营预测与FCFF"
    apply_title_style(ws["A1"])
    ws.merge_cells("A1:J1")

    years = data["years"][-3:] + data["forecast_years"]
    for idx, year in enumerate(["Metric"] + years, start=1):
        cell = ws.cell(row=3, column=idx, value=year)
        apply_header_style(cell)

    metrics = [
        "Revenue",
        "Revenue Growth",
        "EBIT Margin",
        "EBIT",
        "Tax Rate",
        "Tax on EBIT",
        "NOPAT",
        "D&A % Revenue",
        "D&A",
        "Capex % Revenue",
        "Capex",
        "NWC % Revenue",
        "Operating NWC",
        "Change in NWC",
        "FCFF",
    ]

    row_map: Dict[str, int] = {}
    for row_idx, metric in enumerate(metrics, start=4):
        row_map[metric] = row_idx
        ws.cell(row=row_idx, column=1, value=metric)

    hist_years = data["years"][-3:]
    hist_offset = 2
    hist_index_map = {year: data["years"].index(year) for year in hist_years}
    for col_idx, year in enumerate(hist_years, start=2):
        idx = hist_index_map[year]
        ws.cell(row=row_map["Revenue"], column=col_idx, value=data["revenue"][idx])
        prev_revenue = data["revenue"][idx - 1] if idx > 0 else None
        if prev_revenue and abs(prev_revenue) > 1e-9:
            ws.cell(row=row_map["Revenue Growth"], column=col_idx, value=(data["revenue"][idx] / prev_revenue) - 1)
        ws.cell(row=row_map["EBIT Margin"], column=col_idx, value=data["ebit_margin"][idx])
        ws.cell(row=row_map["EBIT"], column=col_idx, value=data["ebit"][idx])
        ws.cell(row=row_map["Tax Rate"], column=col_idx, value=data["tax_rate"][idx])
        ws.cell(row=row_map["Tax on EBIT"], column=col_idx, value=data["ebit"][idx] * data["tax_rate"][idx])
        ws.cell(row=row_map["NOPAT"], column=col_idx, value=data["ebit"][idx] - data["ebit"][idx] * data["tax_rate"][idx])
        da = data["depreciation"][idx] + data["amortization"][idx]
        ws.cell(row=row_map["D&A % Revenue"], column=col_idx, value=da / data["revenue"][idx] if data["revenue"][idx] else 0.0)
        ws.cell(row=row_map["D&A"], column=col_idx, value=da)
        ws.cell(row=row_map["Capex % Revenue"], column=col_idx, value=data["capex"][idx] / data["revenue"][idx] if data["revenue"][idx] else 0.0)
        ws.cell(row=row_map["Capex"], column=col_idx, value=data["capex"][idx])
        ws.cell(row=row_map["NWC % Revenue"], column=col_idx, value=data["nwc_ratio"][idx])
        ws.cell(row=row_map["Operating NWC"], column=col_idx, value=data["nwc"][idx])
        prev_nwc = data["nwc"][idx - 1] if idx > 0 else None
        if prev_nwc is not None:
            ws.cell(row=row_map["Change in NWC"], column=col_idx, value=data["nwc"][idx] - prev_nwc)
        ws.cell(row=row_map["FCFF"], column=col_idx, value=data["fcff_proxy"][idx])

    for forecast_idx, year in enumerate(data["forecast_years"], start=5):
        col = forecast_idx
        ass_col = get_column_letter(forecast_idx - 3)
        prev_col_letter = get_column_letter(col - 1)
        cur_col_letter = get_column_letter(col)
        ws[f"{cur_col_letter}{row_map['Revenue']}"] = f"={prev_col_letter}{row_map['Revenue']}*(1+Assumptions!{ass_col}14)"
        ws[f"{cur_col_letter}{row_map['Revenue Growth']}"] = f"=Assumptions!{ass_col}14"
        ws[f"{cur_col_letter}{row_map['EBIT Margin']}"] = f"=Assumptions!{ass_col}15"
        ws[f"{cur_col_letter}{row_map['EBIT']}"] = f"={cur_col_letter}{row_map['Revenue']}*{cur_col_letter}{row_map['EBIT Margin']}"
        ws[f"{cur_col_letter}{row_map['Tax Rate']}"] = f"=Assumptions!{ass_col}16"
        ws[f"{cur_col_letter}{row_map['Tax on EBIT']}"] = f"={cur_col_letter}{row_map['EBIT']}*{cur_col_letter}{row_map['Tax Rate']}"
        ws[f"{cur_col_letter}{row_map['NOPAT']}"] = f"={cur_col_letter}{row_map['EBIT']}-{cur_col_letter}{row_map['Tax on EBIT']}"
        ws[f"{cur_col_letter}{row_map['D&A % Revenue']}"] = f"=Assumptions!{ass_col}17"
        ws[f"{cur_col_letter}{row_map['D&A']}"] = f"={cur_col_letter}{row_map['Revenue']}*{cur_col_letter}{row_map['D&A % Revenue']}"
        ws[f"{cur_col_letter}{row_map['Capex % Revenue']}"] = f"=Assumptions!{ass_col}18"
        ws[f"{cur_col_letter}{row_map['Capex']}"] = f"={cur_col_letter}{row_map['Revenue']}*{cur_col_letter}{row_map['Capex % Revenue']}"
        ws[f"{cur_col_letter}{row_map['NWC % Revenue']}"] = f"=Assumptions!{ass_col}19"
        ws[f"{cur_col_letter}{row_map['Operating NWC']}"] = f"={cur_col_letter}{row_map['Revenue']}*{cur_col_letter}{row_map['NWC % Revenue']}"
        ws[f"{cur_col_letter}{row_map['Change in NWC']}"] = f"={cur_col_letter}{row_map['Operating NWC']}-{prev_col_letter}{row_map['Operating NWC']}"
        ws[f"{cur_col_letter}{row_map['FCFF']}"] = f"={cur_col_letter}{row_map['NOPAT']}+{cur_col_letter}{row_map['D&A']}-{cur_col_letter}{row_map['Capex']}-{cur_col_letter}{row_map['Change in NWC']}"

    for row in range(4, 19):
        for col in range(2, 10):
            cell = ws.cell(row=row, column=col)
            if row in {
                row_map["Revenue Growth"],
                row_map["EBIT Margin"],
                row_map["Tax Rate"],
                row_map["D&A % Revenue"],
                row_map["Capex % Revenue"],
                row_map["NWC % Revenue"],
            }:
                cell.number_format = "0.0%"
            else:
                cell.number_format = '#,##0.00'
            if col >= 5:
                apply_formula_style(cell)

    add_note(ws, "A20", "左侧为历史，右侧为预测。预测列全部由 Assumptions 页驱动。")
    ws.freeze_panes = "B4"
    set_col_widths(ws, {"A": 20, "B": 13, "C": 13, "D": 13, "E": 13, "F": 13, "G": 13, "H": 13, "I": 13})


def create_dcf_sheet(wb: Workbook, data: Dict[str, object]) -> None:
    ws = wb.create_sheet("DCF")
    ws["A1"] = "DCF估值主表"
    apply_title_style(ws["A1"])
    ws.merge_cells("A1:F1")

    ws["A3"] = "项目"
    ws["B3"] = "数值"
    ws["C3"] = "说明"
    for cell_ref in ["A3", "B3", "C3"]:
        apply_header_style(ws[cell_ref])

    rows = [
        ("B5", "Assumptions!B6", "WACC"),
        ("B6", "Assumptions!B7", "永续增长率"),
        ("B7", "Forecast!I18", "FCFF(终值前一年)"),
        ("B8", "B7*(1+B6)/(B5-B6)", "终值 TV"),
        ("B9", "Forecast!E18/(1+B5)^1", "Year 1 FCFF 现值"),
        ("B10", "Forecast!F18/(1+B5)^2", "Year 2 FCFF 现值"),
        ("B11", "Forecast!G18/(1+B5)^3", "Year 3 FCFF 现值"),
        ("B12", "Forecast!H18/(1+B5)^4", "Year 4 FCFF 现值"),
        ("B13", "Forecast!I18/(1+B5)^5", "Year 5 FCFF 现值"),
        ("B14", "SUM(B9:B13)+B8/(1+B5)^5", "企业价值 EV"),
        ("B16", "-Assumptions!B5", "减：净债务（净现金为负值时会加回）"),
        ("B17", "-Assumptions!B8", "减：少数股东权益"),
        ("B18", "Assumptions!B9", "加：长期金融投资"),
        ("B19", "Assumptions!B10", "加：非经营流动资产"),
        ("B20", "B14+B16+B17+B18+B19", "股东权益价值"),
        ("B21", "B20/Assumptions!B3", "每股内在价值"),
        ("B22", "B21/Assumptions!B4-1", "相对当前股价空间"),
    ]
    for cell_ref, formula_body, label in rows:
        row = int(cell_ref[1:])
        ws[f"A{row}"] = label
        ws[cell_ref] = f"={formula_body}"
        apply_formula_style(ws[cell_ref])

    ws["E3"] = "关键桥接"
    apply_header_style(ws["E3"])
    ws["E4"] = "净债务口径 = 有息短债 + 有息长债 - 现金及短期金融资产"
    ws["E5"] = "权益价值 = EV - 净债务 - 少数股东权益 + 非经营资产"
    ws["E6"] = "FCFF = NOPAT + D&A - Capex - ΔNWC"
    ws["B5"].number_format = "0.0%"
    ws["B6"].number_format = "0.0%"
    for ref in ["B7", "B8", "B9", "B10", "B11", "B12", "B13", "B14", "B16", "B17", "B18", "B19", "B20"]:
        ws[ref].number_format = '#,##0.00'
    ws["B21"].number_format = "0.00"
    ws["B22"].number_format = "0.0%"
    ws.freeze_panes = "A4"
    set_col_widths(ws, {"A": 28, "B": 16, "C": 18, "E": 52})


def create_sensitivity_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Sensitivity")
    ws["A1"] = "WACC / 永续增长率 敏感性"
    apply_title_style(ws["A1"])
    ws.merge_cells("A1:J1")

    growths = [0.01, 0.02, 0.03, 0.04, 0.05]
    waccs = [0.08, 0.09, 0.10, 0.11, 0.12]

    ws["A3"] = "g \\ WACC"
    apply_header_style(ws["A3"])
    for idx, wacc in enumerate(waccs, start=2):
        cell = ws.cell(row=3, column=idx, value=wacc)
        apply_header_style(cell)
        cell.number_format = "0.0%"

    for row_idx, growth in enumerate(growths, start=4):
        cell = ws.cell(row=row_idx, column=1, value=growth)
        apply_header_style(cell)
        cell.number_format = "0.0%"
        for col_idx, wacc in enumerate(waccs, start=2):
            growth_cell = f"$A{row_idx}"
            wacc_cell = f"{get_column_letter(col_idx)}$3"
            formula = (
                f"=(SUM(DCF!$B$9:$B$13)+"
                f"(Forecast!$I$18*(1+{growth_cell})/({wacc_cell}-{growth_cell}))/(1+{wacc_cell})^5"
                f"+DCF!$B$16+DCF!$B$17+DCF!$B$18+DCF!$B$19)/Assumptions!$B$3"
            )
            ws.cell(row=row_idx, column=col_idx, value=f"={formula[1:]}")
            ws.cell(row=row_idx, column=col_idx).number_format = "0.00"

    green_fill = PatternFill("solid", fgColor="E2F0D9")
    red_fill = PatternFill("solid", fgColor="FCE4D6")
    ws.conditional_formatting.add("B4:F8", CellIsRule(operator="greaterThan", formula=["DCF!$B$21"], fill=green_fill))
    ws.conditional_formatting.add("B4:F8", CellIsRule(operator="lessThan", formula=["DCF!$B$21"], fill=red_fill))
    ws.freeze_panes = "B4"
    set_col_widths(ws, {"A": 12, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12})


def create_comparable_sheet(wb: Workbook, data: Dict[str, object]) -> None:
    ws = wb.create_sheet("Comparable")
    ws["A1"] = "相对估值"
    apply_title_style(ws["A1"])
    ws.merge_cells("A1:J1")

    ws["A3"] = "当前市场口径"
    apply_header_style(ws["A3"])
    ws["B3"] = "数值"
    apply_header_style(ws["B3"])
    ws["C3"] = "说明"
    apply_header_style(ws["C3"])

    market_rows = [
        ("A4", "当前股价", "=Assumptions!B4", "来自 Info.csv，可编辑"),
        ("A5", "总股本", "=Assumptions!B3", "来自 Info.csv，可编辑"),
        ("A6", "当前市值", "=B4*B5", "股价 * 总股本"),
        ("A7", "净债务", "=-Assumptions!B5", "净现金为负值，净负债为正值"),
        ("A8", "少数股东权益", "=Assumptions!B8", "来自 BS"),
        ("A9", "企业价值 EV", "=B6+B7+B8-Assumptions!B9-Assumptions!B10", "EV = 市值 + 净债务 + 少数股东权益 - 非经营资产"),
    ]
    for label_cell, label, value_formula, note in market_rows:
        row = int(label_cell[1:])
        ws[label_cell] = label
        ws[f"B{row}"] = value_formula
        ws[f"C{row}"] = note
        apply_formula_style(ws[f"B{row}"])

    ws["E3"] = "可比倍数假设"
    apply_header_style(ws["E3"])
    for col, header in zip(["F", "G", "H"], ["低", "中", "高"]):
        ws[f"{col}3"] = header
        apply_header_style(ws[f"{col}3"])

    assumption_rows = [
        (4, "PE", [18.0, 22.0, 26.0], "x"),
        (5, "PB", [3.0, 3.8, 4.5], "x"),
        (6, "EV/EBIT", [16.0, 20.0, 24.0], "x"),
        (7, "EV/EBITDA", [13.0, 16.0, 19.0], "x"),
    ]
    for row, label, vals, _ in assumption_rows:
        ws[f"E{row}"] = label
        for col, val in zip(["F", "G", "H"], vals):
            ws[f"{col}{row}"] = float(val)
            apply_input_style(ws[f"{col}{row}"])

    ws["A12"] = "历史与当前倍数"
    apply_header_style(ws["A12"])
    for col, header in zip(["B", "C", "D", "E"], ["指标", "数值", "公式说明", "当前倍数"]):
        ws[f"{col}12"] = header
        apply_header_style(ws[f"{col}12"])

    current_metric_rows = [
        (13, "归母净利润", data["parent_net_profit"][-1], "TTM近似采用2024A", "=B6/B13"),
        (14, "股东权益", 0.0, "2024A Total Equity", ""),
        (15, "EBIT", data["ebit"][-1], "2024A", "=B9/B15"),
        (16, "EBITDA", data["ebit"][-1] + data["depreciation"][-1] + data["amortization"][-1], "EBIT + D&A", "=B9/B16"),
    ]
    # Overwrite book value using BS-based equity instead of placeholder market cap.
    book_value = (
        data["cash"][-1]
    )  # temp to keep expression simple before writing real cell formulas below
    for row, label, value, note, multiple_formula in current_metric_rows:
        ws[f"A{row}"] = label
        if row == 14:
            ws[f"B{row}"] = "=Historical!K20"
            ws[f"C{row}"] = note
            ws[f"E{row}"] = "=B6/B14"
        else:
            ws[f"B{row}"] = float(value)
            ws[f"C{row}"] = note
            ws[f"E{row}"] = multiple_formula
        if row != 14:
            ws[f"D{row}"] = ""
        apply_formula_style(ws[f"E{row}"])
        if row == 14:
            apply_formula_style(ws[f"B{row}"])

    ws["D13"] = "PE"
    ws["D14"] = "PB"
    ws["D15"] = "EV/EBIT"
    ws["D16"] = "EV/EBITDA"

    ws["A19"] = "相对估值推导股价"
    apply_header_style(ws["A19"])
    for col, header in zip(["B", "C", "D", "E"], ["方法", "低", "中", "高"]):
        ws[f"{col}19"] = header
        apply_header_style(ws[f"{col}19"])

    # Price targets:
    # PE: Equity Value = PE * Parent NP
    # PB: Equity Value = PB * Book Value
    # EV-based: Equity = EV - NetDebt - Minority + investments + non-op assets
    ws["B20"] = "PE"
    ws["C20"] = "=(F4*$B$13)/$B$5"
    ws["D20"] = "=(G4*$B$13)/$B$5"
    ws["E20"] = "=(H4*$B$13)/$B$5"

    ws["B21"] = "PB"
    ws["C21"] = "=(F5*$B$14)/$B$5"
    ws["D21"] = "=(G5*$B$14)/$B$5"
    ws["E21"] = "=(H5*$B$14)/$B$5"

    ev_bridge = "-$B$7-$B$8+Assumptions!$B$9+Assumptions!$B$10"
    ws["B22"] = "EV/EBIT"
    ws["C22"] = f"=((F6*$B$15){ev_bridge})/$B$5"
    ws["D22"] = f"=((G6*$B$15){ev_bridge})/$B$5"
    ws["E22"] = f"=((H6*$B$15){ev_bridge})/$B$5"

    ws["B23"] = "EV/EBITDA"
    ws["C23"] = f"=((F7*$B$16){ev_bridge})/$B$5"
    ws["D23"] = f"=((G7*$B$16){ev_bridge})/$B$5"
    ws["E23"] = f"=((H7*$B$16){ev_bridge})/$B$5"

    ws["B25"] = "综合参考价"
    ws["C25"] = "=AVERAGE(D20:D23)"
    ws["D25"] = "相对估值中性情景均值"
    apply_formula_style(ws["C25"])

    ws["B26"] = "相对估值空间"
    ws["C26"] = "=C25/Assumptions!B4-1"
    ws["D26"] = "相对当前股价"
    apply_formula_style(ws["C26"])

    add_note(ws, "A28", "F:H 为可编辑可比倍数假设；C:E 会自动反推对应股价。")
    ws.freeze_panes = "A4"
    set_col_widths(ws, {"A": 18, "B": 16, "C": 24, "D": 14, "E": 14, "F": 10, "G": 10, "H": 10})

    for ref in ["B4", "C20", "D20", "E20", "C21", "D21", "E21", "C22", "D22", "E22", "C23", "D23", "E23", "C25"]:
        ws[ref].number_format = "0.00"
    ws["C26"].number_format = "0.0%"
    for ref in ["B5", "B6", "B7", "B8", "B9", "B13", "B14", "B15", "B16"]:
        ws[ref].number_format = '#,##0.00'
    for ref in ["E13", "E14", "E15", "E16"]:
        ws[ref].number_format = "0.00x"


def create_investment_thesis_sheet(wb: Workbook, data: Dict[str, object]) -> None:
    ws = wb.create_sheet("Investment_Thesis")
    ws["A1"] = "投资评级与核心观点"
    apply_title_style(ws["A1"])
    ws.merge_cells("A1:J1")

    ws["A2"] = f"{data['company_name']} ({data['ticker']}) | 估值日期: {data['valuation_date']} | 基于DCF与相对估值的综合判断"
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:J2")

    section_headers = ["A4", "F4", "A11", "F11", "A18"]
    for ref in section_headers:
        apply_header_style(ws[ref])
    ws["A4"] = "评级与目标价"
    ws["F4"] = "核心假设"
    ws["A11"] = "投资逻辑"
    ws["F11"] = "催化剂"
    ws["A18"] = "风险提示"

    ws["A5"] = "当前股价"
    ws["B5"] = "=Assumptions!B4"
    ws["A6"] = "综合目标价"
    ws["B6"] = "=Summary!F7"
    ws["A7"] = "预期空间"
    ws["B7"] = "=Summary!F9"
    ws["A8"] = "DCF目标价"
    ws["B8"] = "=DCF!B21"
    ws["A9"] = "相对估值目标价"
    ws["B9"] = "=Comparable!C25"
    ws["D5"] = "投资评级"
    ws["E5"] = '=IF(B7>=0.30,"买入",IF(B7>=0.15,"增持",IF(B7>=-0.10,"中性","减持")))'
    ws["D6"] = "估值结论"
    ws["E6"] = '=IF(B7>=0.15,"当前股价相对综合估值仍有明显修复空间",IF(B7>=-0.10,"当前股价与模型估值大致匹配","当前股价已高于模型中性估值"))'
    ws["D7"] = "估值方法"
    ws["E7"] = "DCF 60% + Relative 40%"
    ws["D8"] = "目标价区间"
    ws["E8"] = '=TEXT(Summary!E7,"0.00")&" - "&TEXT(Summary!G7,"0.00")'
    ws["D9"] = "评级说明"
    ws["E9"] = "可结合你自己的行业判断调整评级与权重"

    for ref in ["B5", "B6", "B8", "B9"]:
        apply_formula_style(ws[ref])
        ws[ref].number_format = "0.00"
    apply_formula_style(ws["B7"])
    ws["B7"].number_format = "0.0%"
    for ref in ["E5", "E6", "E8"]:
        apply_formula_style(ws[ref])

    assumptions = [
        ("F5", "收入增长"),
        ("F6", "EBIT率"),
        ("F7", "税率"),
        ("F8", "WACC"),
        ("F9", "永续增长率"),
    ]
    assumption_values = {
        "G5": '=TEXT(Assumptions!B14,"0.0%")&" / "&TEXT(Assumptions!F14,"0.0%")',
        "G6": '=TEXT(Assumptions!B15,"0.0%")',
        "G7": '=TEXT(Assumptions!B16,"0.0%")',
        "G8": '=TEXT(Assumptions!B6,"0.0%")',
        "G9": '=TEXT(Assumptions!B7,"0.0%")',
    }
    assumption_notes = {
        "H5": "未来5年收入增速由高到低回落",
        "H6": "采用中性经营利润率假设",
        "H7": "使用历史有效税率近似",
        "H8": "反映资本成本与风险补偿",
        "H9": "决定终值估算弹性",
    }
    for ref, label in assumptions:
        row = ref[1:]
        ws[ref] = label
        ws[f"G{row}"] = assumption_values[f"G{row}"]
        ws[f"H{row}"] = assumption_notes[f"H{row}"]
        apply_formula_style(ws[f"G{row}"])

    thesis_lines = [
        "1. 公司历史收入、EBIT与现金流具备一定延续性，模型假设未来5年仍维持增长但增速逐年回归。",
        "2. DCF 用于刻画长期自由现金流价值，相对估值用于反映市场可交易区间，两者结合可降低单一方法偏差。",
        "3. 目前综合目标价高于现价时，意味着盈利兑现与估值修复均可能贡献收益空间。",
        "4. 若后续经营改善超预期，利润率与自由现金流同步抬升，目标价存在继续上修可能。",
    ]
    catalysts = [
        "1. 收入增速改善或新品/新业务放量，带来盈利预期上修。",
        "2. 毛利率、费用率或资本开支改善，自由现金流表现超出当前模型假设。",
        "3. 行业估值中枢上移，带动PE / EV倍数同步修复。",
        "4. 回购、分红、资产处置等资本运作提升股东回报预期。",
    ]
    risks = [
        "1. 收入增速不及预期，导致利润与估值锚同步下修。",
        "2. 资本开支或营运资本占用持续上升，压缩FCFF与DCF价值。",
        "3. 利率上行或风险偏好走弱，WACC提升将压缩估值。",
        "4. 可比公司估值中枢下移时，相对估值法目标价会受到拖累。",
    ]

    for idx, text in enumerate(thesis_lines, start=12):
        ws[f"A{idx}"] = text
    for idx, text in enumerate(catalysts, start=12):
        ws[f"F{idx}"] = text
    for idx, text in enumerate(risks, start=19):
        ws[f"A{idx}"] = text

    highlight_fill = PatternFill("solid", fgColor="F3F9FD")
    for row in range(5, 10):
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = highlight_fill
    for row in range(5, 10):
        for col in range(6, 9):
            ws.cell(row=row, column=col).fill = highlight_fill

    set_col_widths(ws, {"A": 18, "B": 14, "C": 3, "D": 14, "E": 34, "F": 16, "G": 16, "H": 28, "I": 3, "J": 3})
    ws.freeze_panes = "A5"


def create_charts_sheet(wb: Workbook, data: Dict[str, object]) -> None:
    ws = wb.create_sheet("Charts")
    ws["A1"] = "估值图表"
    apply_title_style(ws["A1"])
    ws.merge_cells("A1:L1")

    years = data["years"][-3:] + data["forecast_years"]
    ws["A3"] = "Year"
    ws["B3"] = "Revenue"
    ws["C3"] = "EBIT"
    ws["D3"] = "FCFF"
    ws["F3"] = "Valuation Method"
    ws["G3"] = "Low"
    ws["H3"] = "Mid"
    ws["I3"] = "High"
    for ref in ["A3", "B3", "C3", "D3", "F3", "G3", "H3", "I3"]:
        apply_header_style(ws[ref])

    for idx, year in enumerate(years, start=4):
        ws[f"A{idx}"] = year
    for idx, col in enumerate(range(2, 10), start=4):
        pass

    source_cols = {"Revenue": "B", "EBIT": "C", "FCFF": "D"}
    forecast_lookup = {"Revenue": 4, "EBIT": 7, "FCFF": 18}
    hist_lookup = {
        "Revenue": ("Historical", 4),
        "EBIT": ("Historical", 5),
        "FCFF": ("Historical", 14),
    }
    for i, year in enumerate(years, start=4):
        if i <= 6:
            hist_col = get_column_letter(i + 7)  # H,I,J -> 2022,2023,2024 in Historical
            ws[f"B{i}"] = f"=Historical!{hist_col}4"
            ws[f"C{i}"] = f"=Historical!{hist_col}5"
            ws[f"D{i}"] = f"=Historical!{hist_col}14"
        else:
            # Forecast sheet columns E:I correspond 2025-2029
            fc_col = get_column_letter(i - 2)
            ws[f"B{i}"] = f"=Forecast!{fc_col}4"
            ws[f"C{i}"] = f"=Forecast!{fc_col}7"
            ws[f"D{i}"] = f"=Forecast!{fc_col}18"
        for col in ["B", "C", "D"]:
            ws[f"{col}{i}"].number_format = '#,##0.00'

    valuation_rows = [
        (4, "DCF", "=Summary!E4", "=Summary!F4", "=Summary!G4"),
        (5, "Relative", "=Summary!E5", "=Summary!F5", "=Summary!G5"),
        (6, "Weighted", "=Summary!E7", "=Summary!F7", "=Summary!G7"),
    ]
    for row, name, low, mid, high in valuation_rows:
        ws[f"F{row}"] = name
        ws[f"G{row}"] = low
        ws[f"H{row}"] = mid
        ws[f"I{row}"] = high
        for col in ["G", "H", "I"]:
            ws[f"{col}{row}"].number_format = "0.00"
            apply_formula_style(ws[f"{col}{row}"])

    line_chart = LineChart()
    line_chart.title = "Revenue / EBIT / FCFF Trend"
    line_chart.y_axis.title = "RMB"
    line_chart.x_axis.title = "Year"
    data_ref = Reference(ws, min_col=2, max_col=4, min_row=3, max_row=11)
    cats_ref = Reference(ws, min_col=1, min_row=4, max_row=11)
    line_chart.add_data(data_ref, titles_from_data=True)
    line_chart.set_categories(cats_ref)
    line_chart.height = 8
    line_chart.width = 16
    ws.add_chart(line_chart, "F9")

    bar_chart = BarChart()
    bar_chart.type = "bar"
    bar_chart.style = 10
    bar_chart.title = "Target Price Range"
    bar_chart.y_axis.title = "Method"
    bar_chart.x_axis.title = "Price"
    price_data = Reference(ws, min_col=7, max_col=9, min_row=3, max_row=6)
    price_cats = Reference(ws, min_col=6, min_row=4, max_row=6)
    bar_chart.add_data(price_data, titles_from_data=True)
    bar_chart.set_categories(price_cats)
    bar_chart.height = 7
    bar_chart.width = 16
    ws.add_chart(bar_chart, "F26")

    add_note(ws, "A10", "左侧数据用于图表引用，打开 Excel 后会随模型公式自动更新。")
    set_col_widths(ws, {"A": 10, "B": 14, "C": 14, "D": 14, "F": 16, "G": 12, "H": 12, "I": 12})
    ws.freeze_panes = "A4"


def create_rawdata_sheet(wb: Workbook, data: Dict[str, object]) -> None:
    ws = wb.create_sheet("Raw_Data")
    ws["A1"] = "模型关键原始口径"
    apply_title_style(ws["A1"])
    ws.merge_cells("A1:F1")
    headers = ["Year", "Revenue", "EBIT", "Capex", "Operating Cash Flow", "Operating NWC"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=idx, value=header)
        apply_header_style(cell)
    for row_idx, year in enumerate(data["years"], start=4):
        idx = data["years"].index(year)
        ws.cell(row=row_idx, column=1, value=year)
        ws.cell(row=row_idx, column=2, value=data["revenue"][idx])
        ws.cell(row=row_idx, column=3, value=data["ebit"][idx])
        ws.cell(row=row_idx, column=4, value=data["capex"][idx])
        ws.cell(row=row_idx, column=5, value=data["cfo"][idx])
        ws.cell(row=row_idx, column=6, value=data["nwc"][idx])
    ws.freeze_panes = "A4"
    set_col_widths(ws, {"A": 10, "B": 14, "C": 14, "D": 14, "E": 18, "F": 16})


def finalize_workbook(wb: Workbook) -> None:
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.column == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif cell.row == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")


def build_workbook(data: Dict[str, object]) -> Workbook:
    wb = Workbook()
    create_summary_sheet(wb, data)
    create_historical_sheet(wb, data)
    create_assumptions_sheet(wb, data)
    create_forecast_sheet(wb, data)
    create_dcf_sheet(wb, data)
    create_sensitivity_sheet(wb)
    create_comparable_sheet(wb, data)
    create_investment_thesis_sheet(wb, data)
    create_charts_sheet(wb, data)
    create_rawdata_sheet(wb, data)
    finalize_workbook(wb)
    return wb


def main() -> None:
    ensure_output_dir()
    data = build_historical_dataset()
    wb = build_workbook(data)
    apply_bilingual_fonts(wb)
    wb.save(OUTPUT_PATH)
    print(f"DCF估值模型已生成：{OUTPUT_PATH}")


if __name__ == "__main__":
    main()
