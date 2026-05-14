"""
Step 9/9 - offline HTML dashboard and interactive DCF valuation tool.

The generated HTML has two tabs:
1. Financial overview: historical financial metrics and charts.
2. Valuation model: editable assumptions, full DCF bridge, forecast table, and
   instantly recalculated intrinsic share price.
"""

from __future__ import annotations

import argparse
import html
import json
import math
import os
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, Optional

from openpyxl import load_workbook


OUTPUT_FILE_NAME = "financial_dcf_dashboard.html"
LOCAL_ECHARTS_SOURCE = Path(__file__).resolve().parent / "assets" / "echarts.min.js"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate an offline HTML financial and DCF dashboard.")
    parser.add_argument(
        "data_dir",
        nargs="?",
        default=None,
        help="Directory containing Info.csv and source files. If omitted, a folder picker will open.",
    )
    parser.add_argument("--data-dir", dest="data_dir_flag", default=None, help="Same as positional data_dir.")
    return parser.parse_args()


def configure_paths(data_dir_value: Optional[str]) -> Path:
    from pipeline_utils import prompt_data_dir_with_dialog, resolve_data_dir, set_results_dir

    data_dir = resolve_data_dir(data_dir_value) if data_dir_value else prompt_data_dir_with_dialog()
    data_dir = data_dir.expanduser().resolve()
    set_results_dir(data_dir / "results")
    return data_dir


def load_dataset(data_dir: Path) -> Dict[str, object]:
    from generate_dcf_valuation import build_historical_dataset
    from pipeline_utils import find_info_file

    info_path = find_info_file(data_dir)
    if info_path is None:
        raise FileNotFoundError(f"Info.csv not found in {data_dir}. Cannot generate DCF HTML dashboard.")
    return build_historical_dataset(data_dir=data_dir, info_path=info_path)


def as_float(value: object, fallback: float = 0.0) -> float:
    try:
        if value is None:
            return fallback
        out = float(value)
        if math.isnan(out) or math.isinf(out):
            return fallback
        return out
    except (TypeError, ValueError):
        return fallback


def read_assumptions(data: Dict[str, object], workbook_path: Path) -> Dict[str, object]:
    assumptions = {
        "shares_outstanding": float(data["shares_outstanding"]),
        "current_price": float(data["current_price"]),
        "net_cash": float(data["cash"][-1] - data["short_debt"][-1] - data["long_debt"][-1]),
        "wacc": 0.10,
        "terminal_growth": 0.03,
        "minority_interest": float(data["minority_interest"][-1]),
        "long_term_investments": float(data["long_term_investments"][-1]),
        "non_op_current_assets": float(data["non_op_current_assets"][-1]),
        "growths": [float(x) for x in data["default_growths"]],
        "ebit_margins": [float(data["base_ebit_margin"])] * 5,
        "tax_rates": [float(data["base_tax_rate"])] * 5,
        "da_ratios": [float(data["base_da_ratio"])] * 5,
        "capex_ratios": [float(data["base_capex_ratio"])] * 5,
        "nwc_ratios": [float(data["base_nwc_ratio"])] * 5,
    }
    if not workbook_path.exists():
        return assumptions

    wb = load_workbook(workbook_path, data_only=False, read_only=True)
    if "Assumptions" not in wb.sheetnames:
        return assumptions
    ws = wb["Assumptions"]
    for key, cell in {
        "shares_outstanding": "B3",
        "current_price": "B4",
        "net_cash": "B5",
        "wacc": "B6",
        "terminal_growth": "B7",
        "minority_interest": "B8",
        "long_term_investments": "B9",
        "non_op_current_assets": "B10",
    }.items():
        assumptions[key] = as_float(ws[cell].value, float(assumptions[key]))

    for key, row in {
        "growths": 14,
        "ebit_margins": 15,
        "tax_rates": 16,
        "da_ratios": 17,
        "capex_ratios": 18,
        "nwc_ratios": 19,
    }.items():
        assumptions[key] = [
            as_float(ws.cell(row=row, column=col).value, float(assumptions[key][col - 2]))
            for col in range(2, 7)
        ]
    return assumptions


def compute_dcf(data: Dict[str, object], assumptions: Dict[str, object]) -> Dict[str, object]:
    forecast_rows = []
    prev_revenue = float(data["revenue"][-1])
    prev_nwc = float(data["nwc"][-1])
    for idx, year in enumerate(data["forecast_years"]):
        revenue = prev_revenue * (1 + float(assumptions["growths"][idx]))
        ebit = revenue * float(assumptions["ebit_margins"][idx])
        tax_on_ebit = ebit * float(assumptions["tax_rates"][idx])
        nopat = ebit - tax_on_ebit
        da = revenue * float(assumptions["da_ratios"][idx])
        capex = revenue * float(assumptions["capex_ratios"][idx])
        operating_nwc = revenue * float(assumptions["nwc_ratios"][idx])
        change_nwc = operating_nwc - prev_nwc
        fcff = nopat + da - capex - change_nwc
        forecast_rows.append(
            {
                "year": int(year),
                "revenue": revenue,
                "growth": float(assumptions["growths"][idx]),
                "ebit_margin": float(assumptions["ebit_margins"][idx]),
                "tax_rate": float(assumptions["tax_rates"][idx]),
                "da_ratio": float(assumptions["da_ratios"][idx]),
                "capex_ratio": float(assumptions["capex_ratios"][idx]),
                "nwc_ratio": float(assumptions["nwc_ratios"][idx]),
                "ebit": ebit,
                "nopat": nopat,
                "da": da,
                "capex": capex,
                "operating_nwc": operating_nwc,
                "change_nwc": change_nwc,
                "fcff": fcff,
            }
        )
        prev_revenue = revenue
        prev_nwc = operating_nwc

    wacc = float(assumptions["wacc"])
    terminal_growth = float(assumptions["terminal_growth"])
    if wacc <= terminal_growth:
        terminal_growth = max(wacc - 0.005, 0.0)
    terminal_value = forecast_rows[-1]["fcff"] * (1 + terminal_growth) / (wacc - terminal_growth)
    pv_fcff = [row["fcff"] / ((1 + wacc) ** (idx + 1)) for idx, row in enumerate(forecast_rows)]
    pv_terminal = terminal_value / ((1 + wacc) ** len(forecast_rows))
    enterprise_value = sum(pv_fcff) + pv_terminal
    equity_value = (
        enterprise_value
        + float(assumptions["net_cash"])
        - float(assumptions["minority_interest"])
        + float(assumptions["long_term_investments"])
        + float(assumptions["non_op_current_assets"])
    )
    intrinsic_price = equity_value / float(assumptions["shares_outstanding"]) if assumptions["shares_outstanding"] else 0.0
    current_price = float(assumptions["current_price"])
    upside = intrinsic_price / current_price - 1 if current_price else 0.0
    return {
        "forecast_rows": forecast_rows,
        "pv_fcff": pv_fcff,
        "pv_terminal": pv_terminal,
        "terminal_value": terminal_value,
        "enterprise_value": enterprise_value,
        "equity_value": equity_value,
        "intrinsic_price": intrinsic_price,
        "upside": upside,
    }


def ensure_echarts_asset(assets_dir: Path) -> str:
    assets_dir.mkdir(parents=True, exist_ok=True)
    target = assets_dir / "echarts.min.js"
    if not LOCAL_ECHARTS_SOURCE.exists():
        raise FileNotFoundError(f"Offline ECharts asset not found: {LOCAL_ECHARTS_SOURCE}")
    shutil.copy2(LOCAL_ECHARTS_SOURCE, target)
    return "../_assets/echarts.min.js"


def money(value: float) -> str:
    abs_value = abs(value)
    if abs_value >= 1e8:
        return f"{value / 1e8:,.2f} 亿"
    if abs_value >= 1e4:
        return f"{value / 1e4:,.2f} 万"
    return f"{value:,.2f}"


def percent(value: float) -> str:
    return f"{value * 100:,.1f}%"


def esc(value: object) -> str:
    return html.escape(str(value), quote=True)


def table(headers: Iterable[str], rows: Iterable[Iterable[str]]) -> str:
    head = "".join(f"<th>{esc(h)}</th>" for h in headers)
    body = "".join("<tr>" + "".join(f"<td>{cell}</td>" for cell in row) + "</tr>" for row in rows)
    return f"<table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>"


def chart_container(chart_id: str, title: str) -> str:
    return f"""
    <div class="chart-card">
      <div class="chart-title">{esc(title)}</div>
      <div id="{esc(chart_id)}" class="echart"></div>
    </div>
    """


def build_assumption_inputs(years: list[int], assumptions: Dict[str, object]) -> str:
    rows = [
        ("growths", "Revenue Growth"),
        ("ebit_margins", "EBIT Margin"),
        ("tax_rates", "Tax Rate"),
        ("da_ratios", "D&A / Revenue"),
        ("capex_ratios", "Capex / Revenue"),
        ("nwc_ratios", "NWC / Revenue"),
    ]
    body = []
    for key, label in rows:
        cells = [f"<td>{esc(label)}</td>"]
        for idx, value in enumerate(assumptions[key]):
            cells.append(
                f'<td><input class="assumption-input" data-array="{key}" data-index="{idx}" '
                f'type="number" step="0.1" value="{float(value) * 100:.2f}"><span class="unit">%</span></td>'
            )
        body.append("<tr>" + "".join(cells) + "</tr>")
    header = "".join(f"<th>{year}E</th>" for year in years)
    return f"""
    <table class="input-table">
      <thead><tr><th>Assumption</th>{header}</tr></thead>
      <tbody>{''.join(body)}</tbody>
    </table>
    """


def build_html(data: Dict[str, object], assumptions: Dict[str, object], dcf: Dict[str, object], echarts_src: str) -> str:
    years = [int(y) for y in data["years"]]
    forecast_years = [int(y) for y in data["forecast_years"]]
    revenue = [float(x) for x in data["revenue"]]
    ebit = [float(x) for x in data["ebit"]]
    cfo = [float(x) for x in data["cfo"]]
    fcff = [float(x) for x in data["fcff_proxy"]]
    net_profit = [float(x) for x in data["parent_net_profit"]]
    ebit_margin = [float(x) for x in data["ebit_margin"]]

    hist_rows = [
        [esc(year), esc(money(revenue[i])), esc(money(ebit[i])), esc(percent(ebit_margin[i])), esc(money(net_profit[i])), esc(money(cfo[i])), esc(money(fcff[i]))]
        for i, year in enumerate(years)
    ]

    payload = {
        "company": {"ticker": data["ticker"], "name": data["company_name"], "valuationDate": data["valuation_date"]},
        "historical": {
            "years": years,
            "revenue": revenue,
            "ebit": ebit,
            "netProfit": net_profit,
            "cfo": cfo,
            "fcffProxy": fcff,
            "ebitMargin": ebit_margin,
            "baseRevenue": revenue[-1],
            "baseNwc": float(data["nwc"][-1]),
        },
        "forecastYears": forecast_years,
        "assumptions": assumptions,
        "initialDcf": dcf,
    }
    payload_json = json.dumps(payload, ensure_ascii=False)
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    initial_badge = "good" if dcf["upside"] >= 0.15 else "neutral" if dcf["upside"] >= -0.10 else "risk"

    cards = [
        ("当前股价", f"{assumptions['current_price']:.2f}", "来自 Info.csv / Assumptions"),
        ("DCF 每股内在价值", f"{dcf['intrinsic_price']:.2f}", f"较现价 {percent(dcf['upside'])}"),
        ("企业价值 EV", money(dcf["enterprise_value"]), "5年 FCFF + Terminal Value"),
        (f"{years[-1]}A 营收", money(revenue[-1]), f"EBIT Margin {percent(ebit_margin[-1])}"),
        (f"{years[-1]}A 经营现金流", money(cfo[-1]), f"FCFF Proxy {money(fcff[-1])}"),
        ("股东权益价值", money(dcf["equity_value"]), "EV 到 Equity Bridge"),
    ]
    card_html = "".join(
        f'<article class="metric-card"><span>{esc(label)}</span><strong>{esc(value)}</strong><small>{esc(note)}</small></article>'
        for label, value, note in cards
    )

    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{esc(data['company_name'])} 财务与DCF估值</title>
  <script src="{esc(echarts_src)}"></script>
  <style>
    :root {{
      --bg:#f6f7fb; --panel:#fff; --ink:#172033; --muted:#687385; --line:#dfe5ef;
      --accent:#2563eb; --teal:#0f766e; --red:#b91c1c; --shadow:0 18px 44px rgba(23,32,51,.08);
    }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; background:var(--bg); color:var(--ink); font-family:"Segoe UI","Microsoft YaHei",Arial,sans-serif; line-height:1.5; }}
    header {{ background:#0f172a; color:#fff; padding:30px 28px 24px; }}
    .wrap {{ max-width:1220px; margin:0 auto; }}
    .topline {{ display:flex; justify-content:space-between; gap:18px; align-items:flex-start; flex-wrap:wrap; }}
    h1 {{ margin:0; font-size:29px; letter-spacing:0; }}
    h2 {{ margin:0 0 14px; font-size:20px; }}
    h3 {{ margin:0 0 10px; font-size:16px; }}
    .subtitle {{ color:#b9c4d8; margin-top:8px; }}
    .badge {{ padding:7px 12px; border-radius:999px; background:rgba(255,255,255,.12); border:1px solid rgba(255,255,255,.18); font-weight:700; }}
    .badge.good {{ color:#bbf7d0; }} .badge.neutral {{ color:#fde68a; }} .badge.risk {{ color:#fecaca; }}
    main {{ padding:22px 28px 42px; }}
    .tabs {{ display:flex; gap:8px; margin:0 auto 16px; max-width:1220px; }}
    .tab-btn {{ border:1px solid var(--line); background:#fff; color:var(--ink); border-radius:8px; padding:10px 14px; font-weight:740; cursor:pointer; }}
    .tab-btn.active {{ background:#172033; color:#fff; border-color:#172033; }}
    .tab-pane {{ display:none; }} .tab-pane.active {{ display:block; }}
    .grid {{ display:grid; gap:16px; }}
    .metrics {{ grid-template-columns:repeat(3,minmax(0,1fr)); }}
    .two {{ grid-template-columns:1.05fr .95fr; }}
    .charts {{ grid-template-columns:repeat(2,minmax(0,1fr)); }}
    .three {{ grid-template-columns:repeat(3,minmax(0,1fr)); }}
    section, .metric-card, .chart-card {{ background:var(--panel); border:1px solid var(--line); border-radius:8px; box-shadow:var(--shadow); }}
    section {{ padding:20px; margin-top:16px; overflow:hidden; }}
    .metric-card {{ padding:17px; min-height:116px; }}
    .metric-card span, .metric-card small {{ color:var(--muted); display:block; }}
    .metric-card strong {{ display:block; font-size:25px; margin:7px 0 5px; }}
    .chart-card {{ padding:16px; box-shadow:none; }}
    .chart-title {{ font-weight:760; margin-bottom:8px; }}
    .echart {{ width:100%; height:320px; }}
    table {{ width:100%; border-collapse:collapse; font-size:14px; }}
    th,td {{ padding:9px 10px; border-bottom:1px solid var(--line); text-align:right; white-space:nowrap; }}
    th:first-child,td:first-child {{ text-align:left; }}
    th {{ color:#465366; background:#f3f6fb; font-weight:740; }}
    #valuation {{ display:grid; grid-template-columns:minmax(0,1fr) 260px; gap:12px; align-items:start; }}
    #valuation section {{ margin-top:16px; }}
    #valuation section:nth-of-type(1) {{ grid-column:1; grid-row:1; }}
    #valuation section:nth-of-type(2) {{ grid-column:2; grid-row:1; padding:14px; }}
    #valuation section:nth-of-type(n+3) {{ grid-column:1 / -1; }}
    #valuation section:nth-of-type(2) h2 {{ font-size:20px; margin-bottom:10px; }}
    #valuation section:nth-of-type(2) .three {{ grid-template-columns:1fr; gap:8px; }}
    #valuation section:nth-of-type(2) .three > div {{ background:#f7f9fc; border:1px solid var(--line); border-radius:8px; padding:9px 10px; }}
    #valuation section:nth-of-type(2) .note {{ font-size:12px; margin-top:0; }}
    .input-grid {{ display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:8px; }}
    label.input-card {{ display:block; background:#f7f9fc; border:1px solid var(--line); border-radius:8px; padding:8px 9px; }}
    label.input-card span {{ display:block; color:var(--muted); font-size:11px; margin-bottom:4px; line-height:1.25; }}
    input {{ width:100%; border:1px solid #cbd5e1; border-radius:7px; padding:6px 7px; font:inherit; font-size:13px; text-align:right; background:#fff; }}
    .input-table {{ font-size:12px; }}
    .input-table th,.input-table td {{ padding:6px 7px; }}
    .input-table input {{ width:64px; display:inline-block; }}
    .unit {{ color:var(--muted); margin-left:4px; }}
    .kpi-large {{ font-size:22px; font-weight:800; }}
    .note {{ color:var(--muted); font-size:13px; margin-top:10px; }}
    .toolbar {{ display:flex; gap:10px; flex-wrap:wrap; margin-top:12px; }}
    .action-btn {{ border:1px solid var(--line); background:#fff; border-radius:8px; padding:9px 12px; cursor:pointer; font-weight:700; }}
    .action-btn.primary {{ background:var(--accent); color:#fff; border-color:var(--accent); }}
    @media(max-width:1100px) {{ #valuation {{ display:block; }} }}
    @media(max-width:900px) {{ .metrics,.two,.charts,.three,.input-grid {{ grid-template-columns:1fr; }} section {{ overflow-x:auto; }} }}
  </style>
</head>
<body>
  <header>
    <div class="wrap topline">
      <div>
        <h1>{esc(data['company_name'])} ({esc(data['ticker'])}) 财务与 DCF 估值</h1>
        <div class="subtitle">估值日期：{esc(data['valuation_date'])} · 生成时间：{esc(generated_at)} · 离线 HTML</div>
      </div>
      <div id="headerBadge" class="badge {initial_badge}">目标价 <span id="headerPrice">{dcf['intrinsic_price']:.2f}</span> · <span id="headerUpside">{percent(dcf['upside'])}</span></div>
    </div>
  </header>
  <main>
    <div class="tabs">
      <button class="tab-btn active" data-tab="financial">财务基本情况</button>
      <button class="tab-btn" data-tab="valuation">估值过程与假设联动</button>
    </div>

    <div id="financial" class="tab-pane active wrap">
      <div class="grid metrics">{card_html}</div>
      <section>
        <h2>历史核心趋势</h2>
        <div class="grid charts">
          {chart_container("revenueTrendChart", "收入、EBIT、归母净利润")}
          {chart_container("cashFlowChart", "CFO 与 FCFF Proxy")}
        </div>
      </section>
      <section>
        <h2>历史关键财务数据</h2>
        {table(["年度", "Revenue", "EBIT", "EBIT Margin", "归母净利润", "CFO", "FCFF Proxy"], hist_rows)}
      </section>
    </div>

    <div id="valuation" class="tab-pane wrap">
      <section>
        <h2>可编辑 Assumptions</h2>
        <div class="input-grid">
          <label class="input-card"><span>当前股价</span><input id="currentPrice" type="number" step="0.01" value="{float(assumptions['current_price']):.4f}"></label>
          <label class="input-card"><span>总股本</span><input id="sharesOutstanding" type="number" step="1" value="{float(assumptions['shares_outstanding']):.4f}"></label>
          <label class="input-card"><span>WACC (%)</span><input id="wacc" type="number" step="0.1" value="{float(assumptions['wacc']) * 100:.2f}"></label>
          <label class="input-card"><span>永续增长率 (%)</span><input id="terminalGrowth" type="number" step="0.1" value="{float(assumptions['terminal_growth']) * 100:.2f}"></label>
          <label class="input-card"><span>净现金 / (净债务)</span><input id="netCash" type="number" step="10000" value="{float(assumptions['net_cash']):.4f}"></label>
          <label class="input-card"><span>少数股东权益</span><input id="minorityInterest" type="number" step="10000" value="{float(assumptions['minority_interest']):.4f}"></label>
          <label class="input-card"><span>长期金融及股权投资</span><input id="longTermInvestments" type="number" step="10000" value="{float(assumptions['long_term_investments']):.4f}"></label>
          <label class="input-card"><span>非经营流动资产</span><input id="nonOpCurrentAssets" type="number" step="10000" value="{float(assumptions['non_op_current_assets']):.4f}"></label>
        </div>
        <div style="margin-top:14px">{build_assumption_inputs(forecast_years, assumptions)}</div>
        <div class="toolbar">
          <button id="resetAssumptions" class="action-btn">恢复初始假设</button>
          <button id="copyAssumptions" class="action-btn primary">使用当前预测数据</button>
        </div>
        <div class="note">修改任意输入后，下面的预测表、估值桥、目标价和图表会自动联动重算。百分比请输入百分数，例如 10 表示 10%。</div>
      </section>

      <section>
        <h2>估值结论</h2>
        <div class="grid three">
          <div><span class="note">DCF 每股内在价值</span><div id="valuationPrice" class="kpi-large"></div></div>
          <div><span class="note">相对当前股价空间</span><div id="valuationUpside" class="kpi-large"></div></div>
          <div><span class="note">企业价值 EV</span><div id="valuationEv" class="kpi-large"></div></div>
        </div>
      </section>

      <section>
        <h2>估值全过程</h2>
        <div class="grid two">
          <div>{chart_container("forecastChart", "预测期 Revenue 与 FCFF")}</div>
          <div>
            <h3>DCF Bridge</h3>
            <table><tbody id="bridgeBody"></tbody></table>
          </div>
        </div>
      </section>

      <section>
        <h2>预测明细</h2>
        <table>
          <thead><tr><th>年度</th><th>Revenue</th><th>增长率</th><th>EBIT Margin</th><th>NOPAT</th><th>D&A</th><th>Capex</th><th>ΔNWC</th><th>FCFF</th><th>PV FCFF</th></tr></thead>
          <tbody id="forecastBody"></tbody>
        </table>
      </section>
    </div>
  </main>

  <script>
    const model = {payload_json};
    let initialAssumptions = JSON.parse(JSON.stringify(model.assumptions));
    const chartRefs = {{}};

    const num = (id) => Number(document.getElementById(id).value || 0);
    const pct = (id) => num(id) / 100;
    const money = (value) => {{
      const n = Number(value || 0), abs = Math.abs(n);
      if (abs >= 100000000) return (n / 100000000).toFixed(2) + ' 亿';
      if (abs >= 10000) return (n / 10000).toFixed(2) + ' 万';
      return n.toLocaleString('zh-CN', {{ maximumFractionDigits: 2 }});
    }};
    const price = (value) => Number(value || 0).toFixed(2);
    const percent = (value) => (Number(value || 0) * 100).toFixed(1) + '%';

    function collectAssumptions() {{
      const a = {{
        current_price: num('currentPrice'),
        shares_outstanding: num('sharesOutstanding'),
        wacc: pct('wacc'),
        terminal_growth: pct('terminalGrowth'),
        net_cash: num('netCash'),
        minority_interest: num('minorityInterest'),
        long_term_investments: num('longTermInvestments'),
        non_op_current_assets: num('nonOpCurrentAssets'),
        growths: [],
        ebit_margins: [],
        tax_rates: [],
        da_ratios: [],
        capex_ratios: [],
        nwc_ratios: []
      }};
      document.querySelectorAll('.assumption-input').forEach(input => {{
        a[input.dataset.array][Number(input.dataset.index)] = Number(input.value || 0) / 100;
      }});
      return a;
    }}

    function compute(a) {{
      const rows = [];
      let prevRevenue = model.historical.baseRevenue;
      let prevNwc = model.historical.baseNwc;
      model.forecastYears.forEach((year, i) => {{
        const revenue = prevRevenue * (1 + a.growths[i]);
        const ebit = revenue * a.ebit_margins[i];
        const taxOnEbit = ebit * a.tax_rates[i];
        const nopat = ebit - taxOnEbit;
        const da = revenue * a.da_ratios[i];
        const capex = revenue * a.capex_ratios[i];
        const operatingNwc = revenue * a.nwc_ratios[i];
        const changeNwc = operatingNwc - prevNwc;
        const fcff = nopat + da - capex - changeNwc;
        rows.push({{ year, revenue, growth:a.growths[i], ebitMargin:a.ebit_margins[i], nopat, da, capex, changeNwc, fcff }});
        prevRevenue = revenue;
        prevNwc = operatingNwc;
      }});
      let terminalGrowth = a.terminal_growth;
      if (a.wacc <= terminalGrowth) terminalGrowth = Math.max(a.wacc - 0.005, 0);
      const terminalValue = rows[rows.length - 1].fcff * (1 + terminalGrowth) / (a.wacc - terminalGrowth);
      const pvFcff = rows.map((row, i) => row.fcff / Math.pow(1 + a.wacc, i + 1));
      const pvTerminal = terminalValue / Math.pow(1 + a.wacc, rows.length);
      const enterpriseValue = pvFcff.reduce((x, y) => x + y, 0) + pvTerminal;
      const equityValue = enterpriseValue + a.net_cash - a.minority_interest + a.long_term_investments + a.non_op_current_assets;
      const intrinsicPrice = a.shares_outstanding ? equityValue / a.shares_outstanding : 0;
      const upside = a.current_price ? intrinsicPrice / a.current_price - 1 : 0;
      return {{ rows, pvFcff, pvTerminal, terminalValue, enterpriseValue, equityValue, intrinsicPrice, upside }};
    }}

    function initChart(id, option) {{
      const el = document.getElementById(id);
      if (!el || !window.echarts) return null;
      const chart = chartRefs[id] || echarts.init(el, null, {{ renderer: 'canvas' }});
      chartRefs[id] = chart;
      chart.setOption(option);
      return chart;
    }}

    function chartBase(xData, series) {{
      return {{
        color: ['#2563eb', '#10b981', '#f97316', '#7c3aed'],
        tooltip: {{ trigger: 'axis', valueFormatter: money }},
        legend: {{ top: 0, textStyle: {{ color: '#687385' }} }},
        grid: {{ left: 72, right: 26, top: 42, bottom: 42 }},
        xAxis: {{ type: 'category', data: xData, axisLabel: {{ color: '#687385' }} }},
        yAxis: {{ type: 'value', axisLabel: {{ color: '#687385', formatter: money }}, splitLine: {{ lineStyle: {{ color: '#e8edf5' }} }} }},
        series
      }};
    }}

    function forecastComboChart(xData, rows) {{
      return {{
        color: ['#2563eb', '#7c3aed'],
        tooltip: {{ trigger: 'axis', valueFormatter: money }},
        legend: {{ top: 0, textStyle: {{ color: '#687385' }} }},
        grid: {{ left: 72, right: 72, top: 42, bottom: 42 }},
        xAxis: {{ type: 'category', data: xData, axisLabel: {{ color: '#687385' }} }},
        yAxis: [
          {{ type: 'value', name: 'Revenue', axisLabel: {{ color: '#687385', formatter: money }}, splitLine: {{ lineStyle: {{ color: '#e8edf5' }} }} }},
          {{ type: 'value', name: 'FCFF', axisLabel: {{ color: '#687385', formatter: money }}, splitLine: {{ show: false }} }}
        ],
        series: [
          {{ name:'Revenue', type:'line', yAxisIndex:0, data:rows.map(r=>r.revenue), smooth:true, symbolSize:7 }},
          {{ name:'FCFF', type:'bar', yAxisIndex:1, data:rows.map(r=>r.fcff), barMaxWidth:38 }}
        ]
      }};
    }}

    function renderFinancialCharts() {{
      initChart('revenueTrendChart', chartBase(model.historical.years, [
        {{ name:'Revenue', type:'line', data:model.historical.revenue, smooth:true, symbolSize:7 }},
        {{ name:'EBIT', type:'line', data:model.historical.ebit, smooth:true, symbolSize:7 }},
        {{ name:'Parent Net Profit', type:'line', data:model.historical.netProfit, smooth:true, symbolSize:7 }}
      ]));
      initChart('cashFlowChart', chartBase(model.historical.years, [
        {{ name:'CFO', type:'bar', data:model.historical.cfo, barMaxWidth:38 }},
        {{ name:'FCFF Proxy', type:'line', data:model.historical.fcffProxy, smooth:true, symbolSize:7 }}
      ]));
    }}

    function renderValuation() {{
      const assumptions = collectAssumptions();
      const dcf = compute(assumptions);
      document.getElementById('valuationPrice').textContent = price(dcf.intrinsicPrice);
      document.getElementById('valuationUpside').textContent = percent(dcf.upside);
      document.getElementById('valuationEv').textContent = money(dcf.enterpriseValue);
      document.getElementById('headerPrice').textContent = price(dcf.intrinsicPrice);
      document.getElementById('headerUpside').textContent = percent(dcf.upside);
      const badge = document.getElementById('headerBadge');
      badge.className = 'badge ' + (dcf.upside >= 0.15 ? 'good' : dcf.upside >= -0.10 ? 'neutral' : 'risk');

      const bridge = [
        ['5年 FCFF 现值合计', money(dcf.pvFcff.reduce((x,y)=>x+y,0))],
        ['终值现值', money(dcf.pvTerminal)],
        ['企业价值 EV', money(dcf.enterpriseValue)],
        ['加：净现金 / 减：净债务', money(assumptions.net_cash)],
        ['减：少数股东权益', money(-assumptions.minority_interest)],
        ['加：长期金融及股权投资', money(assumptions.long_term_investments)],
        ['加：非经营流动资产', money(assumptions.non_op_current_assets)],
        ['股东权益价值', money(dcf.equityValue)],
        ['每股内在价值', price(dcf.intrinsicPrice)]
      ];
      document.getElementById('bridgeBody').innerHTML = bridge.map(r => `<tr><td>${{r[0]}}</td><td>${{r[1]}}</td></tr>`).join('');

      document.getElementById('forecastBody').innerHTML = dcf.rows.map((row, i) => `
        <tr>
          <td>${{row.year}}</td><td>${{money(row.revenue)}}</td><td>${{percent(row.growth)}}</td>
          <td>${{percent(row.ebitMargin)}}</td><td>${{money(row.nopat)}}</td><td>${{money(row.da)}}</td>
          <td>${{money(row.capex)}}</td><td>${{money(row.changeNwc)}}</td><td>${{money(row.fcff)}}</td><td>${{money(dcf.pvFcff[i])}}</td>
        </tr>`).join('');

      initChart('forecastChart', forecastComboChart(model.forecastYears, dcf.rows));
    }}

    document.querySelectorAll('.tab-btn').forEach(btn => {{
      btn.addEventListener('click', () => {{
        document.querySelectorAll('.tab-btn').forEach(x => x.classList.remove('active'));
        document.querySelectorAll('.tab-pane').forEach(x => x.classList.remove('active'));
        btn.classList.add('active');
        document.getElementById(btn.dataset.tab).classList.add('active');
        setTimeout(() => Object.values(chartRefs).forEach(c => c && c.resize()), 0);
      }});
    }});
    document.querySelectorAll('input').forEach(input => input.addEventListener('input', renderValuation));
    document.getElementById('resetAssumptions').addEventListener('click', () => {{
      document.getElementById('currentPrice').value = initialAssumptions.current_price;
      document.getElementById('sharesOutstanding').value = initialAssumptions.shares_outstanding;
      document.getElementById('wacc').value = (initialAssumptions.wacc * 100).toFixed(2);
      document.getElementById('terminalGrowth').value = (initialAssumptions.terminal_growth * 100).toFixed(2);
      document.getElementById('netCash').value = initialAssumptions.net_cash;
      document.getElementById('minorityInterest').value = initialAssumptions.minority_interest;
      document.getElementById('longTermInvestments').value = initialAssumptions.long_term_investments;
      document.getElementById('nonOpCurrentAssets').value = initialAssumptions.non_op_current_assets;
      document.querySelectorAll('.assumption-input').forEach(input => {{
        input.value = (initialAssumptions[input.dataset.array][Number(input.dataset.index)] * 100).toFixed(2);
      }});
      renderValuation();
    }});
    document.getElementById('copyAssumptions').addEventListener('click', () => {{
      initialAssumptions = JSON.parse(JSON.stringify(collectAssumptions()));
      renderValuation();
      alert('已使用当前预测数据作为本页面新的基准假设');
    }});
    window.addEventListener('resize', () => Object.values(chartRefs).forEach(c => c && c.resize()));
    renderFinancialCharts();
    renderValuation();
  </script>
</body>
</html>
"""


def main() -> None:
    args = parse_args()
    data_dir = configure_paths(args.data_dir_flag or args.data_dir)
    from pipeline_utils import ASSETS_DIR, VALUATION_DIR

    data = load_dataset(data_dir)
    workbook_path = VALUATION_DIR / "DCF_valuation_model.xlsx"
    assumptions = read_assumptions(data, workbook_path)
    dcf = compute_dcf(data, assumptions)
    echarts_src = ensure_echarts_asset(ASSETS_DIR)
    output_path = VALUATION_DIR / OUTPUT_FILE_NAME
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(build_html(data, assumptions, dcf, echarts_src), encoding="utf-8")
    print(f"HTML dashboard generated: {output_path}")


if __name__ == "__main__":
    main()
