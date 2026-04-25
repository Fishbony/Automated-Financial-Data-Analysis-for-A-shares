"""
Step 4/8 — 完整财务指标报告（增强版）
======================================
在 step3_extract_metrics.py 的基础上，新增 YoY 增速、CAGR、
ROE、资产负债率、CFO 质量等指标，并输出完整的 Markdown 分析报告。

相比 step3，本脚本新增
-----------------------
- YoY 增速：Revenue / Net Profit / CFO / FCF
- CAGR（全历史期间）
- ROE（归母净利润 / 平均归母权益）
- 资产负债率 / 权益比率
- CFO/Net profit（经营现金含金量）
- CapEX/Revenue（资本开支强度）
- 一致性校验（资产负债表恒等式）
- Summary Analysis 摘要表
- Model Ready 宽表（可直接粘贴进财务模型）
- 缺失科目日志（Missing Log）

输入
----
- results/csv/pl.csv    利润表
- results/csv/bs.csv    资产负债表
- results/csv/cf.csv    现金流量表

输出
----
- results/financial_core_metrics_plus.xlsx
    * ExtractData         原始科目提取值
    * Processed Metrics   全量建模指标
    * Consistency Checks  资产负债表恒等式校验
    * Summary Analysis    CAGR 与最新年度关键指标
    * Model Ready         可直接用于建模的宽表
    * Missing Log         未匹配到的科目日志
- results/financial_core_metrics_report.md
- results/missing_items_log.csv

运行方式
--------
    python step4_metrics_report.py
    # 或通过主管道：
    python run_pipeline.py
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Font as XLFont, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from excel_utils import apply_bilingual_fonts

PL_FILE = "./results/csv/pl.csv"
BS_FILE = "./results/csv/bs.csv"
CF_FILE = "./results/csv/cf.csv"

OUTPUT_XLSX = "./results/financial_core_metrics_plus.xlsx"
OUTPUT_MD = "./results/financial_core_metrics_report.md"
OUTPUT_MISSING = "./results/missing_items_log.csv"

def load_statement(file_path: str) -> pd.DataFrame:
    df = pd.read_csv(file_path)
    df = df.copy()
    first_col = df.columns[0]
    df.rename(columns={first_col: "Item"}, inplace=True)
    df["Item"] = df["Item"].astype(str).str.replace("\ufeff", "", regex=False).str.strip()
    df.columns = ["Item"] + [str(c).strip() for c in df.columns[1:]]
    for col in df.columns[1:]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df.set_index("Item")

def common_years(*dfs):
    cols = set(dfs[0].columns)
    for df in dfs[1:]:
        cols &= set(df.columns)
    return sorted(cols, key=lambda x: str(x))

def find_item(df: pd.DataFrame, candidates):
    for cand in candidates:
        if cand in df.index:
            return cand
    return None

def get_series(df: pd.DataFrame, candidates, years, label=None, missing_log=None, default=np.nan):
    item = find_item(df, candidates)
    if item is not None:
        s = df.loc[item, years].astype(float)
        if missing_log is not None:
            missing_log.append({"label": label or str(candidates[0]), "status": "found", "matched_item": item})
        return s
    if missing_log is not None:
        missing_log.append({"label": label or str(candidates[0]), "status": "missing", "matched_item": ""})
    return pd.Series([default] * len(years), index=years, dtype="float64")

def safe_div(a, b):
    b = b.replace(0, np.nan)
    return a / b

def yoy(series: pd.Series) -> pd.Series:
    return series.astype(float).pct_change()

def cagr(series: pd.Series) -> float:
    s = series.dropna().astype(float)
    if len(s) < 2:
        return np.nan
    start = s.iloc[0]
    end = s.iloc[-1]
    periods = len(s) - 1
    if start <= 0 or end < 0 or periods <= 0:
        return np.nan
    return (end / start) ** (1 / periods) - 1

def fmt_money(x):
    if pd.isna(x):
        return "NaN"
    abs_x = abs(x)
    if abs_x >= 1e8:
        return f"{x/1e8:,.2f} 亿元"
    if abs_x >= 1e4:
        return f"{x/1e4:,.2f} 万元"
    return f"{x:,.2f} 元"

def fmt_pct(x):
    if pd.isna(x):
        return "NaN"
    return f"{x:.2%}"

def _add_charts_sheet(output_path: str, metrics: pd.DataFrame, years: list) -> None:
    """
    在已保存的 Excel 文件中追加 "Charts" 工作表，包含 4 张图表：
    1. Revenue & Net profit — 柱状图（金额绝对值）
    2. Margin trends — 折线图（毛利率、净利率、EBITDA Margin）
    3. CFO & FCF — 柱状图（现金流绝对值）
    4. ROE & 资产负债率 — 折线图（资本结构与回报率）

    Parameters
    ----------
    output_path : str
        已生成的 Excel 文件路径，图表页将追加至该文件。
    metrics : pd.DataFrame
        Processed Metrics 数据框（行为指标名，列为年份）。
    years : list
        覆盖年份列表（字符串格式，与 metrics.columns 一致）。
    """
    wb = load_workbook(output_path)

    # ── 构建图表数据辅助页（隐藏）────────────────────────────────────────
    DATA_SHEET = "_ChartData"
    if DATA_SHEET in wb.sheetnames:
        del wb[DATA_SHEET]
    dws = wb.create_sheet(DATA_SHEET)
    dws.sheet_state = "hidden"

    # 列布局：A=Year, B=Revenue, C=Net profit, D=Gross Margin,
    #         E=Net Margin, F=EBITDA Margin, G=CFO, H=FCF,
    #         I=ROE, J=资产负债率
    headers = ["Year", "Revenue", "Net profit", "Gross Margin",
               "Net Margin", "EBITDA Margin", "CFO", "FCF",
               "ROE", "资产负债率"]
    metric_keys = ["Revenue", "Net profit", "Gross Margin",
                   "Net Margin", "EBITDA Margin", "CFO", "FCF",
                   "ROE", "资产负债率"]

    for col_idx, h in enumerate(headers, start=1):
        dws.cell(row=1, column=col_idx, value=h)

    n_years = len(years)
    for r, yr in enumerate(years, start=2):
        dws.cell(row=r, column=1, value=yr)
        for c_offset, key in enumerate(metric_keys, start=2):
            val = metrics.loc[key, yr] if key in metrics.index else None
            dws.cell(row=r, column=c_offset, value=val if pd.notna(val) else None)

    data_rows = n_years + 1  # including header row

    # ── Charts 展示页 ────────────────────────────────────────────────────
    CHARTS_SHEET = "Charts"
    if CHARTS_SHEET in wb.sheetnames:
        del wb[CHARTS_SHEET]
    cws = wb.create_sheet(CHARTS_SHEET)

    # 标题行
    title_cell = cws["A1"]
    title_cell.value = "核心财务指标图表"
    title_cell.font = XLFont(name="SimHei", bold=True, size=14)
    title_cell.fill = PatternFill("solid", fgColor="1F4E79")
    title_cell.font = XLFont(name="SimHei", bold=True, size=14, color="FFFFFF")

    def _year_ref():
        """返回年份列（A列）的 Reference，用于各图表的 x 轴标签。"""
        return Reference(dws, min_col=1, min_row=2, max_row=data_rows)

    # ── 图表 1：Revenue & Net profit 柱状图 ─────────────────────────────
    bar1 = BarChart()
    bar1.type = "col"
    bar1.grouping = "clustered"
    bar1.title = "Revenue & Net Profit（元）"
    bar1.y_axis.title = "金额（元）"
    bar1.x_axis.title = "年份"
    bar1.style = 10
    bar1.width = 18
    bar1.height = 12

    rev_data = Reference(dws, min_col=2, min_row=1, max_row=data_rows)
    np_data  = Reference(dws, min_col=3, min_row=1, max_row=data_rows)
    bar1.add_data(rev_data, titles_from_data=True)
    bar1.add_data(np_data, titles_from_data=True)
    bar1.set_categories(_year_ref())
    cws.add_chart(bar1, "A3")

    # ── 图表 2：Margin trends 折线图 ─────────────────────────────────────
    line1 = LineChart()
    line1.title = "利润率趋势（Margin）"
    line1.y_axis.title = "比率"
    line1.x_axis.title = "年份"
    line1.y_axis.numFmt = "0%"
    line1.style = 10
    line1.width = 18
    line1.height = 12

    gm_data   = Reference(dws, min_col=4, min_row=1, max_row=data_rows)
    nm_data   = Reference(dws, min_col=5, min_row=1, max_row=data_rows)
    ebit_data = Reference(dws, min_col=6, min_row=1, max_row=data_rows)
    line1.add_data(gm_data,   titles_from_data=True)
    line1.add_data(nm_data,   titles_from_data=True)
    line1.add_data(ebit_data, titles_from_data=True)
    line1.set_categories(_year_ref())
    cws.add_chart(line1, "J3")

    # ── 图表 3：CFO & FCF 柱状图 ─────────────────────────────────────────
    bar2 = BarChart()
    bar2.type = "col"
    bar2.grouping = "clustered"
    bar2.title = "CFO & FCF（元）"
    bar2.y_axis.title = "金额（元）"
    bar2.x_axis.title = "年份"
    bar2.style = 10
    bar2.width = 18
    bar2.height = 12

    cfo_data = Reference(dws, min_col=7, min_row=1, max_row=data_rows)
    fcf_data = Reference(dws, min_col=8, min_row=1, max_row=data_rows)
    bar2.add_data(cfo_data, titles_from_data=True)
    bar2.add_data(fcf_data, titles_from_data=True)
    bar2.set_categories(_year_ref())
    cws.add_chart(bar2, "A23")

    # ── 图表 4：ROE & 资产负债率 折线图 ─────────────────────────────────
    line2 = LineChart()
    line2.title = "ROE & 资产负债率"
    line2.y_axis.title = "比率"
    line2.x_axis.title = "年份"
    line2.y_axis.numFmt = "0%"
    line2.style = 10
    line2.width = 18
    line2.height = 12

    roe_data  = Reference(dws, min_col=9,  min_row=1, max_row=data_rows)
    lev_data  = Reference(dws, min_col=10, min_row=1, max_row=data_rows)
    line2.add_data(roe_data,  titles_from_data=True)
    line2.add_data(lev_data,  titles_from_data=True)
    line2.set_categories(_year_ref())
    cws.add_chart(line2, "J23")

    # 把 Charts 移到所有 Sheet 最前面（紧接在 Missing Log 后）
    wb.move_sheet(CHARTS_SHEET, offset=-(len(wb.sheetnames) - 1))

    wb.save(output_path)


def _apply_fonts_to_file(output_path: str) -> None:
    """重新打开文件，应用双语字体后保存。"""
    wb = load_workbook(output_path)
    apply_bilingual_fonts(wb)
    wb.save(output_path)


def main():
    pl = load_statement(PL_FILE)
    bs = load_statement(BS_FILE)
    cf = load_statement(CF_FILE)
    years = common_years(pl, bs, cf)

    missing_log = []

    ITEM_MAP = {
        "一、营业总收入(元)": {"df": pl, "candidates": ["一、营业总收入(元)", "*一、营业总收入(元)", "营业总收入(元)"]},
        "加：营业外收入(元)": {"df": pl, "candidates": ["加：营业外收入(元)", "营业外收入(元)"]},
        "其中：营业收入(元)": {"df": pl, "candidates": ["其中：营业收入(元)", "营业收入(元)"]},
        "其中：营业成本(元)": {"df": pl, "candidates": ["其中：营业成本(元)", "营业成本(元)"]},
        "四、利润总额(元)": {"df": pl, "candidates": ["四、利润总额(元)", "*四、利润总额(元)", "利润总额(元)"]},
        "营业税金及附加(元)": {"df": pl, "candidates": ["营业税金及附加(元)"]},
        "销售费用(元)": {"df": pl, "candidates": ["销售费用(元)"]},
        "管理费用(元)": {"df": pl, "candidates": ["管理费用(元)"]},
        "研发费用(元)": {"df": pl, "candidates": ["研发费用(元)"]},
        "资产减值损失(元)": {"df": pl, "candidates": ["资产减值损失(元)"]},
        "信用减值损失(元)": {"df": pl, "candidates": ["信用减值损失(元)"]},
        "加：公允价值变动收益(元)": {"df": pl, "candidates": ["加：公允价值变动收益(元)", "公允价值变动收益(元)"]},
        "投资收益(元)": {"df": pl, "candidates": ["投资收益(元)", "加：投资收益(元)"]},
        "资产处置收益(元)": {"df": pl, "candidates": ["资产处置收益(元)", "加：资产处置收益(元)"]},
        "其他收益(元)": {"df": pl, "candidates": ["其他收益(元)", "加：其他收益(元)"]},
        "三、营业利润(元)": {"df": pl, "candidates": ["三、营业利润(元)", "*三、营业利润(元)", "营业利润(元)"]},
        "其中：利息费用(元)": {"df": pl, "candidates": ["其中：利息费用(元)", "利息费用(元)"]},
        "利息收入(元)": {"df": pl, "candidates": ["利息收入(元)", "其中：利息收入(元)"]},
        "归属于母公司所有者的净利润(元)": {"df": pl, "candidates": ["归属于母公司所有者的净利润(元)", "*归属于母公司所有者的净利润(元)"]},
        "扣除非经常性损益后的净利润(元)": {"df": pl, "candidates": ["扣除非经常性损益后的净利润(元)"]},
        "减：所得税费用(元)": {"df": pl, "candidates": ["减：所得税费用(元)", "所得税费用(元)"]},
        "五、净利润(元)": {"df": pl, "candidates": ["五、净利润(元)", "*五、净利润(元)", "净利润(元)"]},
        "固定资产折旧、油气资产折耗、生产性生物资产折旧(元)": {"df": cf, "candidates": ["固定资产折旧、油气资产折耗、生产性生物资产折旧(元)", "固定资产折旧、油气资产折耗、生产性生物资产折旧"]},
        "无形资产摊销(元)": {"df": cf, "candidates": ["无形资产摊销(元)", "无形资产摊销"]},
        "长期待摊费用摊销(元)": {"df": cf, "candidates": ["长期待摊费用摊销(元)", "长期待摊费用摊销"]},
        "购建固定资产、无形资产和其他长期资产支付的现金(元)": {"df": cf, "candidates": ["购建固定资产、无形资产和其他长期资产支付的现金(元)"]},
        "处置固定资产、无形资产和其他长期资产收回的现金净额(元)": {"df": cf, "candidates": ["处置固定资产、无形资产和其他长期资产收回的现金净额(元)"]},
        "*经营活动产生的现金流量净额(元)": {"df": cf, "candidates": ["*经营活动产生的现金流量净额(元)", "经营活动产生的现金流量净额(元)"]},
        "*投资活动产生的现金流量净额(元)": {"df": cf, "candidates": ["*投资活动产生的现金流量净额(元)", "投资活动产生的现金流量净额(元)"]},
        "*筹资活动产生的现金流量净额(元)": {"df": cf, "candidates": ["*筹资活动产生的现金流量净额(元)", "筹资活动产生的现金流量净额(元)"]},
        "归属于母公司所有者权益合计(元)": {"df": bs, "candidates": ["归属于母公司所有者权益合计(元)", "*归属于母公司所有者权益合计(元)"]},
        "资产总计(元)": {"df": bs, "candidates": ["资产总计(元)", "*资产总计(元)", "*资产合计(元)", "资产合计(元)"]},
        "负债合计(元)": {"df": bs, "candidates": ["负债合计(元)", "*负债合计(元)"]},
        "所有者权益合计(元)": {"df": bs, "candidates": ["所有者权益合计(元)", "*所有者权益合计(元)", "*所有者权益（或股东权益）合计(元)", "所有者权益（或股东权益）合计(元)"]},
        "货币资金(元)": {"df": bs, "candidates": ["货币资金(元)", "*货币资金(元)", "货币资金"]},
    }

    extract_df = pd.DataFrame(index=list(ITEM_MAP.keys()), columns=years, dtype="float64")
    for item_name, meta in ITEM_MAP.items():
        extract_df.loc[item_name] = get_series(meta["df"], meta["candidates"], years, item_name, missing_log)

    营业总收入 = extract_df.loc["一、营业总收入(元)"]
    营业外收入 = extract_df.loc["加：营业外收入(元)"]
    营业收入 = extract_df.loc["其中：营业收入(元)"]
    营业成本 = extract_df.loc["其中：营业成本(元)"]
    利润总额 = extract_df.loc["四、利润总额(元)"]
    营业税金及附加 = extract_df.loc["营业税金及附加(元)"]
    销售费用 = extract_df.loc["销售费用(元)"]
    管理费用 = extract_df.loc["管理费用(元)"]
    研发费用 = extract_df.loc["研发费用(元)"]
    资产减值损失 = extract_df.loc["资产减值损失(元)"]
    信用减值损失 = extract_df.loc["信用减值损失(元)"]
    公允价值变动收益 = extract_df.loc["加：公允价值变动收益(元)"]
    投资收益 = extract_df.loc["投资收益(元)"]
    资产处置收益 = extract_df.loc["资产处置收益(元)"]
    其他收益 = extract_df.loc["其他收益(元)"]
    营业利润 = extract_df.loc["三、营业利润(元)"]
    利息费用 = extract_df.loc["其中：利息费用(元)"]
    利息收入 = extract_df.loc["利息收入(元)"]
    折旧 = extract_df.loc["固定资产折旧、油气资产折耗、生产性生物资产折旧(元)"]
    无形摊销 = extract_df.loc["无形资产摊销(元)"]
    长期待摊摊销 = extract_df.loc["长期待摊费用摊销(元)"]
    归母净利润 = extract_df.loc["归属于母公司所有者的净利润(元)"]
    扣非净利润 = extract_df.loc["扣除非经常性损益后的净利润(元)"]
    所得税 = extract_df.loc["减：所得税费用(元)"]
    净利润 = extract_df.loc["五、净利润(元)"]
    CapEX = extract_df.loc["购建固定资产、无形资产和其他长期资产支付的现金(元)"]
    处置长期资产收回现金 = extract_df.loc["处置固定资产、无形资产和其他长期资产收回的现金净额(元)"]
    CFO = extract_df.loc["*经营活动产生的现金流量净额(元)"]
    CFI = extract_df.loc["*投资活动产生的现金流量净额(元)"]
    CFF = extract_df.loc["*筹资活动产生的现金流量净额(元)"]
    归母权益 = extract_df.loc["归属于母公司所有者权益合计(元)"]
    资产总计 = extract_df.loc["资产总计(元)"]
    负债合计 = extract_df.loc["负债合计(元)"]
    所有者权益合计 = extract_df.loc["所有者权益合计(元)"]

    metrics_index = ["营业外收入/营业总收入","营业外收入/利润总额","Revenue","COGS","Gross Profit","Opex","Other operating income","EBIT_建模法","EBIT_报表校验法","D&A","EBITDA","Adjusted EBITDA","Pre-tax Profit","Tax","Net profit","CapEX","Net CapEx","CFO","CFI","CFF","FCF","Gross Margin","Net Margin","EBITDA Margin","Revenue YoY","Net Profit YoY","CFO YoY","FCF YoY","ROE","资产负债率","权益比率","CFO/Net profit","CapEX/Revenue"]
    metrics = pd.DataFrame(index=metrics_index, columns=years, dtype="float64")

    metrics.loc["营业外收入/营业总收入"] = safe_div(营业外收入, 营业总收入)
    metrics.loc["营业外收入/利润总额"] = safe_div(营业外收入, 利润总额)
    metrics.loc["Revenue"] = 营业收入
    metrics.loc["COGS"] = 营业成本
    metrics.loc["Gross Profit"] = 营业收入 - 营业成本
    metrics.loc["Opex"] = 营业税金及附加.fillna(0) + 销售费用.fillna(0) + 管理费用.fillna(0) + 研发费用.fillna(0)
    metrics.loc["Other operating income"] = 其他收益.fillna(0) + 投资收益.fillna(0) + 公允价值变动收益.fillna(0) + 资产处置收益.fillna(0) - 资产减值损失.fillna(0) - 信用减值损失.fillna(0)
    metrics.loc["EBIT_建模法"] = metrics.loc["Gross Profit"].fillna(0) - metrics.loc["Opex"].fillna(0) + metrics.loc["Other operating income"].fillna(0)
    metrics.loc["EBIT_报表校验法"] = 营业利润.fillna(0) + 利息费用.fillna(0) - 利息收入.fillna(0)
    metrics.loc["D&A"] = 折旧.fillna(0) + 无形摊销.fillna(0) + 长期待摊摊销.fillna(0)
    metrics.loc["EBITDA"] = metrics.loc["EBIT_建模法"].fillna(0) + metrics.loc["D&A"].fillna(0)
    metrics.loc["Adjusted EBITDA"] = 扣非净利润.fillna(0) + 所得税.fillna(0) + (利息费用.fillna(0) - 利息收入.fillna(0)) + metrics.loc["D&A"].fillna(0)
    metrics.loc["Pre-tax Profit"] = 利润总额
    metrics.loc["Tax"] = 所得税
    metrics.loc["Net profit"] = 净利润
    metrics.loc["CapEX"] = CapEX
    metrics.loc["Net CapEx"] = CapEX.fillna(0) - 处置长期资产收回现金.fillna(0)
    metrics.loc["CFO"] = CFO
    metrics.loc["CFI"] = CFI
    metrics.loc["CFF"] = CFF
    metrics.loc["FCF"] = CFO.fillna(0) - CapEX.fillna(0)
    metrics.loc["Gross Margin"] = safe_div(metrics.loc["Gross Profit"], metrics.loc["Revenue"])
    metrics.loc["Net Margin"] = safe_div(metrics.loc["Net profit"], metrics.loc["Revenue"])
    metrics.loc["EBITDA Margin"] = safe_div(metrics.loc["EBITDA"], metrics.loc["Revenue"])
    metrics.loc["Revenue YoY"] = yoy(metrics.loc["Revenue"])
    metrics.loc["Net Profit YoY"] = yoy(metrics.loc["Net profit"])
    metrics.loc["CFO YoY"] = yoy(metrics.loc["CFO"])
    metrics.loc["FCF YoY"] = yoy(metrics.loc["FCF"])
    avg_equity = (归母权益.shift(1) + 归母权益) / 2
    metrics.loc["ROE"] = safe_div(归母净利润, avg_equity)
    metrics.loc["资产负债率"] = safe_div(负债合计, 资产总计)
    metrics.loc["权益比率"] = safe_div(所有者权益合计, 资产总计)
    metrics.loc["CFO/Net profit"] = safe_div(CFO, 净利润)
    metrics.loc["CapEX/Revenue"] = safe_div(CapEX, 营业收入)

    checks = pd.DataFrame(index=years)
    checks["资产总计"] = 资产总计
    checks["负债合计+所有者权益合计"] = 负债合计 + 所有者权益合计
    checks["资产负债表差额"] = checks["资产总计"] - checks["负债合计+所有者权益合计"]
    checks["资产负债表通过"] = np.isclose(checks["资产负债表差额"], 0, atol=1e3)

    analysis = pd.DataFrame(index=["Revenue CAGR","Net profit CAGR","CFO CAGR","FCF CAGR","最新年度 ROE","最新年度 Gross Margin","最新年度 Net Margin","最新年度 EBITDA Margin","最新年度 资产负债率","最新年度 CFO/Net profit"], columns=["Value"])
    analysis.loc["Revenue CAGR", "Value"] = cagr(metrics.loc["Revenue"])
    analysis.loc["Net profit CAGR", "Value"] = cagr(metrics.loc["Net profit"])
    analysis.loc["CFO CAGR", "Value"] = cagr(metrics.loc["CFO"])
    analysis.loc["FCF CAGR", "Value"] = cagr(metrics.loc["FCF"])
    last_year = years[-1]
    analysis.loc["最新年度 ROE", "Value"] = metrics.loc["ROE", last_year]
    analysis.loc["最新年度 Gross Margin", "Value"] = metrics.loc["Gross Margin", last_year]
    analysis.loc["最新年度 Net Margin", "Value"] = metrics.loc["Net Margin", last_year]
    analysis.loc["最新年度 EBITDA Margin", "Value"] = metrics.loc["EBITDA Margin", last_year]
    analysis.loc["最新年度 资产负债率", "Value"] = metrics.loc["资产负债率", last_year]
    analysis.loc["最新年度 CFO/Net profit", "Value"] = metrics.loc["CFO/Net profit", last_year]

    model_ready = pd.DataFrame(index=years)
    for col in ["Revenue","COGS","Gross Profit","Opex","Other operating income","EBIT_建模法","D&A","EBITDA","Pre-tax Profit","Tax","Net profit","CapEX","Net CapEx","CFO","CFI","CFF","FCF","Gross Margin","Net Margin","EBITDA Margin","ROE","资产负债率"]:
        name = "EBIT" if col == "EBIT_建模法" else col
        model_ready[name] = metrics.loc[col]

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        extract_df.to_excel(writer, sheet_name="ExtractData")
        metrics.to_excel(writer, sheet_name="Processed Metrics")
        checks.to_excel(writer, sheet_name="Consistency Checks")
        analysis.to_excel(writer, sheet_name="Summary Analysis")
        model_ready.to_excel(writer, sheet_name="Model Ready")
        pd.DataFrame(missing_log).to_excel(writer, sheet_name="Missing Log", index=False)

    pd.DataFrame(missing_log).to_csv(OUTPUT_MISSING, index=False, encoding="utf-8-sig")

    # ── 后处理：添加 Charts 图表页 + 双语字体 ────────────────────────────
    _add_charts_sheet(OUTPUT_XLSX, metrics, years)
    _apply_fonts_to_file(OUTPUT_XLSX)

    found_count = sum(1 for x in missing_log if x["status"] == "found")
    missing_items = [x["label"] for x in missing_log if x["status"] == "missing"]

    lines = []
    lines += ["# 财务核心指标提取与计算报告", ""]
    lines += ["## 1. 文件与年份范围", ""]
    lines += [f"- 利润表文件：`{PL_FILE}`", f"- 资产负债表文件：`{BS_FILE}`", f"- 现金流量表文件：`{CF_FILE}`", f"- 覆盖年份：`{years[0]}` 至 `{years[-1]}`", ""]
    lines += ["## 2. 科目提取情况", ""]
    lines += [f"- 成功匹配科目数：**{found_count}**", f"- 未匹配科目数：**{len(missing_items)}**"]
    lines += [f"- 未匹配科目：{', '.join(missing_items)}" if missing_items else "- 未匹配科目：无", ""]
    lines += ["## 3. 一致性校验", "", "### 3.1 资产负债表恒等式", "", checks.to_markdown(), ""]
    lines += ["### 3.2 结论", ""]
    if bool(checks["资产负债表通过"].all()):
        lines += ["- 资产负债表恒等式整体通过。", ""]
    else:
        failed = [str(i) for i, ok in checks["资产负债表通过"].items() if not ok]
        lines += [f"- 未通过年份：{', '.join(failed)}。建议检查原始报表口径。", ""]
    lines += ["## 4. 核心指标摘要", "", analysis.to_markdown(), ""]
    latest = pd.DataFrame(index=["Revenue","Gross Margin","Net Margin","EBITDA Margin","ROE","资产负债率","CFO/Net profit","FCF"], columns=["Value"])
    latest.loc["Revenue","Value"] = fmt_money(metrics.loc["Revenue", last_year]) if pd.notna(metrics.loc["Revenue", last_year]) else "NaN"
    latest.loc["FCF","Value"] = fmt_money(metrics.loc["FCF", last_year]) if pd.notna(metrics.loc["FCF", last_year]) else "NaN"
    for k in ["Gross Margin","Net Margin","EBITDA Margin","ROE","资产负债率","CFO/Net profit"]:
        latest.loc[k,"Value"] = fmt_pct(metrics.loc[k, last_year]) if pd.notna(metrics.loc[k, last_year]) else "NaN"
    lines += ["## 5. 最新年度关键指标", "", latest.to_markdown(), ""]
    trend = model_ready[["Revenue","Net profit","CFO","FCF","Gross Margin","Net Margin","ROE","资产负债率"]]
    lines += ["## 6. 近年趋势表", "", trend.to_markdown(), "", "## 7. 指标口径说明", ""]
    lines += [
        "- Opex = 营业税金及附加 + 销售费用 + 管理费用 + 研发费用",
        "- Other operating income = 其他收益 + 投资收益 + 公允价值变动收益 + 资产处置收益 - 资产减值损失 - 信用减值损失",
        "- EBIT_建模法 = Gross Profit - Opex + Other operating income",
        "- EBIT_报表校验法 = 营业利润 + 利息费用 - 利息收入",
        "- D&A = 折旧 + 无形资产摊销 + 长期待摊费用摊销",
        "- EBITDA = EBIT_建模法 + D&A",
        "- Adjusted EBITDA = 扣非净利润 + 所得税 + 净利息支出 + D&A",
        "- Net CapEx = CapEX - 处置长期资产收回的现金净额",
        "- FCF = CFO - CapEX",
        "- ROE = 归母净利润 / 平均归母权益",
        "- 资产负债率 = 负债合计 / 资产总计",
        "",
        "## 8. 输出文件",
        "",
        f"- Excel：`{OUTPUT_XLSX}`",
        f"- Markdown：`{OUTPUT_MD}`",
        f"- 缺失科目日志：`{OUTPUT_MISSING}`",
        "",
    ]
    Path(OUTPUT_MD).write_text("\n".join(lines), encoding="utf-8")

    print(OUTPUT_XLSX, OUTPUT_MD, OUTPUT_MISSING)

if __name__ == "__main__":
    main()
