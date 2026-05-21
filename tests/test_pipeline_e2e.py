from __future__ import annotations

import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]
DEMO_DIR = PROJECT_DIR / "demo"
SAMPLE_ROOT = PROJECT_DIR / "tests" / "fixtures" / "company_samples"
REQUIRED_SUFFIXES = ("_debt_year.xls", "_benefit_year.xls", "_cash_year.xls")


def _has_required_statement_files(sample_dir: Path) -> bool:
    return all(any(path.name.endswith(suffix) for path in sample_dir.iterdir()) for suffix in REQUIRED_SUFFIXES)


def _sample_dirs() -> list[Path]:
    samples = [DEMO_DIR]
    if SAMPLE_ROOT.exists():
        samples.extend(path for path in sorted(SAMPLE_ROOT.iterdir()) if path.is_dir() and _has_required_statement_files(path))
    return samples


def _copy_sample_inputs(sample_dir: Path, target_dir: Path) -> None:
    for path in sample_dir.iterdir():
        if path.is_file() and (
            path.name.endswith(REQUIRED_SUFFIXES)
            or path.name.lower() == "info.csv"
            or path.name == "valuation_config.json"
        ):
            shutil.copy2(path, target_dir / path.name)


def _expected_artifacts(results_dir: Path, has_info: bool) -> list[Path]:
    artifacts = [
        results_dir / "01_csv" / "bs.csv",
        results_dir / "01_csv" / "pl.csv",
        results_dir / "01_csv" / "cf.csv",
        results_dir / "02_checks" / "statement_consistency_checks.xlsx",
        results_dir / "03_metrics" / "Core_Metrics.xlsx",
        results_dir
        / "04_rebuilt_statements"
        / "rebuilt_statement_checks"
        / "rebuilt_statement_checks.xlsx",
    ]
    if has_info:
        artifacts.extend(
            [
                results_dir / "05_valuation" / "DCF_valuation_model.xlsx",
                results_dir / "financial_dcf_dashboard.html",
            ]
        )
    return artifacts


def test_company_sample_pipeline_regression(tmp_path: Path) -> None:
    for sample_dir in _sample_dirs():
        data_dir = tmp_path / sample_dir.name
        data_dir.mkdir()
        _copy_sample_inputs(sample_dir, data_dir)

        completed = subprocess.run(
            [sys.executable, "-m", "afda.run_pipeline", str(data_dir)],
            cwd=PROJECT_DIR,
            text=True,
            capture_output=True,
            check=False,
        )

        assert completed.returncode == 0, completed.stdout + completed.stderr

        stdout = completed.stdout
        assert stdout.index("Automated Financial Data Analysis for A-shares") < stdout.index(
            "[RUN] afda.step1_convert_xls_to_csv"
        )
        assert stdout.index("[RUN] afda.step1_convert_xls_to_csv") < stdout.index(
            "Step 1: convert RoyalFlush XLS exports to CSV"
        )
        assert "Pipeline completed successfully." in stdout

        results_dir = data_dir / "results"
        has_info = (data_dir / "Info.csv").exists()
        for path in _expected_artifacts(results_dir, has_info):
            assert path.exists(), f"Missing expected pipeline artifact for {sample_dir.name}: {path}"
            assert path.stat().st_size > 0, f"Pipeline artifact is empty for {sample_dir.name}: {path}"

        if has_info:
            workbook = load_workbook(results_dir / "05_valuation" / "DCF_valuation_model.xlsx", read_only=True)
            assert "Assumption_Audit" in workbook.sheetnames
            html = (results_dir / "financial_dcf_dashboard.html").read_text(encoding="utf-8")
            assert "Assumption Audit" in html
            assert "Valuation Risk Warnings" in html
